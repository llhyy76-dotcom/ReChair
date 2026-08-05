from pathlib import Path
from openpyxl import load_workbook

from core.parser import PartRecord
from core.purchase_analyzer import (
    PurchaseAnalysisResult,
    PurchaseItem,
    _analyze_item,
    write_purchase_analysis,
)


def _row(order_price: float):
    item = PurchaseItem(
        manufacturer="AC", source_sheet="AC", source_row=2,
        part_no="P-100", part_name_kr="패드, 등, A383",
        part_name_en="", model="A383", order_unit_price=order_price,
        order_qty=10, order_amount=order_price * 10,
        usage_type="", note="",
    )
    records = [
        PartRecord(
            base_date="2020-05-19", ir_no="IR20276~300", model="A383",
            part_no="P-100", part_name_kr="패드, 등, A383",
            price=40.0, source_file="old.xlsx", source_row=2,
        ),
        PartRecord(
            base_date="2022-04-04", ir_no="IR22138~141", model="A383",
            part_no="P-100", part_name_kr="패드, 등, A383",
            price=12.85, source_file="new.xlsx", source_row=3,
        ),
    ]
    return _analyze_item(item, records)


def test_negotiation_report_is_created(tmp_path):
    row = _row(40.71)
    result = PurchaseAnalysisResult(source_path=Path("purchase.xlsx"), rows=[row])
    result.total_items = 1
    result.total_qty = 10
    result.total_order_amount = 407.1
    result.matched_items = 1
    result.increased_items = 1
    result.net_impact = 278.6

    output = write_purchase_analysis(result, tmp_path)
    assert output.exists()
    assert result.negotiation_output_path is not None
    assert result.negotiation_output_path.exists()

    wb = load_workbook(result.negotiation_output_path, data_only=False)
    assert "협상자료_요약" in wb.sheetnames
    assert "AC_협상자료" in wb.sheetnames
    assert wb["협상자료_요약"]["O16"].value == "매우비쌈"
    assert "재확인" in wb["협상자료_요약"]["S16"].value


def test_stable_price_is_not_negotiation_candidate(tmp_path):
    row = _row(12.85)
    result = PurchaseAnalysisResult(source_path=Path("purchase.xlsx"), rows=[row])
    result.total_items = 1
    result.total_qty = 10
    result.total_order_amount = 128.5
    result.matched_items = 1
    result.unchanged_items = 1

    write_purchase_analysis(result, tmp_path)
    wb = load_workbook(result.negotiation_output_path, data_only=False)
    assert wb["협상자료_요약"]["A5"].value == 0
    assert "AC_협상자료" not in wb.sheetnames
