from pathlib import Path

from openpyxl import Workbook

from core.parser import parse_money, parse_workbook


def build_legacy(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "단가확인"
    ws.merge_cells("B2:G2")
    ws["B2"] = "XC"
    headers = ["부품명(Eng)", "부품명(Kor)", "1년 사용량", "요청수량", "단가", "합계"]
    for col, value in enumerate(headers, start=2):
        ws.cell(3, col, value)

    rows = [
        ["PCB, Main, 950", "PCB, 메인, 950", 4, 10, "$21,00", "$210,00"],
        ["PU Cover, Leg, 7859", "PU커버, 종아리/다리, 7859", 8, 10, "$21,00", "$210,00"],
        ["Legrest, X3000", "다리부 유닛, X3000", 5, 20, None, "$0,00"],
        ["Backrest Frame, 385N", "등쿠션프레임, BK, 385N", 3, 10, "$11,00", "$110,00"],
        ["Remote controller, 1310", "리모컨, 1310", 5, 10, "$21,87", "$218,70"],
    ]
    for row_no, values in enumerate(rows, start=4):
        for col, value in enumerate(values, start=2):
            ws.cell(row_no, col, value)
    wb.save(path)


def test_decimal_comma_money():
    assert parse_money("$21,00") == 21.0
    assert parse_money("$218,70") == 218.7
    assert parse_money("$1,500.00") == 1500.0
    assert parse_money("$1.500,00") == 1500.0


def test_legacy_price_request_is_parsed(tmp_path):
    path = tmp_path / "XC 2012유상발주2 단가요청 201216.xlsx"
    build_legacy(path)

    records = parse_workbook(path)

    assert len(records) == 5
    assert records[0].ir_no.startswith("XC2012")
    assert records[0].model == "950"
    assert records[0].qty == 10
    assert records[0].price == 21
    assert records[0].amount == 210
    assert records[0].parameter == "1년 사용량: 4"

    assert records[2].model == "X3000"
    assert records[2].qty == 20
    assert records[2].price is None

    assert records[4].model == "1310"
    assert records[4].price == 21.87
    assert records[4].amount == 218.7


def test_part_number_first_column(tmp_path):
    path = tmp_path / "XC2012_유상발주.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["부품명(Eng)", "부품명(Kor)", "1년 사용량", "요청수량", "단가", "합계"])
    ws.append(["C101110-0400", "패드, 등, L50", 10, 50, "$30,00", "$1.500,00"])
    wb.save(path)

    records = parse_workbook(path)

    assert len(records) == 1
    assert records[0].part_no == "C101110-0400"
    assert records[0].part_name_en == ""
    assert records[0].model == "L50"
    assert records[0].price == 30
    assert records[0].amount == 1500
