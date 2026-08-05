from pathlib import Path
from openpyxl import load_workbook
from core.parser import PartRecord
from core.purchase_analyzer import (
    PurchaseAnalysisResult, PurchaseItem, _analyze_item, write_purchase_analysis,
)

def _item(price: float) -> PurchaseItem:
    return PurchaseItem(
        manufacturer="IR", source_sheet="IR", source_row=2,
        part_no="P-1", part_name_kr="테스트", part_name_en="Test Part",
        model="A100", order_unit_price=price, order_qty=10,
        order_amount=price * 10, usage_type="", note="",
    )

def _records():
    return [
        PartRecord(base_date="2024-01-01", ir_no="IR2401", model="A100",
                   part_no="P-1", part_name_en="Test Part", price=10.0),
        PartRecord(base_date="2025-01-01", ir_no="IR2501", model="A100",
                   part_no="P-1", part_name_en="Test Part", price=12.0),
    ]

def test_price_intelligence_flags_large_increase():
    row = _analyze_item(_item(18.0), _records())
    assert row.intelligence_grade == "매우비쌈"
    assert row.latest_deviation_rate == 50.0

def test_price_intelligence_accepts_stable_price():
    row = _analyze_item(_item(12.1), _records())
    assert row.intelligence_grade == "적정"

def test_intelligence_sheet_is_created(tmp_path):
    row = _analyze_item(_item(18.0), _records())
    result = PurchaseAnalysisResult(source_path=Path("sample.xlsx"), rows=[row])
    result.total_items = 1
    result.total_qty = 10
    result.total_order_amount = 180
    result.matched_items = 1
    result.increased_items = 1
    result.net_impact = 60
    output = write_purchase_analysis(result, tmp_path)
    workbook = load_workbook(output, data_only=False)
    assert "가격인텔리전스" in workbook.sheetnames
    assert workbook["가격인텔리전스"]["M2"].value == "매우비쌈"
    assert workbook["전체_상세"]["AH1"].value == "가격판정"
