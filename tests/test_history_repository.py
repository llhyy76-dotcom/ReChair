from openpyxl import Workbook

from core.history_repository import filter_history_rows, load_history_rows


def make_history(path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    price = workbook.create_sheet("가격이력")
    price.append([
        "가격이력이벤트ID", "제조사", "모델", "부품번호", "부품명(영문)",
        "부품명(한글)", "이전단가", "최근단가", "변동액", "변동률(%)",
        "이전요청번호", "최근요청번호", "이전원본파일", "최근원본파일",
        "이전기준일", "최근기준일", "통화",
    ])
    price.append([
        "ID1", "AC", "L500", "P-100", "Cable Assy", "케이블",
        6.41, 6.47, 0.06, 0.936, "AC2401", "AC2502",
        "AC2401.xls", "AC2502.xls", "2024-01-01", "2025-02-01", "USD",
    ])

    change = workbook.create_sheet("변경이력")
    change.append([
        "이벤트ID", "실행ID", "기록일시", "제조사", "요청번호", "원본파일",
        "원본행", "변경유형", "변경항목", "모델", "부품번호", "부품명(영문)",
        "부품명(한글)", "이전값", "변경값", "기준년도", "통화",
    ])
    change.append([
        "ID2", "RUN1", "2026-07-22 12:00:00", "IR", "IR26100",
        "IR26100.xlsx", 10, "변경", "단가", "A430", "P-200",
        "Air Bag", "에어백", "10", "11", "2026", "USD",
    ])
    workbook.save(path)


def test_load_and_filter_history(tmp_path):
    make_history(tmp_path / "History_2026.xlsx")
    rows = load_history_rows(tmp_path)
    assert len(rows) == 2

    assert len(filter_history_rows(rows, keyword="AC2502")) == 1
    assert len(filter_history_rows(rows, keyword="에어백")) == 1
    assert len(filter_history_rows(rows, manufacturer="AC")) == 1
    assert len(filter_history_rows(rows, category="변경이력")) == 1
