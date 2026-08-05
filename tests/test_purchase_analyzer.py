from openpyxl import Workbook

from core.parser import PartRecord
from core.purchase_analyzer import PurchaseItem, _analyze_item, parse_purchase_workbook


def test_purchase_workbook_parser(tmp_path):
    path = tmp_path / "purchase.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "IR"
    ws.append(["부품코드","부품명","부품명(Eng.)","AS사용","택배발송","재작업","물류팀대여","부품원가","판매수량","매출","재고","예약","미입고","신규\n부품발주 수량\n(EA)","신규\n부품발주 비용\n(달러)","원화","사용 유형","비고"])
    ws.append(["P-1","케이블, A100","Cable, A100",0,0,0,0,6.5,0,0,0,0,0,10,65,0,"유상",""])
    wb.save(path)
    items = parse_purchase_workbook(path)
    assert len(items) == 1
    assert items[0].part_no == "P-1"
    assert items[0].order_qty == 10
    assert items[0].order_unit_price == 6.5
    assert items[0].model == "A100"


def test_exact_part_number_price_increase():
    item = PurchaseItem("IR","IR",2,"P-1","케이블","Cable","A100",7.0,10,70,"","")
    records = [
        PartRecord(base_date="2025-01-01", ir_no="IR2501", model="A100", part_no="P-1", part_name_en="Cable", price=5.0),
        PartRecord(base_date="2026-01-01", ir_no="IR2601", model="A100", part_no="P-1", part_name_en="Cable", price=6.0),
    ]
    row = _analyze_item(item, records)
    assert row.status == "상승"
    assert row.master_latest_price == 6.0
    assert row.previous_price == 5.0
    assert row.impact_amount == 10.0
    assert row.match_method == "부품코드"


def test_unmatched_item_is_new():
    item = PurchaseItem("AC","AC",3,"","신규부품","New Part","L500",2.5,20,50,"","")
    row = _analyze_item(item, [])
    assert row.status == "신규부품"
    assert "Master 미등록" in row.risk


def test_multiple_prices_use_immediate_previous_and_keep_full_history(tmp_path):
    from core.purchase_analyzer import write_purchase_analysis, PurchaseAnalysisResult
    from openpyxl import load_workbook

    item = PurchaseItem(
        manufacturer="IR", source_sheet="IR", source_row=10,
        part_no="P-77", part_name_kr="테스트부품", part_name_en="Test Part",
        model="A210", order_unit_price=14.0, order_qty=10,
        order_amount=140, usage_type="", note="",
    )
    records = [
        PartRecord(base_date="2024-01-01", ir_no="IR24001", model="A210", part_no="P-77", part_name_en="Test Part", price=10.0),
        PartRecord(base_date="2024-03-01", ir_no="IR24058", model="A210", part_no="P-77", part_name_en="Test Part", price=12.0),
        PartRecord(base_date="2025-01-01", ir_no="IR25032", model="A210", part_no="P-77", part_name_en="Test Part", price=13.2),
        PartRecord(base_date="2026-02-01", ir_no="IR26012", model="A210", part_no="P-77", part_name_en="Test Part", price=14.0),
    ]
    row = _analyze_item(item, records)
    assert row.previous_price == 13.2
    assert row.master_latest_price == 14.0
    assert row.price_change_count == 3
    assert row.latest_request_no == "IR26012"
    assert row.previous_request_no == "IR25032"
    assert row.first_price == 10.0
    assert row.minimum_price == 10.0
    assert row.maximum_price == 14.0
    assert round(row.average_price, 2) == 12.30
    assert len(row.price_history) == 4

    result = PurchaseAnalysisResult(source_path=tmp_path / "source.xlsx", rows=[row])
    result.total_items = 1
    result.total_qty = 10
    result.total_order_amount = 140
    result.matched_items = 1
    result.unchanged_items = 1
    output = write_purchase_analysis(result, tmp_path)
    wb = load_workbook(output, data_only=False)
    assert "가격이력_상세" in wb.sheetnames
    detail = wb["전체_상세"]
    assert detail["I1"].value == "비교기준단가(Master 최신)"
    assert detail["L1"].value == "가격변경횟수"
    assert detail["P1"].value == "최근 변경일"
    assert detail["Q1"].value == "최근 변경 요청 No."
    assert detail["X2"].hyperlink is not None
    assert wb["가격이력_상세"].max_row == 5
