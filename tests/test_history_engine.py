from dataclasses import replace

from openpyxl import load_workbook

from core.history_engine import (
    append_history, compare_snapshots, make_history_run, make_run_id, summarize_events,
)
from core.parser import PartRecord


def record(**kwargs):
    base = PartRecord(
        ir_no="IR26001", model="A430", part_name_en="AIR BAG",
        part_name_kr="에어백", price=10.0, qty=1.0, amount=10.0,
        part_no="P-001", source_file="IR26001.xlsx", source_row=8,
        base_year="2026", currency="USD",
    )
    return replace(base, **kwargs)


def test_transaction_records_changes_and_dedup(tmp_path):
    old = record()
    changed = record(price=12.0)
    new = record(source_row=9, part_no="P-002", part_name_en="MOTOR")
    run_id = make_run_id("IR", ["IR26001.xlsx"])

    events = compare_snapshots("IR", [old], [changed, new], run_id=run_id)
    assert summarize_events(events) == (1, 0, 1)

    run = make_history_run(
        manufacturer="IR", run_id=run_id, source_files=["IR26001.xlsx"],
        request_nos=["IR26001"], before_rows=1, after_rows=2,
        events=events, reprocessed_rows=1, master_path="Master.xlsx",
        backup_path="Backup.xlsx",
    )
    path, added = append_history(tmp_path, events, run)
    assert path is not None
    assert added == len(events)

    _, added_again = append_history(tmp_path, events, run)
    assert added_again == 0

    wb = load_workbook(path, data_only=True)
    assert wb["실행이력"].max_row == 2
    assert wb["변경이력"].max_row == len(events) + 1


def test_zero_change_run_is_still_audited(tmp_path):
    old = record()
    run_id = make_run_id("IR", ["IR26001.xlsx"])
    events = compare_snapshots("IR", [old], [old], run_id=run_id)
    assert events == []
    run = make_history_run(
        manufacturer="IR", run_id=run_id, source_files=["IR26001.xlsx"],
        request_nos=["IR26001"], before_rows=1, after_rows=1,
        events=events, reprocessed_rows=1, master_path="Master.xlsx",
        backup_path="Backup.xlsx",
    )
    path, added = append_history(tmp_path, events, run)
    assert path is not None
    assert added == 0
    wb = load_workbook(path, data_only=True)
    assert wb["실행이력"].max_row == 2
    assert wb["변경이력"].max_row == 1
