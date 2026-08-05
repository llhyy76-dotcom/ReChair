from dataclasses import replace
from openpyxl import load_workbook

from core.comparison_engine import build_price_history, compare_master
from core.history_engine import append_history, make_history_run, make_run_id
from core.parser import PartRecord


def rec(**kwargs):
    base = PartRecord(
        ir_no="AC2401", model="L500", part_name_en="Cable Assy",
        part_name_kr="케이블", price=6.41, qty=1, amount=6.41,
        part_no="P-100", source_file="AC2401.xls", source_row=10,
        base_year=2024, base_date="2024-01-01", currency="USD",
    )
    return replace(base, **kwargs)


def test_price_history_matches_distinct_latest_transition():
    rows = [
        rec(),
        rec(ir_no="AC2502", source_file="AC2502.xls", source_row=11,
            base_year=2025, base_date="2025-02-01", price=6.47),
    ]
    events = build_price_history("AC", rows)
    assert len(events) == 1
    assert events[0].previous_price == 6.41
    assert events[0].latest_price == 6.47
    assert events[0].latest_request_no == "AC2502"


def test_price_history_sheet_is_persisted_when_snapshot_unchanged(tmp_path):
    rows = [
        rec(),
        rec(ir_no="AC2502", source_file="AC2502.xls", source_row=11,
            base_year=2025, base_date="2025-02-01", price=6.47),
    ]
    run_id = make_run_id("AC", ["AC2502.xls"])
    comparison = compare_master("AC", rows, rows, run_id=run_id)
    assert comparison.snapshot_events == []
    assert comparison.cumulative_price_variations == 1

    run = make_history_run(
        manufacturer="AC", run_id=run_id, source_files=["AC2502.xls"],
        request_nos=["AC2502"], before_rows=2, after_rows=2,
        events=[], reprocessed_rows=2, master_path="Master.xlsx",
        backup_path="Backup.xlsx", cumulative_price_variations=1,
    )
    path, added = append_history(
        tmp_path, comparison.snapshot_events, run, comparison.price_history_events
    )
    assert added == 0
    wb = load_workbook(path, data_only=True)
    assert wb["가격이력"].max_row == 2
    assert wb["실행이력"].max_row == 2
