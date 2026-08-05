from __future__ import annotations

import os
from pathlib import Path

from PySide6.QtCore import Qt, Slot
from PySide6.QtWidgets import (
    QComboBox, QDialog, QGridLayout, QGroupBox, QHBoxLayout, QLabel,
    QMessageBox, QPushButton, QPlainTextEdit, QVBoxLayout,
)
from openpyxl import load_workbook

from core.pdf_auto_approval import Decision, evaluate, threshold
from core.pdf_approval import process_approved_pdf_reviews


class PdfReviewManagerDialog(QDialog):
    """Human-in-the-loop PDF review and approval manager.

    The "AI 검토" is a deterministic local validation engine. It does not call
    an external ChatGPT API and never updates a Master by itself.
    """

    def __init__(self, root: Path, parent=None) -> None:
        super().__init__(parent)
        self.root = root
        self.review_dir = root / "PDF_Review"
        self.current_path: Path | None = None
        self.current_decision: Decision | None = None

        self.setWindowTitle("CPMS PDF 검토 및 승인")
        self.resize(900, 650)

        layout = QVBoxLayout(self)

        title = QLabel("PDF 검토 및 승인")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("font-size: 22px; font-weight: 700;")
        layout.addWidget(title)

        file_box = QGroupBox("검토파일")
        file_layout = QGridLayout(file_box)
        self.file_combo = QComboBox()
        self.btn_refresh = QPushButton("새로고침")
        self.btn_open = QPushButton("검토 Excel 열기")
        file_layout.addWidget(QLabel("파일"), 0, 0)
        file_layout.addWidget(self.file_combo, 0, 1, 1, 3)
        file_layout.addWidget(self.btn_refresh, 1, 0)
        file_layout.addWidget(self.btn_open, 1, 1)
        layout.addWidget(file_box)

        result_box = QGroupBox("자동 검토 결과")
        result_layout = QGridLayout(result_box)
        self.grade_label = QLabel("-")
        self.score_label = QLabel("-")
        self.request_label = QLabel("-")
        self.manufacturer_label = QLabel("-")
        self.rows_label = QLabel("-")
        self.confidence_label = QLabel("-")
        values = [
            ("등급", self.grade_label), ("점수", self.score_label),
            ("요청번호", self.request_label), ("제조사", self.manufacturer_label),
            ("품목수", self.rows_label), ("최저 신뢰도", self.confidence_label),
        ]
        for index, (name, widget) in enumerate(values):
            result_layout.addWidget(QLabel(name), index // 3 * 2, index % 3)
            result_layout.addWidget(widget, index // 3 * 2 + 1, index % 3)
            widget.setStyleSheet("font-size: 16px; font-weight: 700;")
        layout.addWidget(result_box)

        self.reason = QPlainTextEdit()
        self.reason.setReadOnly(True)
        self.reason.setPlaceholderText("검토 결과와 차단 사유가 표시됩니다.")
        layout.addWidget(self.reason, stretch=1)

        help_label = QLabel(
            "등급 A/B라도 검토 Excel을 직접 확인한 뒤 승인하세요. "
            "등급 C/D는 승인 표시가 차단됩니다. Master 반영 전 최종 검증이 다시 수행됩니다."
        )
        help_label.setWordWrap(True)
        layout.addWidget(help_label)

        buttons = QHBoxLayout()
        self.btn_grade = QPushButton("자동 검토 실행")
        self.btn_mark_approved = QPushButton("전체 행 승인 표시")
        self.btn_apply = QPushButton("승인 파일 Master 반영")
        self.btn_close = QPushButton("닫기")
        buttons.addWidget(self.btn_grade)
        buttons.addWidget(self.btn_mark_approved)
        buttons.addWidget(self.btn_apply)
        buttons.addStretch(1)
        buttons.addWidget(self.btn_close)
        layout.addLayout(buttons)

        self.btn_refresh.clicked.connect(self.refresh_files)
        self.btn_open.clicked.connect(self.open_excel)
        self.btn_grade.clicked.connect(self.run_grade)
        self.btn_mark_approved.clicked.connect(self.mark_approved)
        self.btn_apply.clicked.connect(self.apply_approved)
        self.btn_close.clicked.connect(self.accept)
        self.file_combo.currentTextChanged.connect(self.on_file_changed)

        self.refresh_files()

    @Slot()
    def refresh_files(self) -> None:
        self.review_dir.mkdir(parents=True, exist_ok=True)
        files = sorted(self.review_dir.glob("PDF_검토대기_*.xlsx"))
        current = self.file_combo.currentText()
        self.file_combo.blockSignals(True)
        self.file_combo.clear()
        self.file_combo.addItems([path.name for path in files])
        index = self.file_combo.findText(current)
        self.file_combo.setCurrentIndex(index if index >= 0 else (0 if files else -1))
        self.file_combo.blockSignals(False)
        self.on_file_changed(self.file_combo.currentText())

    @Slot(str)
    def on_file_changed(self, filename: str) -> None:
        self.current_path = self.review_dir / filename if filename else None
        self.current_decision = None
        self._display(None)

    def _display(self, decision: Decision | None) -> None:
        if decision is None:
            for label in (
                self.grade_label, self.score_label, self.request_label,
                self.manufacturer_label, self.rows_label, self.confidence_label,
            ):
                label.setText("-")
            self.reason.clear()
            self.btn_mark_approved.setEnabled(False)
            self.btn_apply.setEnabled(False)
            self.btn_apply.setText("승인 파일 Master 반영")
            self.btn_apply.setStyleSheet("")
            self.btn_apply.setToolTip("자동 검토를 먼저 실행하세요.")
            return

        self.grade_label.setText(decision.grade)
        self.score_label.setText(str(decision.score))
        self.request_label.setText(decision.request_no or "-")
        self.manufacturer_label.setText(decision.manufacturer or "-")
        self.rows_label.setText(str(decision.row_count))
        self.confidence_label.setText(f"{decision.min_confidence:.1f}%")
        self.reason.setPlainText(
            f"판정: {decision.decision}\n\n{decision.reason}"
        )
        approvable = decision.grade in {"A", "B"}
        self.btn_mark_approved.setEnabled(approvable)
        # C/D도 버튼을 눌러 차단 사유를 확인할 수 있게 유지합니다.
        self.btn_apply.setEnabled(True)
        if approvable:
            self.btn_apply.setText("승인 파일 Master 반영")
            self.btn_apply.setStyleSheet("")
            self.btn_apply.setToolTip("승인 표시된 검토파일을 최종 검증 후 Master에 반영합니다.")
        else:
            self.btn_apply.setText(f"승인 불가 ({decision.grade})")
            self.btn_apply.setStyleSheet(
                "QPushButton { color: #b00020; border: 1px solid #b00020; font-weight: 700; }"
            )
            self.btn_apply.setToolTip("클릭하면 Master 반영이 차단된 구체적인 사유를 확인할 수 있습니다.")

    @Slot()
    def run_grade(self) -> None:
        if not self.current_path or not self.current_path.exists():
            QMessageBox.information(self, "PDF 검토", "검토파일을 선택하세요.")
            return
        try:
            decision = evaluate(self.current_path, threshold(self.root))
            self.current_decision = decision
            self._display(decision)
        except Exception as exc:
            self.current_decision = None
            self._display(None)
            QMessageBox.critical(self, "자동 검토 실패", str(exc))

    @Slot()
    def open_excel(self) -> None:
        if not self.current_path or not self.current_path.exists():
            return
        try:
            os.startfile(self.current_path)  # type: ignore[attr-defined]
        except OSError as exc:
            QMessageBox.warning(self, "파일 열기 실패", str(exc))

    @Slot()
    def mark_approved(self) -> None:
        decision = self.current_decision
        if not decision or decision.grade not in {"A", "B"}:
            QMessageBox.warning(
                self, "승인 차단",
                "자동 검토 등급이 A 또는 B인 파일만 승인 표시할 수 있습니다."
            )
            return
        if not self.current_path:
            return

        reply = QMessageBox.question(
            self,
            "승인 표시",
            f"{decision.request_no} / {decision.manufacturer}\n"
            f"{decision.row_count}개 품목을 승인 상태로 표시할까요?\n\n"
            "아직 Master에는 반영되지 않습니다.",
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            wb = load_workbook(self.current_path)
            ws = wb["PDF_검토대기"]
            headers = {
                str(ws.cell(1, col).value or "").strip(): col
                for col in range(1, ws.max_column + 1)
            }
            status_col = headers["검토상태"]
            part_col = headers["부품명(영어)"]
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row, part_col).value or "").strip():
                    ws.cell(row, status_col, "승인")
            wb.save(self.current_path)
            QMessageBox.information(
                self, "승인 표시 완료",
                "검토파일의 모든 유효 행을 '승인'으로 표시했습니다.\n"
                "검토 Excel을 닫은 상태에서 Master 반영을 진행하세요.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "승인 표시 실패", str(exc))

    @Slot()
    def apply_approved(self) -> None:
        # 선택 파일이 C/D 등급이면 아무 반응 없이 끝내지 않고 차단 사유를 안내합니다.
        decision = self.current_decision
        if self.current_path and self.current_path.exists() and decision is None:
            try:
                decision = evaluate(self.current_path, threshold(self.root))
                self.current_decision = decision
                self._display(decision)
            except Exception as exc:
                QMessageBox.critical(self, "자동 검토 실패", str(exc))
                return

        if decision and decision.grade not in {"A", "B"}:
            reasons = [item.strip() for item in decision.reason.split(";") if item.strip()]
            reason_text = "\n".join(f"• {item}" for item in reasons) or "• 승인 기준을 충족하지 못했습니다."
            QMessageBox.warning(
                self,
                "Master 반영 차단",
                f"이 검토파일은 Master에 반영할 수 없습니다.\n\n"
                f"등급: {decision.grade}\n"
                f"판정: {decision.decision}\n"
                f"점수: {decision.score}점\n\n"
                f"차단 사유\n{reason_text}\n\n"
                "검토 Excel에서 제조사, 요청번호, 적용모델, 부품명, 수량·단가·금액을 보완한 뒤 "
                "자동 검토를 다시 실행해 주세요.",
            )
            return

        reply = QMessageBox.question(
            self,
            "Master 반영 확인",
            "승인 표시된 PDF 검토파일을 제조사 Master에 반영합니다.\n\n"
            "요청번호의 제조사와 검토파일 제조사가 다르거나, 수량·단가·금액·"
            "적용모델 검증에 실패하면 자동으로 차단됩니다.\n\n계속할까요?",
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            result = process_approved_pdf_reviews(self.root)
            lines = [
                f"검토파일: {result.scanned_files}",
                f"반영 완료: {result.approved_files}",
                f"승인 대기: {result.waiting_files}",
                f"실패/차단: {result.failed_files}",
                f"반영 품목: {result.total_rows}",
                f"신규부품: {result.total_new_parts}",
                f"가격변동: {result.total_price_changes}",
            ]
            for item in result.files:
                if item.errors:
                    lines.append(
                        f"\n{item.review_file.name}: {item.status}\n"
                        + "\n".join(item.errors)
                    )
            QMessageBox.information(
                self, "PDF 승인 반영 결과", "\n".join(lines)
            )
            self.refresh_files()
        except Exception as exc:
            QMessageBox.critical(self, "PDF 승인 반영 실패", str(exc))
