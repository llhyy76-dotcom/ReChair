from __future__ import annotations

import math
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.parser import PartRecord, clean_text, norm_model, norm_txt, parse_master, parse_number
from core.matching_engine import infer_item_model, infer_record_model, select_match


ProgressCallback = Callable[[int, str], None]
SUPPORTED_MANUFACTURERS = ("IR", "XC", "AC", "KA", "NC")


def _compact(value: object) -> str:
    return re.sub(r"[\s\-_./(),]+", "", clean_text(value)).upper()


def _header(value: object) -> str:
    return re.sub(r"\s+", "", clean_text(value)).replace("\n", "")


def _number(value: object) -> float | None:
    return parse_number(value)


def _safe_float(value: object) -> float | None:
    number = _number(value)
    if number is None or not math.isfinite(number):
        return None
    return float(number)


@dataclass(frozen=True)
class PurchaseItem:
    manufacturer: str
    source_sheet: str
    source_row: int
    part_no: str
    part_name_kr: str
    part_name_en: str
    model: str
    order_unit_price: float | None
    order_qty: float
    order_amount: float | None
    usage_type: str
    note: str


@dataclass(frozen=True)
class MasterMatch:
    method: str
    confidence: int
    records: tuple[PartRecord, ...]


@dataclass(frozen=True)
class PriceHistoryEntry:
    sequence: int
    price: float
    request_no: str
    base_date: str
    source_file: str
    source_sheet: str
    source_row: int
    change_amount: float | None
    change_rate: float | None


@dataclass(frozen=True)
class PurchaseAnalysisRow:
    manufacturer: str
    part_no: str
    part_name_kr: str
    part_name_en: str
    model: str
    order_qty: float
    order_unit_price: float | None
    order_amount: float | None
    previous_price: float | None
    master_latest_price: float | None
    first_price: float | None
    minimum_price: float | None
    maximum_price: float | None
    average_price: float | None
    previous_request_no: str
    previous_base_date: str
    unit_difference: float | None
    change_rate: float | None
    impact_amount: float | None
    status: str
    match_method: str
    match_confidence: int
    latest_request_no: str
    latest_base_date: str
    latest_source_file: str
    price_change_count: int
    price_history_summary: str
    price_history: tuple[PriceHistoryEntry, ...]
    risk: str
    usage_type: str
    note: str
    source_sheet: str
    source_row: int
    intelligence_grade: str
    intelligence_reason: str
    recommendation: str
    latest_deviation_rate: float | None
    average_deviation_rate: float | None


@dataclass
class PurchaseAnalysisResult:
    source_path: Path
    output_path: Path | None = None
    negotiation_output_path: Path | None = None
    rows: list[PurchaseAnalysisRow] = field(default_factory=list)
    total_items: int = 0
    total_qty: float = 0
    total_order_amount: float = 0
    matched_items: int = 0
    increased_items: int = 0
    decreased_items: int = 0
    unchanged_items: int = 0
    new_items: int = 0
    review_items: int = 0
    net_impact: float = 0
    errors: list[str] = field(default_factory=list)


HEADER_ALIASES = {
    "part_no": ("부품코드", "품번", "품번/코드", "PARTNO", "PARTNUMBER", "CODE"),
    "part_name_kr": ("부품명", "부품명(한글)", "부품명(KR)"),
    "part_name_en": ("부품명(ENG.)", "부품명(영어)", "부품명(EN)", "PARTNAME", "DESCRIPTION"),
    "unit_price": ("부품원가", "신규부품발주단가", "발주단가", "UNITPRICE", "PRICE"),
    "order_qty": ("신규부품발주수량(EA)", "신규부품발주수량", "발주수량", "요청수량", "요청 수량", "ORDERQTY", "QTY", "Q'TY"),
    "order_amount": ("신규부품발주비용(달러)", "신규부품발주비용", "발주금액(달러)", "합계", "총액", "AMOUNT(USD)", "AMOUNT", "TOTAL"),
    "usage_type": ("사용유형", "유형"),
    "note": ("비고", "NOTE", "REMARK"),
}


def _find_columns(ws) -> tuple[int, dict[str, int]]:
    """Find purchase columns while preserving duplicate visible headers.

    Some legacy XC sheets have two adjacent columns both labeled '부품명':
    the first contains English names or part numbers and the second contains
    Korean names. A dict keyed only by header text loses the first occurrence.
    """
    for row in range(1, min(ws.max_row, 30) + 1):
        header_cells = [
            (_header(ws.cell(row, col).value), col)
            for col in range(1, ws.max_column + 1)
        ]
        found: dict[str, int] = {}

        # First resolve unique/explicit aliases.
        for field, aliases in HEADER_ALIASES.items():
            alias_keys = {_header(alias) for alias in aliases}
            matches = [col for key, col in header_cells if key in alias_keys]
            if matches:
                found[field] = matches[0]

        # Legacy duplicate '부품명' columns:
        duplicate_part_cols = [
            col for key, col in header_cells
            if key in {_header("부품명"), _header("PART NAME")}
        ]
        if len(duplicate_part_cols) >= 2:
            found["part_name_en"] = duplicate_part_cols[0]
            found["part_name_kr"] = duplicate_part_cols[1]
            found["_mixed_part_column"] = 1
            # Do not treat the first duplicate column as a definite part number;
            # row-level logic decides whether it is a code or an English name.

        # Explicit English/Korean labels override the legacy fallback.
        for field, aliases in {
            "part_name_en": HEADER_ALIASES["part_name_en"],
            "part_name_kr": HEADER_ALIASES["part_name_kr"],
            "part_no": HEADER_ALIASES["part_no"],
        }.items():
            alias_keys = {_header(alias) for alias in aliases}
            matches = [col for key, col in header_cells if key in alias_keys]
            if matches:
                # Generic '부품명' alone should not override duplicate handling.
                explicit = [
                    col for key, col in header_cells
                    if key in alias_keys and key not in {_header("부품명"), _header("PART NAME")}
                ]
                if explicit:
                    found[field] = explicit[0]

        if "order_qty" in found and any(
            field in found for field in ("part_no", "part_name_kr", "part_name_en")
        ):
            return row, found
    raise ValueError(f"{ws.title}: 발주 수량 및 부품 열을 찾지 못했습니다.")


def _looks_like_part_code(value: object) -> bool:
    text = clean_text(value).strip().upper()
    if not text:
        return False
    # Examples: C101110-0400, XC-P-100, 12345-0670.
    return bool(
        re.fullmatch(r"[A-Z]{0,5}\d{3,}(?:[-_/]\d{2,})+", text)
        or re.fullmatch(r"[A-Z]{1,5}[-_/]\d{1,}(?:[-_/][A-Z0-9]+)*", text)
        or re.fullmatch(r"[A-Z]{2,5}[-_/][A-Z0-9]+(?:[-_/][A-Z0-9]+)+", text)
    )


def _extract_model_candidates(*texts: object) -> tuple[str, ...]:
    joined = " ".join(clean_text(value) for value in texts if clean_text(value)).upper()
    candidates: list[str] = []

    # Composite comma tokens: BR,1320 -> BR1320.
    tokens = [token.strip() for token in re.split(r"[,/]", joined) if token.strip()]
    for index, token in enumerate(tokens):
        if re.fullmatch(r"[A-Z]{1,3}", token) and index + 1 < len(tokens):
            nxt = re.sub(r"\(G\)", "", tokens[index + 1]).strip()
            if re.fullmatch(r"\d{2,5}[A-Z]?", nxt):
                candidates.append(f"{token}{nxt}")

    # Standalone model-like tokens.
    for token in tokens:
        normalized = norm_model(token)
        if normalized and re.fullmatch(r"[A-Z]{0,3}\d{2,5}[A-Z]?", normalized):
            candidates.append(normalized)

    # General embedded model patterns.
    for match in re.findall(
        r"(?:CMC-|CMS-|HM-)?([A-Z]{0,3}\d{2,5}[A-Z]?)(?:\(G\))?",
        joined,
    ):
        normalized = norm_model(match)
        if normalized:
            candidates.append(normalized)

    # Preserve order, preferring composite variants first.
    result: list[str] = []
    for model in candidates:
        if model not in result:
            result.append(model)
    return tuple(result)



def _infer_model(*texts: object) -> str:
    candidates = _extract_model_candidates(*texts)
    return "/".join(candidates[:3]) if candidates else ""


def parse_purchase_workbook(path: str | Path) -> list[PurchaseItem]:
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    items: list[PurchaseItem] = []
    try:
        for ws in workbook.worksheets:
            manufacturer = ws.title.strip().upper()
            if manufacturer not in SUPPORTED_MANUFACTURERS:
                continue
            try:
                header_row, columns = _find_columns(ws)
            except ValueError:
                continue

            for row in range(header_row + 1, ws.max_row + 1):
                part_no = clean_text(ws.cell(row, columns.get("part_no", 0)).value) if columns.get("part_no") else ""
                kr = clean_text(ws.cell(row, columns.get("part_name_kr", 0)).value) if columns.get("part_name_kr") else ""
                en = clean_text(ws.cell(row, columns.get("part_name_en", 0)).value) if columns.get("part_name_en") else ""

                # In legacy XC files the first duplicate '부품명' column may
                # contain either an English name or an actual part code.
                if columns.get("_mixed_part_column"):
                    if not part_no and _looks_like_part_code(en):
                        part_no, en = en, ""
                elif part_no and not en and not _looks_like_part_code(part_no):
                    en, part_no = part_no, ""

                if any(value.upper() in {"합계", "TOTAL"} for value in (part_no, kr, en) if value):
                    continue

                qty = _safe_float(ws.cell(row, columns["order_qty"]).value)
                if qty is None or qty <= 0:
                    continue
                if not any((part_no, kr, en)):
                    continue

                unit_price = _safe_float(ws.cell(row, columns.get("unit_price", 0)).value) if columns.get("unit_price") else None
                amount = _safe_float(ws.cell(row, columns.get("order_amount", 0)).value) if columns.get("order_amount") else None
                if unit_price is None and amount is not None and qty:
                    unit_price = amount / qty
                if amount is None and unit_price is not None:
                    amount = unit_price * qty

                usage = clean_text(ws.cell(row, columns.get("usage_type", 0)).value) if columns.get("usage_type") else ""
                note = clean_text(ws.cell(row, columns.get("note", 0)).value) if columns.get("note") else ""
                items.append(PurchaseItem(
                    manufacturer=manufacturer,
                    source_sheet=ws.title,
                    source_row=row,
                    part_no=part_no,
                    part_name_kr=kr,
                    part_name_en=en,
                    model=_infer_model(kr, en),
                    order_unit_price=unit_price,
                    order_qty=qty,
                    order_amount=amount,
                    usage_type=usage,
                    note=note,
                ))
    finally:
        workbook.close()
    if not items:
        raise ValueError("발주수량이 입력된 품목을 찾지 못했습니다.")
    return items


def _sort_key(record: PartRecord) -> tuple[str, int, str, int]:
    return (
        str(record.base_date or ""),
        int(record.base_year or 0),
        str(record.ir_no or ""),
        int(record.source_row or 0),
    )


def _build_indexes(records: Iterable[PartRecord]) -> dict[str, dict[str, list[PartRecord]]]:
    indexes: dict[str, dict[str, list[PartRecord]]] = {
        "part_no": defaultdict(list),
        "en_model": defaultdict(list),
        "kr_model": defaultdict(list),
        "en": defaultdict(list),
        "kr": defaultdict(list),
        "model": defaultdict(list),
        "all": {"records": []},
    }
    for record in records:
        indexes["all"]["records"].append(record)
        if clean_text(record.part_no):
            indexes["part_no"][_compact(record.part_no)].append(record)

        model = infer_record_model(record)
        en = _compact(record.part_name_en)
        kr = _compact(record.part_name_kr)
        if model:
            indexes["model"][model].append(record)
        if en and model:
            indexes["en_model"][f"{en}|{model}"].append(record)
        if kr and model:
            indexes["kr_model"][f"{kr}|{model}"].append(record)
        if en:
            indexes["en"][en].append(record)
        if kr:
            indexes["kr"][kr].append(record)
    return indexes


def _match(item: PurchaseItem, indexes: dict[str, dict[str, list[PartRecord]]]) -> MasterMatch:
    method, confidence, records = select_match(
        indexes["all"].get("records", []),
        item_model=item.model,
        part_no=item.part_no,
        part_name_en=item.part_name_en,
        part_name_kr=item.part_name_kr,
        minimum_score=82,
    )
    return MasterMatch(method, confidence, records)


def _expand_history_records(
    item: PurchaseItem,
    primary: MasterMatch,
    indexes: dict[str, dict[str, list[PartRecord]]],
) -> MasterMatch:
    """Expand history when the code changed but normalized part name and model stayed identical."""
    if not primary.records:
        return primary

    candidates: list[PartRecord] = list(primary.records)
    model = infer_item_model(item.model, item.part_name_en, item.part_name_kr)
    en = _compact(item.part_name_en)
    kr = _compact(item.part_name_kr)

    if en and model:
        candidates.extend(indexes["en_model"].get(f"{en}|{model}", []))
    if kr and model:
        candidates.extend(indexes["kr_model"].get(f"{kr}|{model}", []))

    unique: list[PartRecord] = []
    seen: set[tuple[str, str, int, str, float | None]] = set()
    for record in candidates:
        key = (
            clean_text(record.source_file),
            clean_text(getattr(record, "source_sheet", "")),
            int(record.source_row or 0),
            clean_text(record.ir_no),
            float(record.price) if record.price is not None else None,
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(record)

    if len(unique) > len(primary.records):
        return MasterMatch(f"{primary.method}+이력확장", primary.confidence, tuple(unique))
    return primary


def _price_history(records: Iterable[PartRecord]) -> list[PartRecord]:
    ordered = sorted((r for r in records if r.price is not None), key=_sort_key)
    distinct: list[PartRecord] = []
    for record in ordered:
        price = round(float(record.price), 8)
        if not distinct or round(float(distinct[-1].price), 8) != price:
            distinct.append(record)
    return distinct


def _history_entries(history: list[PartRecord]) -> tuple[PriceHistoryEntry, ...]:
    entries: list[PriceHistoryEntry] = []
    previous_price: float | None = None
    for sequence, record in enumerate(history, start=1):
        price = float(record.price)
        change_amount = None if previous_price is None else price - previous_price
        change_rate = None if previous_price in (None, 0) else change_amount / previous_price * 100
        entries.append(PriceHistoryEntry(
            sequence=sequence,
            price=price,
            request_no=clean_text(record.ir_no),
            base_date=str(record.base_date or record.base_year or ""),
            source_file=clean_text(record.source_file),
            source_sheet=clean_text(getattr(record, "source_sheet", "")),
            source_row=int(record.source_row or 0),
            change_amount=change_amount,
            change_rate=change_rate,
        ))
        previous_price = price
    return tuple(entries)



def _evaluate_price_intelligence(
    *,
    order_price: float | None,
    latest_price: float | None,
    average_price: float | None,
    maximum_price: float | None,
    match_confidence: int,
    status: str,
) -> tuple[str, str, str, float | None, float | None]:
    """Evaluate the purchase price using only CPMS price history."""
    if status == "신규부품":
        return (
            "자료부족",
            "Master에 비교 가능한 동일 부품 이력이 없습니다.",
            "신규 부품 여부와 공급사 견적 근거를 확인하세요.",
            None,
            None,
        )
    if order_price is None or latest_price is None:
        return (
            "확인필요",
            "발주단가 또는 Master 최신단가가 없어 자동 판정할 수 없습니다.",
            "누락 단가와 매칭 결과를 확인하세요.",
            None,
            None,
        )
    if match_confidence < 80:
        return (
            "매칭확인",
            f"Master 매칭 신뢰도가 {match_confidence}%로 낮습니다.",
            "부품코드·모델·부품명이 같은 품목인지 먼저 확인하세요.",
            None,
            None,
        )

    latest_rate = None if latest_price == 0 else (order_price - latest_price) / latest_price * 100
    average_rate = (
        None if average_price in (None, 0)
        else (order_price - float(average_price)) / float(average_price) * 100
    )

    if latest_rate >= 30:
        grade = "매우비쌈"
        reason = f"Master 최신단가보다 {latest_rate:.1f}% 높습니다."
        recommendation = "발주 전 공급사 단가 재확인 및 인상 사유 확인을 권장합니다."
    elif latest_rate >= 10:
        grade = "비쌈"
        reason = f"Master 최신단가보다 {latest_rate:.1f}% 높습니다."
        recommendation = "최근 요청 No.와 견적서를 대조하고 단가 협의를 검토하세요."
    elif latest_rate > 3:
        grade = "주의"
        reason = f"Master 최신단가보다 {latest_rate:.1f}% 높습니다."
        recommendation = "소폭 인상 근거를 확인한 뒤 발주하세요."
    elif latest_rate >= -3:
        grade = "적정"
        reason = f"Master 최신단가 대비 차이가 {latest_rate:.1f}%로 안정 범위입니다."
        recommendation = "현 단가 기준 발주 검토가 가능합니다."
    elif latest_rate > -10:
        grade = "저렴"
        reason = f"Master 최신단가보다 {abs(latest_rate):.1f}% 낮습니다."
        recommendation = "단가 인하가 맞는지 품질·사양 변경 여부를 함께 확인하세요."
    else:
        grade = "매우저렴"
        reason = f"Master 최신단가보다 {abs(latest_rate):.1f}% 낮습니다."
        recommendation = "오입력 또는 사양 변경 가능성이 있으므로 견적 내용을 확인하세요."

    maximum_excess_rate = (
        None
        if maximum_price in (None, 0)
        else (order_price - float(maximum_price)) / float(maximum_price) * 100
    )
    if maximum_excess_rate is not None and maximum_excess_rate > 3:
        reason += f" 과거 최고단가도 {maximum_excess_rate:.1f}% 초과했습니다."
        if maximum_excess_rate >= 10 and grade in {"적정", "주의", "비쌈"}:
            grade = "매우비쌈"
            recommendation = "과거 최고단가 초과 사유를 확인한 뒤 승인하세요."

    if average_rate is not None:
        reason += f" 과거 평균단가 대비 {average_rate:+.1f}%입니다."

    return grade, reason, recommendation, latest_rate, average_rate

def _analyze_item(item: PurchaseItem, records: list[PartRecord]) -> PurchaseAnalysisRow:
    indexes = _build_indexes(records)
    match = _match(item, indexes)
    match = _expand_history_records(item, match, indexes)
    history = _price_history(match.records)

    latest = history[-1] if history else None
    previous = history[-2] if len(history) >= 2 else None
    latest_price = float(latest.price) if latest and latest.price is not None else None
    previous_price = float(previous.price) if previous and previous.price is not None else None
    prices = [float(record.price) for record in history if record.price is not None]
    first_price = prices[0] if prices else None
    minimum_price = min(prices) if prices else None
    maximum_price = max(prices) if prices else None
    average_price = sum(prices) / len(prices) if prices else None
    price_history = _history_entries(history)
    price_history_summary = " → ".join(f"{price:.4f}" for price in prices)

    difference = None
    rate = None
    impact = None
    if item.order_unit_price is not None and latest_price is not None:
        difference = item.order_unit_price - latest_price
        rate = None if latest_price == 0 else difference / latest_price * 100
        impact = difference * item.order_qty

    if not match.records:
        status = "신규부품"
    elif item.order_unit_price is None or latest_price is None:
        status = "확인필요"
    elif abs(difference or 0) < 0.000001:
        status = "동일"
    elif (difference or 0) > 0:
        status = "상승"
    else:
        status = "하락"

    risks: list[str] = []
    if status == "신규부품":
        risks.append("Master 미등록")
    if match.records and match.confidence < 82:
        risks.append("매칭 확인")
    if item.model and match.method == "미매칭":
        risks.append(f"{item.model} 모델 내 동일부품 미확인")
    if rate is not None and rate >= 10:
        risks.append("10% 이상 상승")
    if len(history) - 1 >= 3:
        risks.append("가격변경 3회 이상")
    if item.part_no == "":
        risks.append("부품코드 없음")

    (
        intelligence_grade,
        intelligence_reason,
        recommendation,
        latest_deviation_rate,
        average_deviation_rate,
    ) = _evaluate_price_intelligence(
        order_price=item.order_unit_price,
        latest_price=latest_price,
        average_price=average_price,
        maximum_price=maximum_price,
        match_confidence=match.confidence,
        status=status,
    )

    return PurchaseAnalysisRow(
        manufacturer=item.manufacturer,
        part_no=item.part_no,
        part_name_kr=item.part_name_kr,
        part_name_en=item.part_name_en,
        model=item.model,
        order_qty=item.order_qty,
        order_unit_price=item.order_unit_price,
        order_amount=item.order_amount,
        previous_price=previous_price,
        master_latest_price=latest_price,
        first_price=first_price,
        minimum_price=minimum_price,
        maximum_price=maximum_price,
        average_price=average_price,
        previous_request_no=previous.ir_no if previous else "",
        previous_base_date=str(previous.base_date or previous.base_year or "") if previous else "",
        unit_difference=difference,
        change_rate=rate,
        impact_amount=impact,
        status=status,
        match_method=match.method,
        match_confidence=match.confidence,
        latest_request_no=latest.ir_no if latest else "",
        latest_base_date=str(latest.base_date or latest.base_year or "") if latest else "",
        latest_source_file=latest.source_file if latest else "",
        price_change_count=max(len(history) - 1, 0),
        price_history_summary=price_history_summary,
        price_history=price_history,
        risk=", ".join(risks),
        usage_type=item.usage_type,
        note=item.note,
        source_sheet=item.source_sheet,
        source_row=item.source_row,
        intelligence_grade=intelligence_grade,
        intelligence_reason=intelligence_reason,
        recommendation=recommendation,
        latest_deviation_rate=latest_deviation_rate,
        average_deviation_rate=average_deviation_rate,
    )


def analyze_purchase(
    source_path: str | Path,
    master_paths: dict[str, str | Path],
    progress: ProgressCallback | None = None,
) -> PurchaseAnalysisResult:
    source_path = Path(source_path)
    items = parse_purchase_workbook(source_path)
    records_by_manufacturer: dict[str, list[PartRecord]] = {}
    errors: list[str] = []

    manufacturers = sorted({item.manufacturer for item in items})
    for index, manufacturer in enumerate(manufacturers, start=1):
        if progress:
            progress(int(index / max(len(manufacturers), 1) * 25), f"{manufacturer} Master 읽는 중")
        path = Path(master_paths.get(manufacturer, ""))
        if not path.exists():
            records_by_manufacturer[manufacturer] = []
            errors.append(f"{manufacturer} Master 없음: {path}")
            continue
        try:
            records_by_manufacturer[manufacturer] = parse_master(path)
        except Exception as exc:
            records_by_manufacturer[manufacturer] = []
            errors.append(f"{manufacturer} Master 읽기 실패: {exc}")

    rows: list[PurchaseAnalysisRow] = []
    total = len(items)
    for index, item in enumerate(items, start=1):
        rows.append(_analyze_item(item, records_by_manufacturer.get(item.manufacturer, [])))
        if progress and (index == total or index % 10 == 0):
            progress(25 + int(index / total * 55), f"발주품목 비교 {index}/{total}")

    result = PurchaseAnalysisResult(source_path=source_path, rows=rows, errors=errors)
    result.total_items = len(rows)
    result.total_qty = sum(row.order_qty for row in rows)
    result.total_order_amount = sum(row.order_amount or 0 for row in rows)
    result.matched_items = sum(row.status not in {"신규부품"} for row in rows)
    result.increased_items = sum(row.status == "상승" for row in rows)
    result.decreased_items = sum(row.status == "하락" for row in rows)
    result.unchanged_items = sum(row.status == "동일" for row in rows)
    result.new_items = sum(row.status == "신규부품" for row in rows)
    result.review_items = sum(row.status == "확인필요" or "매칭 확인" in row.risk for row in rows)
    result.net_impact = sum(row.impact_amount or 0 for row in rows)
    return result


DETAIL_HEADERS = [
    "제조사", "부품코드", "부품명", "부품명(Eng.)", "모델",
    "발주수량", "발주단가", "발주금액",
    "비교기준단가(Master 최신)", "직전 이력단가", "단가이력 요약", "가격변경횟수",
    "단가차이", "변동률(%)", "발주영향금액",
    "최근 변경일", "최근 변경 요청 No.",
    "직전단가 기준일", "직전단가 요청 No.",
    "최초단가", "최저단가", "최고단가", "평균단가", "가격이력 보기",
    "판정", "매칭방식", "매칭신뢰도", "최근 원본파일",
    "위험표시", "사용유형", "비고", "원본시트", "원본행",
    "가격판정", "판정근거", "권고사항", "최신단가 대비(%)", "평균단가 대비(%)",
]



def _style_header(ws, row: int, last_col: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B7C9E2")
    for col in range(1, last_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _history_anchor(row: PurchaseAnalysisRow) -> str:
    return f"{row.manufacturer}|{row.part_no}|{row.model}|{row.source_sheet}|{row.source_row}"


def _write_detail_sheet(
    wb: Workbook,
    name: str,
    rows: list[PurchaseAnalysisRow],
    history_links: dict[str, int],
) -> None:
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for col, header in enumerate(DETAIL_HEADERS, start=1):
        ws.cell(1, col).value = header
    _style_header(ws, 1, len(DETAIL_HEADERS))
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(rows, start=2):
        values = [
            row.manufacturer, row.part_no, row.part_name_kr, row.part_name_en, row.model,
            row.order_qty, row.order_unit_price, row.order_amount,
            row.master_latest_price, row.previous_price, row.price_history_summary,
            row.price_change_count, row.unit_difference, row.change_rate, row.impact_amount,
            row.latest_base_date, row.latest_request_no,
            row.previous_base_date, row.previous_request_no,
            row.first_price, row.minimum_price, row.maximum_price, row.average_price,
            "이력보기" if row.price_history else "",
            row.status, row.match_method, row.match_confidence, row.latest_source_file,
            row.risk, row.usage_type, row.note, row.source_sheet, row.source_row,
            row.intelligence_grade, row.intelligence_reason, row.recommendation,
            row.latest_deviation_rate, row.average_deviation_rate,
        ]
        for c_idx, value in enumerate(values, start=1):
            ws.cell(r_idx, c_idx).value = value
        anchor_row = history_links.get(_history_anchor(row))
        if anchor_row:
            cell = ws.cell(r_idx, 24)
            cell.hyperlink = f"#'가격이력_상세'!A{anchor_row}"
            cell.style = "Hyperlink"

    status_fill = {
        "상승": "F4CCCC", "하락": "D9EAF7", "동일": "E2F0D9",
        "신규부품": "FFF2CC", "확인필요": "FCE4D6",
    }
    thin = Side(style="thin", color="D9E2F3")
    for row_idx in range(2, ws.max_row + 1):
        status = clean_text(ws.cell(row_idx, 25).value)
        fill = PatternFill("solid", fgColor=status_fill.get(status, "FFFFFF"))
        for col in range(1, len(DETAIL_HEADERS) + 1):
            cell = ws.cell(row_idx, col)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="맑은 고딕", size=9)
        ws.cell(row_idx, 25).fill = fill
        if clean_text(ws.cell(row_idx, 29).value):
            ws.cell(row_idx, 29).fill = PatternFill("solid", fgColor="FCE4D6")

    price_columns = (7, 8, 9, 10, 13, 15, 20, 21, 22, 23)
    for col in price_columns:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row_idx, col).number_format = '#,##0.0000;[Red]-#,##0.0000'
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, 14).number_format = '0.00%;[Red]-0.00%'
        if ws.cell(row_idx, 14).value is not None:
            ws.cell(row_idx, 14).value = float(ws.cell(row_idx, 14).value) / 100
    for col in (6, 12, 27, 33):
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row_idx, col).number_format = '#,##0.##'

    grade_fill = {
        "매우비쌈": "F4CCCC", "비쌈": "FCE4D6", "주의": "FFF2CC",
        "적정": "E2F0D9", "저렴": "D9EAF7", "매우저렴": "DDEBF7",
        "자료부족": "E7E6E6", "확인필요": "FCE4D6", "매칭확인": "FCE4D6",
    }
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, 34).fill = PatternFill(
            "solid", fgColor=grade_fill.get(clean_text(ws.cell(row_idx, 34).value), "FFFFFF")
        )
        for col in (37, 38):
            ws.cell(row_idx, col).number_format = '0.00%;[Red]-0.00%'
            if ws.cell(row_idx, col).value is not None:
                ws.cell(row_idx, col).value = float(ws.cell(row_idx, col).value) / 100

    widths = {
        1: 9, 2: 19, 3: 29, 4: 42, 5: 12, 6: 11, 7: 12, 8: 14,
        9: 20, 10: 15, 11: 42, 12: 12, 13: 12, 14: 11, 15: 15,
        16: 14, 17: 18, 18: 15, 19: 18, 20: 12, 21: 12, 22: 12,
        23: 12, 24: 12, 25: 11, 26: 18, 27: 11, 28: 42, 29: 25,
        30: 14, 31: 25, 32: 11, 33: 9,
        34: 13, 35: 45, 36: 48, 37: 16, 38: 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_HEADERS))}{max(ws.max_row, 1)}"


def _write_price_history_sheet(
    wb: Workbook,
    rows: list[PurchaseAnalysisRow],
) -> dict[str, int]:
    ws = wb.create_sheet("가격이력_상세")
    ws.sheet_view.showGridLines = False
    headers = [
        "제조사", "부품코드", "부품명", "부품명(Eng.)", "모델",
        "순번", "기준일", "요청 No.", "단가", "직전대비", "변동률(%)",
        "원본파일", "원본시트", "원본행", "발주서 원본시트", "발주서 원본행",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    links: dict[str, int] = {}
    out_row = 2
    for analysis_row in rows:
        if not analysis_row.price_history:
            continue
        links[_history_anchor(analysis_row)] = out_row
        for entry in analysis_row.price_history:
            values = [
                analysis_row.manufacturer, analysis_row.part_no, analysis_row.part_name_kr,
                analysis_row.part_name_en, analysis_row.model, entry.sequence, entry.base_date,
                entry.request_no, entry.price, entry.change_amount, entry.change_rate,
                entry.source_file, entry.source_sheet, entry.source_row,
                analysis_row.source_sheet, analysis_row.source_row,
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(out_row, col).value = value
            if entry.sequence == len(analysis_row.price_history):
                for col in range(1, len(headers) + 1):
                    ws.cell(out_row, col).fill = PatternFill("solid", fgColor="E2F0D9")
            out_row += 1

    thin = Side(style="thin", color="D9E2F3")
    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row_idx, col)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.font = Font(name="맑은 고딕", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for col in (9, 10):
            ws.cell(row_idx, col).number_format = '#,##0.0000;[Red]-#,##0.0000'
        ws.cell(row_idx, 11).number_format = '0.00%;[Red]-0.00%'
        if ws.cell(row_idx, 11).value is not None:
            ws.cell(row_idx, 11).value = float(ws.cell(row_idx, 11).value) / 100

    widths = [9, 19, 29, 42, 12, 8, 14, 18, 12, 12, 11, 42, 18, 9, 15, 11]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
    return links



def _write_price_intelligence_sheet(
    wb: Workbook,
    rows: list[PurchaseAnalysisRow],
) -> None:
    ws = wb.create_sheet("가격인텔리전스")
    ws.sheet_view.showGridLines = False
    headers = [
        "우선순위", "제조사", "부품코드", "부품명", "모델",
        "발주수량", "발주단가", "Master 최신단가", "과거 평균단가",
        "최신 대비(%)", "평균 대비(%)", "발주영향금액",
        "가격판정", "판정근거", "권고사항", "최근 요청 No.", "가격변경횟수",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col).value = header
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    priority = {
        "매우비쌈": 1, "비쌈": 2, "주의": 3, "매칭확인": 4,
        "확인필요": 5, "자료부족": 6, "적정": 7, "저렴": 8, "매우저렴": 9,
    }
    sorted_rows = sorted(
        rows,
        key=lambda row: (
            priority.get(row.intelligence_grade, 99),
            -(row.impact_amount or 0),
            row.manufacturer,
            row.part_no,
        ),
    )

    for r_idx, row in enumerate(sorted_rows, start=2):
        values = [
            priority.get(row.intelligence_grade, 99),
            row.manufacturer, row.part_no, row.part_name_kr, row.model,
            row.order_qty, row.order_unit_price, row.master_latest_price, row.average_price,
            row.latest_deviation_rate, row.average_deviation_rate, row.impact_amount,
            row.intelligence_grade, row.intelligence_reason, row.recommendation,
            row.latest_request_no, row.price_change_count,
        ]
        for c_idx, value in enumerate(values, start=1):
            ws.cell(r_idx, c_idx).value = value

    grade_fill = {
        "매우비쌈": "F4CCCC", "비쌈": "FCE4D6", "주의": "FFF2CC",
        "적정": "E2F0D9", "저렴": "D9EAF7", "매우저렴": "DDEBF7",
        "자료부족": "E7E6E6", "확인필요": "FCE4D6", "매칭확인": "FCE4D6",
    }
    thin = Side(style="thin", color="D9E2F3")
    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row_idx, col)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="맑은 고딕", size=9)
        ws.cell(row_idx, 13).fill = PatternFill(
            "solid", fgColor=grade_fill.get(clean_text(ws.cell(row_idx, 13).value), "FFFFFF")
        )
        for col in (7, 8, 9, 12):
            ws.cell(row_idx, col).number_format = '#,##0.0000;[Red]-#,##0.0000'
        for col in (10, 11):
            ws.cell(row_idx, col).number_format = '0.00%;[Red]-0.00%'
            if ws.cell(row_idx, col).value is not None:
                ws.cell(row_idx, col).value = float(ws.cell(row_idx, col).value) / 100

    widths = {
        1: 10, 2: 9, 3: 19, 4: 30, 5: 12, 6: 11, 7: 12, 8: 14, 9: 14,
        10: 13, 11: 13, 12: 15, 13: 13, 14: 48, 15: 52, 16: 18, 17: 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"


def _negotiation_candidates(rows: list[PurchaseAnalysisRow]) -> list[PurchaseAnalysisRow]:
    accepted = {"매우비쌈", "비쌈", "주의"}
    return sorted(
        [row for row in rows if row.intelligence_grade in accepted],
        key=lambda row: (
            {"매우비쌈": 1, "비쌈": 2, "주의": 3}.get(row.intelligence_grade, 9),
            -(row.impact_amount or 0),
            row.manufacturer,
            row.part_no,
        ),
    )


def _negotiation_message(row: PurchaseAnalysisRow) -> str:
    latest = (
        f"{row.master_latest_price:,.4f} USD"
        if row.master_latest_price is not None else "확인 불가"
    )
    current = (
        f"{row.order_unit_price:,.4f} USD"
        if row.order_unit_price is not None else "확인 불가"
    )
    rate = (
        f"{row.latest_deviation_rate:+.1f}%"
        if row.latest_deviation_rate is not None else "계산 불가"
    )
    return (
        f"{row.part_name_kr or row.part_name_en or row.part_no}의 신규 발주단가는 "
        f"{current}이며, CPMS Master 최신단가 {latest} 대비 {rate}입니다. "
        f"{row.intelligence_reason} "
        "기존 가격 이력과 차이가 큰 만큼 적용 단가 및 인상 근거를 재확인해 주시기 바랍니다."
    )


def write_vendor_negotiation_report(
    result: PurchaseAnalysisResult,
    output_dir: str | Path,
) -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"공급사_단가협상자료_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    candidates = _negotiation_candidates(result.rows)
    wb = Workbook()
    summary = wb.active
    summary.title = "협상자료_요약"
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:H1")
    summary["A1"] = "CPMS 공급사 단가 협상자료"
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary["A1"].font = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 30

    summary.merge_cells("A2:H2")
    summary["A2"] = (
        f"원본: {result.source_path.name} / 작성일시: {datetime.now():%Y-%m-%d %H:%M:%S}"
    )
    summary["A2"].fill = PatternFill("solid", fgColor="D9EAF7")

    total_impact = sum(row.impact_amount or 0 for row in candidates)
    very_expensive = sum(row.intelligence_grade == "매우비쌈" for row in candidates)
    expensive = sum(row.intelligence_grade == "비쌈" for row in candidates)
    caution = sum(row.intelligence_grade == "주의" for row in candidates)

    kpis = [
        ("협상검토 품목", len(candidates)),
        ("매우비쌈", very_expensive),
        ("비쌈", expensive),
        ("주의", caution),
        ("예상 추가비용(USD)", total_impact),
        ("전체 발주금액(USD)", result.total_order_amount),
    ]
    for index, (label, value) in enumerate(kpis):
        row = 4 + (index // 3) * 2
        col = 1 + (index % 3) * 3
        summary.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        summary.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        summary.cell(row, col).value = label
        summary.cell(row, col).fill = PatternFill("solid", fgColor="5B9BD5")
        summary.cell(row, col).font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
        summary.cell(row, col).alignment = Alignment(horizontal="center")
        summary.cell(row + 1, col).value = value
        summary.cell(row + 1, col).font = Font(name="맑은 고딕", size=14, bold=True)
        summary.cell(row + 1, col).alignment = Alignment(horizontal="center")
        if "USD" in label:
            summary.cell(row + 1, col).number_format = '#,##0.00;[Red]-#,##0.00'

    summary["A10"] = "협상 권고 요약"
    summary["A10"].font = Font(name="맑은 고딕", size=12, bold=True)
    summary.merge_cells("A11:H13")
    if candidates:
        top = candidates[0]
        summary["A11"] = (
            f"우선 검토 대상은 {top.manufacturer} / "
            f"{top.part_name_kr or top.part_name_en or top.part_no}입니다. "
            f"발주 영향금액은 {(top.impact_amount or 0):,.2f} USD이며, "
            f"{top.intelligence_reason} "
            "가격 상승폭이 큰 품목부터 공급사 견적 및 인상 근거를 재확인하는 것을 권장합니다."
        )
    else:
        summary["A11"] = "현재 기준으로 협상 우선 검토 대상 품목이 없습니다."
    summary["A11"].alignment = Alignment(wrap_text=True, vertical="top")
    summary["A11"].fill = PatternFill("solid", fgColor="FFF2CC")

    headers = [
        "우선순위", "제조사", "부품코드", "부품명", "모델",
        "발주수량", "발주단가", "Master 최신단가", "직전 이력단가",
        "과거 평균단가", "과거 최고단가", "최신 대비(%)", "평균 대비(%)",
        "발주영향금액", "가격판정", "최근 요청 No.", "가격변경횟수",
        "단가이력 요약", "협상요청 문구",
    ]
    start_row = 15
    for col, header in enumerate(headers, start=1):
        summary.cell(start_row, col).value = header
    _style_header(summary, start_row, len(headers))
    summary.freeze_panes = f"A{start_row + 1}"

    for r_idx, row in enumerate(candidates, start=start_row + 1):
        values = [
            r_idx - start_row,
            row.manufacturer,
            row.part_no,
            row.part_name_kr or row.part_name_en,
            row.model,
            row.order_qty,
            row.order_unit_price,
            row.master_latest_price,
            row.previous_price,
            row.average_price,
            row.maximum_price,
            row.latest_deviation_rate,
            row.average_deviation_rate,
            row.impact_amount,
            row.intelligence_grade,
            row.latest_request_no,
            row.price_change_count,
            row.price_history_summary,
            _negotiation_message(row),
        ]
        for c_idx, value in enumerate(values, start=1):
            summary.cell(r_idx, c_idx).value = value

    grade_fill = {"매우비쌈": "F4CCCC", "비쌈": "FCE4D6", "주의": "FFF2CC"}
    thin = Side(style="thin", color="D9E2F3")
    for row_idx in range(start_row + 1, summary.max_row + 1):
        for col in range(1, len(headers) + 1):
            cell = summary.cell(row_idx, col)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="맑은 고딕", size=9)
        summary.cell(row_idx, 15).fill = PatternFill(
            "solid", fgColor=grade_fill.get(clean_text(summary.cell(row_idx, 15).value), "FFFFFF")
        )
        for col in (7, 8, 9, 10, 11, 14):
            summary.cell(row_idx, col).number_format = '#,##0.0000;[Red]-#,##0.0000'
        for col in (12, 13):
            summary.cell(row_idx, col).number_format = '0.00%;[Red]-0.00%'
            if summary.cell(row_idx, col).value is not None:
                summary.cell(row_idx, col).value = float(summary.cell(row_idx, col).value) / 100

    widths = {
        1: 10, 2: 9, 3: 19, 4: 30, 5: 12, 6: 11, 7: 12, 8: 14, 9: 14,
        10: 14, 11: 14, 12: 13, 13: 13, 14: 15, 15: 13, 16: 18, 17: 12,
        18: 42, 19: 70,
    }
    for col, width in widths.items():
        summary.column_dimensions[get_column_letter(col)].width = width
    summary.auto_filter.ref = (
        f"A{start_row}:{get_column_letter(len(headers))}{max(summary.max_row, start_row)}"
    )

    # Manufacturer-specific sheets serve as ready-to-send negotiation appendices.
    by_mfr: dict[str, list[PurchaseAnalysisRow]] = defaultdict(list)
    for row in candidates:
        by_mfr[row.manufacturer].append(row)

    for manufacturer, rows in sorted(by_mfr.items()):
        ws = wb.create_sheet(f"{manufacturer}_협상자료")
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:H1")
        ws["A1"] = f"{manufacturer} 공급사 단가 재협의 요청자료"
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        ws["A1"].font = Font(name="맑은 고딕", size=15, bold=True, color="FFFFFF")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:H7")
        impact = sum(row.impact_amount or 0 for row in rows)
        ws["A3"] = (
            f"금번 발주 검토 결과, {manufacturer} 관련 {len(rows)}개 품목에서 "
            f"CPMS Master 최신단가 대비 인상 또는 주의 수준의 차이가 확인되었습니다.\n"
            f"해당 품목의 예상 추가비용은 총 {impact:,.2f} USD입니다.\n"
            "아래 품목별 기존 단가 이력과 비교 내용을 검토하시어 "
            "적용 단가 및 인상 근거를 회신해 주시기 바랍니다."
        )
        ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        ws["A3"].fill = PatternFill("solid", fgColor="FFF2CC")

        vendor_headers = [
            "부품코드", "부품명", "모델", "발주수량", "발주단가",
            "Master 최신단가", "최신 대비(%)", "과거 평균단가",
            "과거 최고단가", "발주영향금액", "최근 요청 No.",
            "단가이력", "재협의 요청 내용",
        ]
        header_row = 9
        for col, header in enumerate(vendor_headers, start=1):
            ws.cell(header_row, col).value = header
        _style_header(ws, header_row, len(vendor_headers))
        ws.freeze_panes = f"A{header_row + 1}"

        for r_idx, row in enumerate(rows, start=header_row + 1):
            values = [
                row.part_no,
                row.part_name_kr or row.part_name_en,
                row.model,
                row.order_qty,
                row.order_unit_price,
                row.master_latest_price,
                row.latest_deviation_rate,
                row.average_price,
                row.maximum_price,
                row.impact_amount,
                row.latest_request_no,
                row.price_history_summary,
                _negotiation_message(row),
            ]
            for c_idx, value in enumerate(values, start=1):
                ws.cell(r_idx, c_idx).value = value

        for row_idx in range(header_row + 1, ws.max_row + 1):
            for col in range(1, len(vendor_headers) + 1):
                cell = ws.cell(row_idx, col)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.font = Font(name="맑은 고딕", size=9)
            for col in (5, 6, 8, 9, 10):
                ws.cell(row_idx, col).number_format = '#,##0.0000;[Red]-#,##0.0000'
            ws.cell(row_idx, 7).number_format = '0.00%;[Red]-0.00%'
            if ws.cell(row_idx, 7).value is not None:
                ws.cell(row_idx, 7).value = float(ws.cell(row_idx, 7).value) / 100

        vendor_widths = {
            1: 19, 2: 30, 3: 12, 4: 11, 5: 12, 6: 14, 7: 13,
            8: 14, 9: 14, 10: 15, 11: 18, 12: 42, 13: 72,
        }
        for col, width in vendor_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(vendor_headers))}{max(ws.max_row, header_row)}"
        )

    wb.save(output_path)
    result.negotiation_output_path = output_path
    return output_path

def write_purchase_analysis(result: PurchaseAnalysisResult, output_dir: str | Path) -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"발주단가_분석_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = Workbook()
    summary = wb.active
    summary.title = "발주분석_요약"
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:H1")
    summary["A1"] = "CPMS 발주 단가 변동 분석"
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary["A1"].font = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 30
    summary.merge_cells("A2:H2")
    summary["A2"] = f"원본: {result.source_path.name} / 분석일시: {datetime.now():%Y-%m-%d %H:%M:%S}"
    summary["A2"].fill = PatternFill("solid", fgColor="D9EAF7")

    kpis = [
        ("총 발주품목", result.total_items), ("총 발주수량", result.total_qty),
        ("발주금액(USD)", result.total_order_amount), ("Master 매칭", result.matched_items),
        ("가격상승", result.increased_items), ("가격하락", result.decreased_items),
        ("동일단가", result.unchanged_items), ("신규부품", result.new_items),
        ("확인필요", result.review_items), ("과거 대비 순영향(USD)", result.net_impact),
    ]
    for index, (label, value) in enumerate(kpis):
        row = 4 + index // 2 * 2
        col = 1 + (index % 2) * 4
        summary.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        summary.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        summary.cell(row, col).value = label
        summary.cell(row, col).fill = PatternFill("solid", fgColor="5B9BD5")
        summary.cell(row, col).font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
        summary.cell(row, col).alignment = Alignment(horizontal="center")
        summary.cell(row + 1, col).value = value
        summary.cell(row + 1, col).font = Font(name="맑은 고딕", size=14, bold=True)
        summary.cell(row + 1, col).alignment = Alignment(horizontal="center")
        if "USD" in label:
            summary.cell(row + 1, col).number_format = '#,##0.00;[Red]-#,##0.00'

    start_row = 16
    summary.cell(start_row, 1).value = "제조사별 요약"
    summary.cell(start_row, 1).font = Font(name="맑은 고딕", size=12, bold=True)
    headers = ["제조사", "품목", "수량", "발주금액", "상승", "하락", "신규", "순영향"]
    for col, header in enumerate(headers, start=1):
        summary.cell(start_row + 1, col).value = header
    _style_header(summary, start_row + 1, len(headers))

    by_mfr: dict[str, list[PurchaseAnalysisRow]] = defaultdict(list)
    for row in result.rows:
        by_mfr[row.manufacturer].append(row)
    for r_idx, manufacturer in enumerate(sorted(by_mfr), start=start_row + 2):
        rows = by_mfr[manufacturer]
        values = [
            manufacturer, len(rows), sum(r.order_qty for r in rows),
            sum(r.order_amount or 0 for r in rows),
            sum(r.status == "상승" for r in rows), sum(r.status == "하락" for r in rows),
            sum(r.status == "신규부품" for r in rows), sum(r.impact_amount or 0 for r in rows),
        ]
        for c_idx, value in enumerate(values, start=1):
            summary.cell(r_idx, c_idx).value = value
        summary.cell(r_idx, 4).number_format = '#,##0.00'
        summary.cell(r_idx, 8).number_format = '#,##0.00;[Red]-#,##0.00'

    if result.errors:
        error_row = start_row + 3 + len(by_mfr)
        summary.cell(error_row, 1).value = "확인사항"
        summary.cell(error_row, 1).font = Font(name="맑은 고딕", bold=True, color="C00000")
        for idx, error in enumerate(result.errors, start=error_row + 1):
            summary.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)
            summary.cell(idx, 1).value = error

    for col, width in enumerate([13, 13, 13, 16, 11, 11, 11, 16], start=1):
        summary.column_dimensions[get_column_letter(col)].width = width

    history_links = _write_price_history_sheet(wb, result.rows)
    _write_detail_sheet(wb, "전체_상세", result.rows, history_links)
    for manufacturer in sorted(by_mfr):
        _write_detail_sheet(wb, manufacturer, by_mfr[manufacturer], history_links)

    _write_price_intelligence_sheet(wb, result.rows)

    wb.save(output_path)
    result.output_path = output_path
    write_vendor_negotiation_report(result, output_dir)
    return output_path
