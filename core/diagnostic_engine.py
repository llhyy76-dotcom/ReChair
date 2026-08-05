from __future__ import annotations

import csv
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from core.compare import merge_records
from core.header_fingerprint import classify_file
from core.parser import (
    diagnose_workbook,
    infer_ir_no,
    infer_manufacturer,
    parse_master,
    parse_workbook,
)


EXCEL_SUFFIXES = {".xls", ".xlsx", ".xlsm", ".xltx", ".xltm"}
PDF_SUFFIXES = {".pdf"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".tif", ".tiff"}
INTERNAL_DOCUMENT_SUFFIXES = {".hwp", ".hwpx", ".doc", ".docx"}

EXCLUDE_NAME_PATTERNS = re.compile(
    r"(택배|지사.*발주|재고|사용량|고객만족팀|필요부품|품의)",
    re.I,
)
PRICE_NAME_PATTERNS = re.compile(
    r"(단가|PRICE|PI_|ORDER_|오더|회신|유상.*발주)",
    re.I,
)


@dataclass(frozen=True)
class DiagnosticItem:
    path: Path
    category: str
    action: str
    manufacturer: str
    request_no: str
    extracted_rows: int
    priced_rows: int
    confidence: int
    reason: str
    sheet_summary: str = ""

    @property
    def safe_to_exclude(self) -> bool:
        return self.category == "D" and self.action in {
            "단가 Master 반영 제외", "단가 미회신 요청서 보관"
        }


def is_unpriced_request_file(path: str | Path) -> tuple[bool, str]:
    """Return True when a request template has part rows but no supplier price.

    This is not a parser failure. It is an unanswered price-request form and
    must be kept out of the price Master while being preserved separately.
    """
    path = Path(path)
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return False, ""

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        for sheet in workbook.worksheets:
            header_row = 0
            part_cols: list[int] = []
            price_cols: list[int] = []
            qty_cols: list[int] = []
            amount_cols: list[int] = []

            for row_no, values in enumerate(
                sheet.iter_rows(
                    min_row=1,
                    max_row=min(sheet.max_row, 50),
                    max_col=min(sheet.max_column, 30),
                    values_only=True,
                ),
                start=1,
            ):
                normalized = [
                    re.sub(r"[\s\n\r\t()_\-./]+", "", str(value or "")).upper()
                    for value in values
                ]
                current_part = [
                    index + 1 for index, value in enumerate(normalized)
                    if any(token in value for token in ("부품명", "PARTNAME", "DESCRIPTION"))
                ]
                current_price = [
                    index + 1 for index, value in enumerate(normalized)
                    if any(token in value for token in ("단가", "UNITPRICE", "PRICEUSD", "PRICE"))
                ]
                current_qty = [
                    index + 1 for index, value in enumerate(normalized)
                    if any(token in value for token in ("요청수량", "QTYPCS", "QUANTITY", "QTY"))
                ]
                current_amount = [
                    index + 1 for index, value in enumerate(normalized)
                    if any(token in value for token in ("합계", "AMOUNT", "TOTAL"))
                ]
                if current_part and current_price:
                    header_row = row_no
                    part_cols = current_part
                    price_cols = current_price
                    qty_cols = current_qty
                    amount_cols = current_amount
                    break

            if not header_row:
                continue

            part_rows = 0
            priced_rows = 0
            qty_rows = 0
            amount_rows = 0
            for values in sheet.iter_rows(
                min_row=header_row + 1,
                max_row=sheet.max_row,
                max_col=min(sheet.max_column, 30),
                values_only=True,
            ):
                def value_at(col: int):
                    return values[col - 1] if col and col - 1 < len(values) else None

                parts = [
                    str(value_at(col) or "").strip()
                    for col in part_cols
                ]
                joined = " ".join(parts).strip()
                if not joined or joined.upper().startswith("TOTAL") or joined.startswith("합계"):
                    continue
                part_rows += 1

                for col in price_cols:
                    value = value_at(col)
                    try:
                        number = float(str(value).replace("$", "").replace(",", "").strip())
                    except (TypeError, ValueError):
                        number = None
                    if number is not None and number > 0:
                        priced_rows += 1
                        break

                for col in qty_cols:
                    value = value_at(col)
                    try:
                        number = float(str(value).replace(",", "").strip())
                    except (TypeError, ValueError):
                        number = None
                    if number is not None and number > 0:
                        qty_rows += 1
                        break

                for col in amount_cols:
                    value = value_at(col)
                    try:
                        number = float(str(value).replace("$", "").replace(",", "").strip())
                    except (TypeError, ValueError):
                        number = None
                    if number is not None and number > 0:
                        amount_rows += 1
                        break

            if part_rows > 0 and priced_rows == 0:
                return True, (
                    f"{sheet.title}: 부품 {part_rows}행이 있으나 "
                    f"단가 0행, 수량 {qty_rows}행, 금액 {amount_rows}행"
                )
    finally:
        workbook.close()
    return False, ""


def _quick_xlsx_profile(path: Path) -> dict[str, object]:
    """Inspect only the first 40 rows of each sheet for fast classification."""
    result = {
        "price_headers": 0,
        "qty_headers": 0,
        "part_headers": 0,
        "manufacturer_sheets": set(),
        "summary": [],
    }
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return result

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            sheet_name = str(sheet.title).strip().upper()
            if sheet_name in {"IR", "XC", "AC", "KA", "NC"}:
                result["manufacturer_sheets"].add(sheet_name)

            best = {"part": False, "qty": False, "price": False, "amount": False}
            best_row = 0
            for row_no, values in enumerate(
                sheet.iter_rows(
                    min_row=1,
                    max_row=min(sheet.max_row, 40),
                    max_col=min(sheet.max_column, 20),
                    values_only=True,
                ),
                start=1,
            ):
                joined = " | ".join(str(value or "") for value in values).upper()
                flags = {
                    "part": any(token in joined for token in (
                        "부품명", "PART NAME", "DESCRIPTION", "품명",
                    )),
                    "qty": any(token in joined for token in (
                        "요청수량", "수량", "QTY", "QUANTITY", "Q'TY",
                    )),
                    "price": any(token in joined for token in (
                        "단가", "UNIT PRICE", "PRICE(USD)", "PRICE",
                    )),
                    "amount": any(token in joined for token in (
                        "합계", "금액", "AMOUNT", "TOTAL",
                    )),
                }
                if sum(flags.values()) > sum(best.values()):
                    best, best_row = flags, row_no
            result["part_headers"] += int(best["part"])
            result["qty_headers"] += int(best["qty"])
            result["price_headers"] += int(best["price"])
            result["summary"].append(
                f"{sheet.title}: 헤더 {best_row}행 "
                f"(부품={best['part']}, 수량={best['qty']}, "
                f"단가={best['price']}, 금액={best['amount']})"
            )
    finally:
        workbook.close()
    return result


def _sheet_summary(diagnosis: dict[str, object]) -> str:
    parts: list[str] = []
    for sheet in diagnosis.get("sheets", []):
        status = str(sheet.get("status", ""))
        parser = str(sheet.get("parser", ""))
        extracted = int(sheet.get("extracted_rows", 0) or 0)
        parts.append(f"{sheet.get('sheet')}: {parser}, {status}, {extracted}행")
    return " / ".join(parts)


def diagnose_file(path: str | Path) -> DiagnosticItem:
    path = Path(path)
    suffix = path.suffix.lower()
    filename_manufacturer = infer_manufacturer(path)
    classification = None
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        try:
            classification = classify_file(path, filename_manufacturer)
        except Exception:
            classification = None
    manufacturer = (
        str(classification.get("manufacturer"))
        if classification and classification.get("manufacturer")
        else (filename_manufacturer or "UNKNOWN")
    )
    request_no = infer_ir_no(path)

    if suffix in PDF_SUFFIXES:
        return DiagnosticItem(
            path, "C", "PDF/OCR 검토", manufacturer, request_no,
            0, 0, 70,
            "PDF는 OCR 또는 내장 표 추출 후 사람의 검토·승인이 필요합니다.",
        )

    if suffix in IMAGE_SUFFIXES:
        return DiagnosticItem(
            path, "C", "이미지 OCR 검토", manufacturer, request_no,
            0, 0, 60,
            "이미지 주문서는 OCR 검토파일로 변환한 뒤 승인해야 합니다.",
        )

    if suffix in INTERNAL_DOCUMENT_SUFFIXES:
        return DiagnosticItem(
            path, "D", "단가 Master 반영 제외", manufacturer, request_no,
            0, 0, 95,
            "HWP/문서형 구매품의는 내부 결재자료입니다. 연결된 Excel·PI·ORDER를 단가 근거로 사용합니다.",
        )

    if suffix not in EXCEL_SUFFIXES:
        return DiagnosticItem(
            path, "D", "지원하지 않는 형식", manufacturer, request_no,
            0, 0, 100, f"지원하지 않는 확장자: {suffix}",
        )

    unpriced, unpriced_reason = is_unpriced_request_file(path)
    if unpriced:
        return DiagnosticItem(
            path, "D", "단가 미회신 요청서 보관", manufacturer, request_no,
            0, 0, 99,
            "공급사 단가가 입력되지 않은 요청서입니다. "
            + unpriced_reason,
        )

    # Fast preflight avoids fully parsing huge inventory/shipping workbooks.
    profile = _quick_xlsx_profile(path)
    manufacturer_sheets = set(profile.get("manufacturer_sheets", set()))
    quick_summary = " / ".join(profile.get("summary", []))

    if len(manufacturer_sheets) > 1:
        return DiagnosticItem(
            path, "B", "제조사별 시트 분리 필요", manufacturer, request_no,
            0, 0, 65,
            "한 파일에 여러 제조사 시트가 있습니다. 각 시트를 제조사별 요청으로 분리해야 안전합니다.",
            quick_summary,
        )

    if (
        int(profile.get("part_headers", 0)) > 0
        and int(profile.get("qty_headers", 0)) > 0
        and int(profile.get("price_headers", 0)) == 0
        and EXCLUDE_NAME_PATTERNS.search(path.stem)
        and not PRICE_NAME_PATTERNS.search(path.stem)
    ):
        return DiagnosticItem(
            path, "D", "단가 Master 반영 제외", manufacturer, request_no,
            0, 0, 98,
            "부품·수량 헤더는 있으나 단가 열이 없는 재고·택배·지사 요청 자료입니다.",
            quick_summary,
        )

    if suffix == ".xls":
        if EXCLUDE_NAME_PATTERNS.search(path.stem) and not PRICE_NAME_PATTERNS.search(path.stem):
            return DiagnosticItem(
                path, "D", "단가 Master 반영 제외", manufacturer, request_no,
                0, 0, 95,
                "구형 XLS이며 단가 근거가 아닌 택배·재고·지사 요청 자료입니다.",
            )
        return DiagnosticItem(
            path, "B", "구형 XLS 변환/전용 파서", manufacturer, request_no,
            0, 0, 50,
            "구형 XLS 파일은 진단 중 자동 변환하지 않습니다. 가격자료 여부 확인 후 XLSX로 변환합니다.",
        )

    try:
        diagnosis = diagnose_workbook(path)
        rows = int(diagnosis.get("total_rows", 0) or 0)
        priced = int(diagnosis.get("priced_rows", 0) or 0)
        summary = _sheet_summary(diagnosis)
    except Exception as exc:
        return DiagnosticItem(
            path, "B", "전용 파서 수정 필요", manufacturer, request_no,
            0, 0, 35,
            f"Excel 열기 또는 파싱 실패: {type(exc).__name__}: {exc}",
        )

    if rows > 0 and priced > 0:
        confidence = 99 if priced == rows else max(82, round(priced / rows * 100))
        fingerprint_text = ""
        if classification:
            fingerprint_text = (
                f" / 양식 {classification.get('parser_type')} "
                f"{classification.get('parser_confidence')}% / "
                f"제조사판정 {classification.get('manufacturer_confidence')}%"
            )
        return DiagnosticItem(
            path, "A", "자동 반영 가능", manufacturer, request_no,
            rows, priced, confidence,
            f"단가 포함 품목 {priced:,}/{rows:,}행을 추출했습니다."
            + fingerprint_text,
            summary,
        )

    if rows > 0 and priced == 0:
        # Stock/usage/branch request sheets must not pollute price history.
        if EXCLUDE_NAME_PATTERNS.search(path.stem) and not PRICE_NAME_PATTERNS.search(path.stem):
            return DiagnosticItem(
                path, "D", "단가 Master 반영 제외", manufacturer, request_no,
                rows, 0, 98,
                "수량·재고·사용량 자료이며 단가 근거가 없습니다.",
                summary,
            )
        return DiagnosticItem(
            path, "B", "수량이력 전용 또는 수동검토", manufacturer, request_no,
            rows, 0, 55,
            "부품·수량은 읽었지만 단가가 없습니다. 가격 Master에는 자동 반영하지 않습니다.",
            summary,
        )

    if EXCLUDE_NAME_PATTERNS.search(path.stem) and not PRICE_NAME_PATTERNS.search(path.stem):
        return DiagnosticItem(
            path, "D", "단가 Master 반영 제외", manufacturer, request_no,
            0, 0, 95,
            "파일명과 표 구조상 지사·재고·택배·사용량 자료로 판단됩니다.",
            summary,
        )

    return DiagnosticItem(
        path, "B", "전용 파서 수정 필요", manufacturer, request_no,
        0, 0, 30,
        "핵심 단가 행을 추출하지 못했습니다.",
        summary,
    )


def diagnose_folder(folder: str | Path) -> list[DiagnosticItem]:
    folder = Path(folder)
    if not folder.exists():
        return []
    return [
        diagnose_file(path)
        for path in sorted(folder.iterdir())
        if path.is_file() and not path.name.startswith("~$") and path.name != ".keep"
    ]


def write_diagnostic_csv(
    items: list[DiagnosticItem],
    output_dir: str | Path,
) -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"file_diagnostic_{datetime.now():%Y%m%d_%H%M%S}.csv"
    with path.open("w", newline="", encoding="utf-8-sig") as fp:
        writer = csv.writer(fp)
        writer.writerow([
            "등급", "조치", "파일명", "제조사", "요청번호",
            "추출행", "단가행", "신뢰도", "판정사유", "시트진단", "원본경로",
        ])
        for item in items:
            writer.writerow([
                item.category, item.action, item.path.name,
                item.manufacturer, item.request_no,
                item.extracted_rows, item.priced_rows,
                item.confidence, item.reason,
                item.sheet_summary, str(item.path),
            ])
    return path


def quarantine_excluded(
    items: list[DiagnosticItem],
    excluded_root: str | Path,
) -> tuple[list[Path], list[str]]:
    excluded_root = Path(excluded_root)
    target_dir = excluded_root / datetime.now().strftime("%Y%m%d")
    target_dir.mkdir(parents=True, exist_ok=True)

    moved: list[Path] = []
    errors: list[str] = []
    for item in items:
        if not item.safe_to_exclude or not item.path.exists():
            continue
        target = target_dir / item.path.name
        counter = 1
        while target.exists():
            target = target_dir / f"{item.path.stem}_{counter}{item.path.suffix}"
            counter += 1
        try:
            shutil.move(str(item.path), str(target))
            moved.append(target)
        except Exception as exc:
            errors.append(f"{item.path.name}: {exc}")
    return moved, errors


@dataclass(frozen=True)
class SimulationResult:
    manufacturer: str
    files: int
    parsed_rows: int
    expected_new_parts: int
    expected_price_changes: int
    expected_duplicates: int
    note: str = ""


def simulate_update(root: str | Path, items: list[DiagnosticItem] | None = None) -> list[SimulationResult]:
    """Dry-run A-grade files without changing Master, History or Archive."""
    root = Path(root)
    items = items if items is not None else diagnose_folder(root / "Update")
    grouped: dict[str, list[Path]] = {}
    for item in items:
        if item.category != "A" or item.manufacturer not in {"IR", "XC", "AC", "KA", "NC"}:
            continue
        grouped.setdefault(item.manufacturer, []).append(item.path)

    results: list[SimulationResult] = []
    for manufacturer in ("IR", "XC", "AC", "KA", "NC"):
        paths = grouped.get(manufacturer, [])
        if not paths:
            continue
        incoming = []
        errors: list[str] = []
        for path in paths:
            try:
                incoming.extend(parse_workbook(path))
            except Exception as exc:
                errors.append(f"{path.name}: {exc}")

        master_path = root / "Manufacturers" / manufacturer / "Master" / "Master.xlsx"
        existing = parse_master(master_path) if master_path.exists() else []
        _all, new_records, price_changes, duplicates, _updated = merge_records(
            existing, incoming
        )
        results.append(
            SimulationResult(
                manufacturer=manufacturer,
                files=len(paths),
                parsed_rows=len(incoming),
                expected_new_parts=len(new_records),
                expected_price_changes=len(price_changes),
                expected_duplicates=duplicates,
                note=" / ".join(errors),
            )
        )
    return results
