from __future__ import annotations

from collections import defaultdict
from copy import copy
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.parser import PartRecord, clean_text, norm_model, norm_txt, parse_master

try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    RICH_TEXT_AVAILABLE = True
except Exception:
    RICH_TEXT_AVAILABLE = False

PRICE_COLORS = ["0070C0", "C00000", "7030A0", "00A050", "ED7D31", "0000FF", "FF0000", "806000"]
HEADER_FILL = "1F4E79"; LIGHT_BLUE = "DDEBF7"; TITLE_FILL = "FFF2CC"; WHITE = "FFFFFF"; GRID = "D9E2F3"

DETAIL_HEADERS = [
    "기준년도", "기준일자", "요청 No.", "모델", "부품명(한글)", "부품명(영어)", "부품명(중국어)",
    "단가(USD)", "수량", "금액(USD)", "품번/코드", "파라미터/비고", "원본파일", "원본행", "추출방식"
]
BASE_PRODUCT_HEADERS = ["적용모델", "부품명(한글)", "부품명(영어)", "부품명(중국어)", "품번/코드", "파라미터/비고"]
TAIL_PRODUCT_HEADERS = ["수량합계", "금액합계(USD)", "요청 No.", "기준일자", "원본파일", "비고"]
CHANGE_HEADERS = ["적용모델", "부품명(한글)", "부품명(영어)", "변동구분/년도", "단가목록(USD)", "요청 No.", "기준일자", "원본파일"]
COMPARE_HEADERS = ["부품명(한글)", "부품명(영어)", "부품명(중국어)", "적용모델 수", "최저 USD", "최고 USD", "차이", "단가 동일여부", "최근 기준년도", "최근년도 모델별 단가", "전체 적용모델"]
MODEL_HEADERS = ["적용모델", "부품명 수", "상세 행수", "반영 기준년도", "요청 건수", "수량합계", "금액합계(USD)", "최저 USD", "최고 USD", "평균 USD"]
FILE_HEADERS = ["원본파일", "요청 No.", "적용모델", "시트명", "기준일자", "기준년도", "추출 행수", "상태"]
HISTORY_HEADERS = ["적용모델", "부품명(한글)", "부품명(영어)", "기준년도", "단가(USD)", "요청 No.", "기준일자", "원본파일", "수량합계", "금액합계(USD)"]
VALIDATION_HEADERS = ["검증항목", "값", "비고"]
DASHBOARD_TOP_HEADERS = ["순위", "적용모델", "부품명(한글)", "부품명(영어)", "이전단가", "최근단가", "변동액", "변동률", "최근 요청 No."]


def _num(value: float | None) -> float | None:
    return None if value is None else round(float(value), 4)


def _price_text(value: float | None) -> str:
    if value is None: return ""
    return f"{float(value):.2f}".rstrip("0").rstrip(".")


def _unique_sorted(values: Iterable[str]) -> list[str]:
    out, seen = [], set()
    for value in values:
        v = clean_text(value)
        if v and v not in seen:
            seen.add(v); out.append(v)
    return out


def _set_cell_rich_text(cell, chunks: list[tuple[str, str | None]], default_color: str = "000000") -> None:
    if not chunks:
        cell.value = ""; return
    if not RICH_TEXT_AVAILABLE:
        cell.value = "".join(t for t, _ in chunks)
        colors = {c for _, c in chunks if c}
        if len(colors) == 1:
            cell.font = Font(name="맑은 고딕", size=10, color=next(iter(colors)))
        return
    rt = CellRichText()
    for text, color in chunks:
        if text:
            rt.append(TextBlock(InlineFont(color=color or default_color), text))
    cell.value = rt


def _date_short(s: str) -> str:
    s = clean_text(s)
    return s[5:7] + "/" + s[8:10] if len(s) >= 10 and s[4] == "-" else s


def _record_sort_key(r: PartRecord):
    return (r.base_date or "", r.ir_no or "", r.source_row or 0)


def _group_key(r: PartRecord):
    return (norm_model(r.model), norm_txt(r.part_name_en), norm_txt(r.part_name_kr), norm_txt(r.part_no), norm_txt(r.parameter))


def _group_key_simple(r: PartRecord):
    return (norm_txt(r.part_name_en), norm_txt(r.part_name_kr), norm_txt(r.part_no), norm_txt(r.parameter))


def _latest_nonblank(records: list[PartRecord], attr: str) -> str:
    for r in sorted(records, key=_record_sort_key, reverse=True):
        v = clean_text(getattr(r, attr))
        if v: return v
    return ""


def years_from_records(records: list[PartRecord]) -> list[int]:
    years = sorted({int(r.base_year) for r in records if r.base_year})
    return years or []


def _style_header_row(ws, row: int, max_col: int) -> None:
    fill = PatternFill("solid", fgColor=HEADER_FILL); font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
    border = Border(bottom=Side(style="thin", color="BFBFBF"))
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c); cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = border


def _style_title(ws, title: str, subtitle: str | None, max_col: int) -> None:
    ws.cell(1, 1).value = title; ws.cell(1, 1).font = Font(name="맑은 고딕", size=15, bold=True, color=WHITE)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col); ws.row_dimensions[1].height = 25
    if subtitle is not None:
        ws.cell(2, 1).value = subtitle; ws.cell(2, 1).font = Font(name="맑은 고딕", size=10, bold=True, color="7F6000")
        ws.cell(2, 1).fill = PatternFill("solid", fgColor=TITLE_FILL); ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col); ws.row_dimensions[2].height = 35


def _finish_sheet(ws, header_row: int, max_col: int, freeze: str = "A5") -> None:
    widths = {1: 13, 2: 26, 3: 40, 4: 22, 5: 16, 6: 28, 7: 15, 8: 15, 9: 15, 10: 12, 11: 15, 12: 58, 13: 25, 14: 55, 15: 18}
    for c in range(1, max_col + 1): ws.column_dimensions[get_column_letter(c)].width = widths.get(c, 18)
    ws.freeze_panes = freeze; ws.sheet_view.showGridLines = False
    last = max(ws.max_row, header_row); thin = Side(style="thin", color=GRID)
    for row in ws.iter_rows(min_row=header_row, max_row=last, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row > header_row:
                # Do not overwrite rich-text colors.
                if not (RICH_TEXT_AVAILABLE and isinstance(cell.value, CellRichText)):
                    cell.font = Font(name="맑은 고딕", size=10, color="000000")
        if row[0].row > header_row and row[0].row % 2 == 0:
            for cell in row: cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{last}"


def _reset_sheet(wb, name: str, index: int | None = None):
    if name in wb.sheetnames: wb.remove(wb[name])
    return wb.create_sheet(name, index) if index is not None else wb.create_sheet(name)


def write_original_detail(wb, records: list[PartRecord]) -> None:
    ws = _reset_sheet(wb, "원본상세", 4)
    for c, h in enumerate(DETAIL_HEADERS, 1): ws.cell(1, c).value = h
    _style_header_row(ws, 1, len(DETAIL_HEADERS))
    for r_idx, r in enumerate(sorted(records, key=lambda x: (x.base_year or 0, x.base_date or "", x.ir_no, x.source_row or 0)), 2):
        vals = [r.base_year, r.base_date, r.ir_no, r.model, r.part_name_kr, r.part_name_en, r.part_name_cn, r.price, r.qty, r.amount, r.part_no, r.parameter, r.source_file, r.source_row, r.extraction_method]
        for c, v in enumerate(vals, 1): ws.cell(r_idx, c).value = v
    _finish_sheet(ws, 1, len(DETAIL_HEADERS), "A2")
    for c in [8, 10]:
        for row in range(2, ws.max_row + 1): ws.cell(row, c).number_format = '#,##0.00'
    for c in [9, 14]:
        for row in range(2, ws.max_row + 1): ws.cell(row, c).number_format = '#,##0'


def build_product_sheet(wb, records: list[PartRecord], manufacturer: str = "") -> list[dict]:
    years = years_from_records(records)
    headers = BASE_PRODUCT_HEADERS + [f"{y}년 단가(USD)" for y in years] + TAIL_PRODUCT_HEADERS
    ws = _reset_sheet(wb, "제품별_단가표", 0)
    title_prefix = f"{manufacturer} " if manufacturer else ""
    _style_title(ws, f"{title_prefix}Request 기준 제품별 부품 단가표", "원본상세 전체 기준입니다. 단가 변동 시 단가와 요청 No.를 같은 글자색으로 매칭합니다. 기준년도는 원본상세 기준으로 자동 생성됩니다.", len(headers))
    for c, h in enumerate(headers, 1): ws.cell(4, c).value = h
    _style_header_row(ws, 4, len(headers))
    groups: dict[tuple, list[PartRecord]] = defaultdict(list)
    for r in records: groups[_group_key(r)].append(r)
    change_rows: list[dict] = []
    row = 5; year_start = len(BASE_PRODUCT_HEADERS) + 1; tail_start = year_start + len(years)
    for key, recs in sorted(groups.items(), key=lambda kv: (kv[0][0], kv[0][1], kv[0][2])):
        recs = sorted(recs, key=_record_sort_key)
        model = _latest_nonblank(recs, "model") or key[0]; kr = _latest_nonblank(recs, "part_name_kr"); en = _latest_nonblank(recs, "part_name_en"); cn = _latest_nonblank(recs, "part_name_cn")
        part_no = _latest_nonblank(recs, "part_no"); param = _latest_nonblank(recs, "parameter")
        unique_prices_all = []
        for r in recs:
            p = _num(r.price)
            if p is not None and p not in unique_prices_all: unique_prices_all.append(p)
        has_change = len(set(unique_prices_all)) > 1
        color_by_price = {p: PRICE_COLORS[i % len(PRICE_COLORS)] for i, p in enumerate(unique_prices_all)} if has_change else {}
        for c, v in enumerate([model, kr, en, cn, part_no, param], 1): ws.cell(row, c).value = v
        for idx, year in enumerate(years, 0):
            yr_recs = [r for r in recs if r.base_year == year and r.price is not None]
            price_order = []
            for r in yr_recs:
                p = _num(r.price)
                if p not in price_order: price_order.append(p)
            chunks = []
            for i, p in enumerate(price_order):
                if i: chunks.append((" / ", None))
                chunks.append((_price_text(p), color_by_price.get(p)))
            _set_cell_rich_text(ws.cell(row, year_start + idx), chunks) if chunks else None
            if (len(set(price_order)) > 1 or has_change) and yr_recs:
                change_rows.append({"model": model, "kr": kr, "en": en, "year": year, "prices": price_order, "records": yr_recs, "color_by_price": color_by_price})
        ws.cell(row, tail_start).value = sum(r.qty or 0 for r in recs)
        ws.cell(row, tail_start + 1).value = sum(r.amount if r.amount is not None else ((r.price or 0) * (r.qty or 0)) for r in recs)
        if has_change:
            chunks = []; first = True
            for p in unique_prices_all:
                prs = [r for r in recs if _num(r.price) == p]
                if not prs: continue
                if not first: chunks.append((" / ", None))
                first = False
                refs = ", ".join(_unique_sorted(f"{r.ir_no}({_date_short(r.base_date)})" for r in prs))
                chunks.append((f"{_price_text(p)}: {refs}", color_by_price.get(p)))
            _set_cell_rich_text(ws.cell(row, tail_start + 2), chunks)
        else:
            ws.cell(row, tail_start + 2).value = ", ".join(_unique_sorted(r.ir_no for r in recs))
        ws.cell(row, tail_start + 3).value = ", ".join(_unique_sorted(r.base_date for r in recs))
        ws.cell(row, tail_start + 4).value = ", ".join(_unique_sorted(r.source_file for r in recs))
        ws.cell(row, tail_start + 5).value = "가격변동" if has_change else ""
        row += 1
    _finish_sheet(ws, 4, len(headers), "A5")
    for c in list(range(year_start, year_start + len(years))) + [tail_start + 1]:
        for r in range(5, ws.max_row + 1): ws.cell(r, c).number_format = '#,##0.00'
    return change_rows


def build_change_sheet(wb, change_rows: list[dict]) -> None:
    ws = _reset_sheet(wb, "가격변동_확인", 1)
    _style_title(ws, "가격변동 확인", "제품별_단가표에서 글자색으로 매칭 표시된 가격변동 단가와 요청 No.의 근거 목록입니다.", len(CHANGE_HEADERS))
    for c, h in enumerate(CHANGE_HEADERS, 1): ws.cell(4, c).value = h
    _style_header_row(ws, 4, len(CHANGE_HEADERS)); row = 5; seen = set()
    for item in change_rows:
        prices, recs = item["prices"], item["records"]
        if not prices or not recs: continue
        key = (item["model"], item["en"], item["year"], tuple(prices))
        if key in seen: continue
        seen.add(key)
        for c, v in enumerate([item["model"], item["kr"], item["en"], item["year"]], 1): ws.cell(row, c).value = v
        p_chunks = []
        for i, p in enumerate(prices):
            if i: p_chunks.append((" / ", None))
            p_chunks.append((_price_text(p), item["color_by_price"].get(p)))
        _set_cell_rich_text(ws.cell(row, 5), p_chunks)
        ir_chunks = []
        for i, p in enumerate(prices):
            if i: ir_chunks.append((" / ", None))
            prs = [r for r in recs if _num(r.price) == p]
            refs = ", ".join(_unique_sorted(f"{r.ir_no}({_date_short(r.base_date)})" for r in prs))
            ir_chunks.append((f"{_price_text(p)}: {refs}", item["color_by_price"].get(p)))
        _set_cell_rich_text(ws.cell(row, 6), ir_chunks)
        ws.cell(row, 7).value = ", ".join(_unique_sorted(r.base_date for r in recs))
        ws.cell(row, 8).value = ", ".join(_unique_sorted(r.source_file for r in recs))
        row += 1
    _finish_sheet(ws, 4, len(CHANGE_HEADERS), "A5"); ws.column_dimensions['F'].width = 65; ws.column_dimensions['H'].width = 70


def build_history_sheet(wb, records: list[PartRecord]) -> None:
    ws = _reset_sheet(wb, "가격_History", 2)
    _style_title(ws, "가격 History", "원본상세 기준 모든 연도/단가 이력을 보관합니다.", len(HISTORY_HEADERS))
    for c, h in enumerate(HISTORY_HEADERS, 1): ws.cell(4, c).value = h
    _style_header_row(ws, 4, len(HISTORY_HEADERS)); row = 5
    groups: dict[tuple, list[PartRecord]] = defaultdict(list)
    for r in records:
        if r.price is not None: groups[(_group_key(r), r.base_year, _num(r.price))].append(r)
    for (_, year, price), recs in sorted(groups.items(), key=lambda kv: (kv[0][0], kv[0][1] or 0, kv[0][2] or 0)):
        vals = [_latest_nonblank(recs,"model"), _latest_nonblank(recs,"part_name_kr"), _latest_nonblank(recs,"part_name_en"), year, _price_text(price), ", ".join(_unique_sorted(r.ir_no for r in recs)), ", ".join(_unique_sorted(r.base_date for r in recs)), ", ".join(_unique_sorted(r.source_file for r in recs)), sum(r.qty or 0 for r in recs), sum(r.amount if r.amount is not None else ((r.price or 0)*(r.qty or 0)) for r in recs)]
        for c, v in enumerate(vals, 1): ws.cell(row, c).value = v
        row += 1
    _finish_sheet(ws, 4, len(HISTORY_HEADERS), "A5")


def build_compare_sheet(wb, records: list[PartRecord]) -> None:
    ws = _reset_sheet(wb, "부품별_가격비교", 3)
    _style_title(ws, "부품별 가격 비교", "동일 부품명 기준으로 모델별 단가 차이를 확인하는 비교용 시트입니다.", len(COMPARE_HEADERS))
    for c, h in enumerate(COMPARE_HEADERS, 1): ws.cell(4, c).value = h
    _style_header_row(ws, 4, len(COMPARE_HEADERS)); row = 5
    groups: dict[tuple, list[PartRecord]] = defaultdict(list)
    for r in records: groups[(norm_txt(r.part_name_en), norm_txt(r.part_name_kr), norm_txt(r.part_name_cn))].append(r)
    for key, recs in sorted(groups.items(), key=lambda kv: kv[0][0]):
        prices = [r.price for r in recs if r.price is not None]
        if not prices: continue
        models = sorted(set(norm_model(r.model) for r in recs if r.model)); latest_year = max([r.base_year for r in recs if r.base_year] or [None])
        latest_recs = [r for r in recs if r.base_year == latest_year] if latest_year else recs
        latest_by_model = {}
        for r in sorted(latest_recs, key=_record_sort_key):
            if r.price is not None: latest_by_model[norm_model(r.model)] = _price_text(r.price)
        min_p, max_p = min(prices), max(prices)
        vals = [_latest_nonblank(recs,"part_name_kr"), _latest_nonblank(recs,"part_name_en"), _latest_nonblank(recs,"part_name_cn"), len(models), _price_text(min_p), _price_text(max_p), _price_text(max_p-min_p), "동일" if round(max_p-min_p,6)==0 else "상이", latest_year, ", ".join(f"{m}: {p}" for m,p in sorted(latest_by_model.items())), ", ".join(models)]
        for c, v in enumerate(vals, 1): ws.cell(row, c).value = v
        row += 1
    _finish_sheet(ws, 4, len(COMPARE_HEADERS), "A5")


def build_model_summary(wb, records: list[PartRecord]) -> None:
    ws = _reset_sheet(wb, "모델별_요약", 5)
    _style_title(ws, "모델별 요약", None, len(MODEL_HEADERS))
    for c, h in enumerate(MODEL_HEADERS, 1): ws.cell(3, c).value = h
    _style_header_row(ws, 3, len(MODEL_HEADERS)); row = 4
    groups: dict[str, list[PartRecord]] = defaultdict(list)
    for r in records: groups[norm_model(r.model)].append(r)
    for model, recs in sorted(groups.items()):
        prices = [r.price for r in recs if r.price is not None]
        vals = [model, len(set(_group_key_simple(r) for r in recs)), len(recs), ", ".join(str(y) for y in sorted(set(r.base_year for r in recs if r.base_year))), len(set(r.ir_no for r in recs if r.ir_no)), sum(r.qty or 0 for r in recs), sum(r.amount if r.amount is not None else ((r.price or 0)*(r.qty or 0)) for r in recs), _price_text(min(prices)) if prices else "", _price_text(max(prices)) if prices else "", _price_text(sum(prices)/len(prices)) if prices else ""]
        for c, v in enumerate(vals, 1): ws.cell(row, c).value = v
        row += 1
    _finish_sheet(ws, 3, len(MODEL_HEADERS), "A4")


def build_file_year_sheet(wb, records: list[PartRecord]) -> None:
    ws = _reset_sheet(wb, "파일_기준년도", 6)
    for c, h in enumerate(FILE_HEADERS, 1): ws.cell(1, c).value = h
    _style_header_row(ws, 1, len(FILE_HEADERS)); row = 2
    groups: dict[str, list[PartRecord]] = defaultdict(list)
    for r in records: groups[r.source_file or "(원본파일 없음)"].append(r)
    for sf, recs in sorted(groups.items()):
        vals = [sf, ", ".join(_unique_sorted(r.ir_no for r in recs)), ", ".join(_unique_sorted(norm_model(r.model) for r in recs)), "Request", ", ".join(_unique_sorted(r.base_date for r in recs)), ", ".join(str(y) for y in sorted(set(r.base_year for r in recs if r.base_year))), len(recs), "정상" if all(r.base_year for r in recs) and sf != "(원본파일 없음)" else "확인필요"]
        for c, v in enumerate(vals, 1): ws.cell(row, c).value = v
        row += 1
    _finish_sheet(ws, 1, len(FILE_HEADERS), "A2"); ws.column_dimensions['A'].width = 70


def build_work_rules(wb) -> None:
    ws = _reset_sheet(wb, "작업기준", 7)
    rows = [
        ["가격변동 표시 기준", ""], ["제조사", "IR/XC/AC/KA/NC 모두 같은 공통 엔진을 사용합니다."],
        ["기준년도", "원본상세 기준일자/요청번호/원본파일 날짜에서 자동 산출합니다. 2020년 이하와 2027년 이후도 자동 확장됩니다."],
        ["적용모델", "부품명(영어)의 마지막 모델값을 우선 적용합니다."],
        ["단가 색상", "동일 부품에서 단가가 2개 이상이면 단가와 요청 No.를 같은 글자색으로 매칭합니다."],
        ["업데이트", "공통 Update 폴더의 모든 xlsx를 파일명 접두어(IR/XC/AC/KA/NC)로 자동 분류합니다."],
        ["제한", "파일 개수 제한은 없습니다. 오류 파일은 건너뛰고 로그에 남깁니다."],
    ]
    for r, vals in enumerate(rows, 1):
        for c, v in enumerate(vals, 1): ws.cell(r, c).value = v
    _style_header_row(ws, 1, 2); _finish_sheet(ws, 1, 2, "A2"); ws.column_dimensions['B'].width = 110


def build_validation_sheet(wb, records: list[PartRecord], appended_count: int, selected_count: int, skipped_over_limit: int = 0) -> None:
    ws = _reset_sheet(wb, "업데이트_검증", 8)
    for c, h in enumerate(VALIDATION_HEADERS, 1): ws.cell(1, c).value = h
    _style_header_row(ws, 1, len(VALIDATION_HEADERS))
    by_key: dict[tuple, set] = defaultdict(set)
    for r in records:
        if r.price is not None: by_key[_group_key(r)].add(_num(r.price))
    rows = [
        ["처리 파일 수", selected_count, "이번 실행에서 처리된 파일 수"], ["신규 원본상세 추가 행", appended_count, "중복 raw row 제외"],
        ["원본상세 총 행", len(records), "헤더 제외"], ["제품별 단가표 행", len(set(_group_key(r) for r in records)), "동일 모델/부품/품번/비고 기준"],
        ["가격변동 부품 수", sum(1 for s in by_key.values() if len(s) > 1), "동일 부품 내 단가 2개 이상"],
        ["기준년도 누락 행", sum(1 for r in records if not r.base_year), "0이어야 정상"], ["기준일자 누락 행", sum(1 for r in records if not r.base_date), "필요 시 원본파일명 날짜 확인"],
    ]
    for r, vals in enumerate(rows, 2):
        for c, v in enumerate(vals, 1): ws.cell(r, c).value = v
    _finish_sheet(ws, 1, len(VALIDATION_HEADERS), "A2"); ws.column_dimensions['A'].width = 24; ws.column_dimensions['C'].width = 70



def _price_change_summary(records: list[PartRecord]) -> list[dict]:
    groups: dict[tuple, list[PartRecord]] = defaultdict(list)
    for record in records:
        if record.price is not None:
            groups[_group_key(record)].append(record)

    summaries: list[dict] = []
    for recs in groups.values():
        ordered = sorted(recs, key=_record_sort_key)
        distinct: list[tuple[float, PartRecord]] = []
        for record in ordered:
            price = _num(record.price)
            if price is None:
                continue
            if not distinct or distinct[-1][0] != price:
                distinct.append((price, record))
        if len(distinct) < 2:
            continue

        previous_price, _ = distinct[-2]
        latest_price, latest_record = distinct[-1]
        rate = None if previous_price == 0 else ((latest_price - previous_price) / previous_price) * 100
        summaries.append({
            "model": _latest_nonblank(recs, "model"),
            "kr": _latest_nonblank(recs, "part_name_kr"),
            "en": _latest_nonblank(recs, "part_name_en"),
            "previous_price": previous_price,
            "latest_price": latest_price,
            "difference": latest_price - previous_price,
            "rate": rate,
            "request_no": latest_record.ir_no,
        })
    return summaries


def build_dashboard_sheet(wb, records: list[PartRecord], manufacturer: str = "") -> None:
    ws = _reset_sheet(wb, "Dashboard", 0)
    title = f"{manufacturer} MASTER DASHBOARD" if manufacturer else "CPMS MASTER DASHBOARD"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A9"

    ws.merge_cells("A1:I2")
    ws["A1"] = title
    ws["A1"].font = Font(name="맑은 고딕", size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    changes = _price_change_summary(records)
    increases = sorted([x for x in changes if x["difference"] > 0], key=lambda x: x["rate"] or -999999, reverse=True)
    decreases = sorted([x for x in changes if x["difference"] < 0], key=lambda x: x["rate"] or 999999)

    kpis = [
        ("원본상세 행", len(records), "#,##0"),
        ("부품 그룹", len(set(_group_key(r) for r in records)), "#,##0"),
        ("적용모델", len(set(norm_model(r.model) for r in records if r.model)), "#,##0"),
        ("요청 건수", len(set(r.ir_no for r in records if r.ir_no)), "#,##0"),
        ("원본파일", len(set(r.source_file for r in records if r.source_file)), "#,##0"),
        ("가격변동", len(changes), "#,##0"),
        ("수량합계", sum(r.qty or 0 for r in records), "#,##0"),
        ("금액합계(USD)", sum(r.amount if r.amount is not None else ((r.price or 0) * (r.qty or 0)) for r in records), '#,##0.00'),
        ("기준년도 누락", sum(1 for r in records if not r.base_year), "#,##0"),
        ("단가 누락", sum(1 for r in records if r.price is None), "#,##0"),
    ]

    for i, (label, value, fmt) in enumerate(kpis):
        col = 1 + (i % 5) * 2
        row = 4 + (i // 5) * 2
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.cell(row, col).value = label
        ws.cell(row, col).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws.cell(row, col).font = Font(name="맑은 고딕", size=10, bold=True, color=HEADER_FILL)
        ws.cell(row, col).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        ws.cell(row + 1, col).value = value
        ws.cell(row + 1, col).number_format = fmt
        ws.cell(row + 1, col).font = Font(name="맑은 고딕", size=15, bold=True)
        ws.cell(row + 1, col).alignment = Alignment(horizontal="center")

    def write_table(start_row: int, title_text: str, items: list[dict], color: str) -> int:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
        ws.cell(start_row, 1).value = title_text
        ws.cell(start_row, 1).fill = PatternFill("solid", fgColor=color)
        ws.cell(start_row, 1).font = Font(name="맑은 고딕", size=12, bold=True, color=WHITE)

        header_row = start_row + 1
        for c, h in enumerate(DASHBOARD_TOP_HEADERS, 1):
            ws.cell(header_row, c).value = h
        _style_header_row(ws, header_row, 9)

        row = header_row + 1
        for rank, item in enumerate(items[:10], 1):
            vals = [
                rank, item["model"], item["kr"], item["en"],
                item["previous_price"], item["latest_price"],
                item["difference"], None if item["rate"] is None else item["rate"] / 100,
                item["request_no"],
            ]
            for c, value in enumerate(vals, 1):
                ws.cell(row, c).value = value
                ws.cell(row, c).alignment = Alignment(vertical="center", wrap_text=True)
            for c in (5, 6, 7):
                ws.cell(row, c).number_format = '#,##0.00'
            ws.cell(row, 8).number_format = '0.0%'
            row += 1

        if not items:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            ws.cell(row, 1).value = "해당 항목 없음"
            ws.cell(row, 1).alignment = Alignment(horizontal="center")
            row += 1
        return row + 1

    next_row = write_table(9, "가격 상승 TOP 10", increases, "C00000")
    write_table(next_row, "가격 하락 TOP 10", decreases, "0070C0")

    for col, width in {"A":8,"B":16,"C":28,"D":40,"E":14,"F":14,"G":14,"H":12,"I":22}.items():
        ws.column_dimensions[col].width = width

def rebuild_all_sheets(master_path: str | Path, records: list[PartRecord] | None = None, appended_count: int = 0, selected_count: int = 0, skipped_over_limit: int = 0, manufacturer: str = "") -> None:
    master_path = Path(master_path); wb = load_workbook(master_path)
    if records is None: records = parse_master(master_path)
    change_rows = build_product_sheet(wb, records, manufacturer=manufacturer)
    build_change_sheet(wb, change_rows)
    build_history_sheet(wb, records)
    build_compare_sheet(wb, records)
    write_original_detail(wb, records)
    build_model_summary(wb, records)
    build_file_year_sheet(wb, records)
    build_work_rules(wb)
    build_validation_sheet(wb, records, appended_count, selected_count, skipped_over_limit)
    build_dashboard_sheet(wb, records, manufacturer=manufacturer)
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1: del wb["Sheet"]
    wb.save(master_path)
