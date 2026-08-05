from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from core.parser import clean_text, infer_model, parse_workbook


@dataclass(frozen=True)
class HeaderFingerprint:
    parser_type: str
    confidence: int
    header_row: int
    headers: tuple[str, ...]
    reason: str


def _norm(value: object) -> str:
    return re.sub(r"[^A-Z0-9가-힣]+", "", clean_text(value).upper())


def fingerprint_workbook(path: str | Path) -> HeaderFingerprint:
    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        best: HeaderFingerprint | None = None
        for sheet in workbook.worksheets:
            for row_no, values in enumerate(
                sheet.iter_rows(
                    min_row=1,
                    max_row=min(sheet.max_row, 40),
                    max_col=min(sheet.max_column, 30),
                    values_only=True,
                ),
                start=1,
            ):
                headers = tuple(clean_text(value) for value in values if clean_text(value))
                normalized = {_norm(value) for value in headers}
                has_part = any("PARTNAME" in value or "부품명" in value for value in normalized)
                has_kr = any("KOREAN" in value or "KOR" in value or "한글" in value for value in normalized)
                has_qty = any("QTY" in value or "QUANTITY" in value or "수량" in value for value in normalized)
                has_price = any("PRICE" in value or "단가" in value for value in normalized)
                has_amount = any("TOTAL" in value or "AMOUNT" in value or "합계" in value for value in normalized)
                image_count = sum(1 for value in normalized if value.startswith("IMAGE"))

                parser_type = ""
                confidence = 0
                reason = ""

                if has_part and has_price and image_count >= 2 and not has_qty:
                    parser_type = "LEGACY-C"
                    confidence = 98
                    reason = "QTY 없음 + IMAGE 열 2개 이상 + PRICE 존재"
                elif has_part and has_qty and has_price:
                    parser_type = "LEGACY-B"
                    confidence = 97
                    reason = "PART NAME + QTY + PRICE 표준 발주형"
                elif has_part and has_price and not has_qty:
                    parser_type = "LEGACY-A"
                    confidence = 93
                    reason = "QTY 없이 공급사 단가만 회신된 가격이력형"
                elif has_part and has_qty and not has_price:
                    parser_type = "QUANTITY-ONLY"
                    confidence = 90
                    reason = "부품·수량은 있으나 단가 없음"
                else:
                    continue

                candidate = HeaderFingerprint(
                    parser_type=parser_type,
                    confidence=confidence,
                    header_row=row_no,
                    headers=headers,
                    reason=reason,
                )
                if best is None or candidate.confidence > best.confidence:
                    best = candidate

        return best or HeaderFingerprint(
            parser_type="UNKNOWN",
            confidence=20,
            header_row=0,
            headers=(),
            reason="알려진 헤더 패턴을 찾지 못함",
        )
    finally:
        workbook.close()


def infer_manufacturer_from_content(path: str | Path) -> tuple[str | None, int, str]:
    """Infer manufacturer conservatively from all parsed model names.

    Only a unanimous and known model-family signal is accepted. Mixed or
    ambiguous files remain UNKNOWN.
    """
    path = Path(path)
    try:
        records = parse_workbook(path)
    except Exception as exc:
        return None, 0, f"내용 분석 실패: {exc}"

    models = [record.model for record in records if record.model]
    if not models:
        return None, 0, "모델을 찾지 못함"

    families: set[str] = set()
    for model in models:
        normalized = re.sub(r"[^A-Z0-9]+", "", model.upper())
        if normalized.startswith("K"):
            families.add("KA")
        elif normalized.startswith("X") or normalized.startswith("L"):
            families.add("XC")
        elif normalized.startswith("A"):
            families.add("IR")
        elif normalized.startswith("N"):
            families.add("NC")

    if len(families) == 1:
        manufacturer = next(iter(families))
        confidence = 96 if len(set(models)) == 1 else 90
        return manufacturer, confidence, f"모델군 일치: {', '.join(sorted(set(models)))}"
    if len(families) > 1:
        return None, 0, f"여러 제조사 모델군 혼재: {', '.join(sorted(families))}"
    return None, 0, f"모델명으로 제조사를 확정할 수 없음: {', '.join(sorted(set(models)))}"


def classify_file(path: str | Path, filename_manufacturer: str | None) -> dict[str, object]:
    fingerprint = fingerprint_workbook(path)
    manufacturer = filename_manufacturer
    manufacturer_confidence = 100 if manufacturer else 0
    manufacturer_reason = "파일명 기준" if manufacturer else ""

    if not manufacturer:
        manufacturer, manufacturer_confidence, manufacturer_reason = (
            infer_manufacturer_from_content(path)
        )

    return {
        "manufacturer": manufacturer or "UNKNOWN",
        "manufacturer_confidence": manufacturer_confidence,
        "manufacturer_reason": manufacturer_reason,
        "parser_type": fingerprint.parser_type,
        "parser_confidence": fingerprint.confidence,
        "header_row": fingerprint.header_row,
        "headers": fingerprint.headers,
        "parser_reason": fingerprint.reason,
    }
