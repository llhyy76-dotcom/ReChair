from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable, Mapping, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER = "1F4E79"
LIGHT_BLUE = "DDEBF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_RED = "FCE4D6"
LIGHT_YELLOW = "FFF2CC"
WHITE = "FFFFFF"
GRID = "D9E2F3"


def _value(obj: Any, name: str, default: Any = "") -> Any:
    return getattr(obj, name, default)


def _style_header(ws, row: int, max_col: int) -> None:
    thin = Side(style="thin", color=GRID)
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _box(ws, label_cell: str, value_cell: str, label: str, value: Any, fill: str = LIGHT_BLUE, number_format: str = "#,##0") -> None:
    ws[label_cell] = label
    ws[label_cell].fill = PatternFill("solid", fgColor=fill)
    ws[label_cell].font = Font(name="맑은 고딕", size=10, bold=True, color=HEADER)
    ws[label_cell].alignment = Alignment(horizontal="center", vertical="center")
    ws[value_cell] = value
    ws[value_cell].font = Font(name="맑은 고딕", size=16, bold=True)
    ws[value_cell].alignment = Alignment(horizontal="center", vertical="center")
    ws[value_cell].number_format = number_format


def write_update_result_report(
    *,
    log_dir: str | Path,
    scanned_files: int,
    selected_files: int,
    duplicates: int,
    revisions: int,
    new_parts: int,
    updated_rows: int,
    kr_completed: int,
    price_changes: int,
    skipped_over_limit: int,
    by_manufacturer: Mapping[str, Any],
    rejected_entries: Iterable[Any],
    audit_entries: Iterable[Any],
    manifest_path: str | Path,
    rejected_manifest_path: str | Path | None = None,
    file_audit_path: str | Path | None = None,
    history_events: int = 0,
    history_path: str | Path | None = None,
    mode: str = "업데이트",
) -> Path:
    """Create an Excel report summarizing one CPMS run."""
    log_dir = Path(log_dir)
    log_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = log_dir / f"Update_Result_{stamp}.xlsx"

    rejected = list(rejected_entries)
    audits = list(audit_entries)
    completed = sum(1 for entry in audits if _value(entry, "status") == "처리 완료")
    pdf_review = sum(1 for entry in audits if "PDF" in str(_value(entry, "status")))
    failed = sum(
        1 for entry in audits
        if _value(entry, "status") in {"파싱 실패", "제조사 판정 불가", "Master 저장 실패", "PDF OCR 실패"}
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "업데이트_요약"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J2")
    ws["A1"] = f"CPMS {mode} 결과 보고서"
    ws["A1"].fill = PatternFill("solid", fgColor=HEADER)
    ws["A1"].font = Font(name="맑은 고딕", size=20, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"] = f"실행일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A3"].font = Font(name="맑은 고딕", size=10, bold=True)

    boxes = [
        ("A5", "A6", "Update 후보", scanned_files, LIGHT_BLUE),
        ("C5", "C6", "처리 완료 파일", selected_files, LIGHT_GREEN),
        ("E5", "E6", "신규 원본행", new_parts, LIGHT_GREEN),
        ("G5", "G6", "기존행 보정", updated_rows, LIGHT_YELLOW),
        ("I5", "I6", "가격변동", price_changes, LIGHT_YELLOW),
        ("A8", "A9", "KR 보완", kr_completed, LIGHT_BLUE),
        ("C8", "C9", "중복 raw", duplicates, LIGHT_BLUE),
        ("E8", "E9", "수정본 제외", revisions, LIGHT_BLUE),
        ("G8", "G9", "처리 보류", skipped_over_limit, LIGHT_YELLOW),
        ("I8", "I9", "History 기록", history_events, LIGHT_GREEN),
    ]
    for label_cell, value_cell, label, value, fill in boxes:
        _box(ws, label_cell, value_cell, label, value, fill)
        # Merge KPI boxes across two columns.
        start_col = ws[label_cell].column
        start_row = ws[label_cell].row
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
        ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 1, end_column=start_col + 1)

    row = 12
    headers = ["제조사", "스캔", "선택", "신규", "보정", "KR보완", "가격변동", "중복", "수정본 제외", "상태"]
    for col, header in enumerate(headers, 1):
        ws.cell(row, col).value = header
    _style_header(ws, row, len(headers))
    row += 1

    manufacturers = ["IR", "XC", "AC", "KA", "NC"]
    for manufacturer in manufacturers:
        result = by_manufacturer.get(manufacturer)
        values = [
            manufacturer,
            _value(result, "scanned_files", 0),
            _value(result, "selected_files", 0),
            _value(result, "new_parts", 0),
            _value(result, "updated_rows", 0),
            _value(result, "kr_completed", 0),
            _value(result, "price_changes", 0),
            _value(result, "duplicates", 0),
            _value(result, "revisions", 0),
            _value(result, "status", "처리 없음") or "처리 없음",
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col).value = value
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
        row += 1

    row += 1
    ws.cell(row, 1).value = "관련 결과 파일"
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=HEADER)
    ws.cell(row, 1).font = Font(name="맑은 고딕", bold=True, color=WHITE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    paths = [
        ("처리명세", manifest_path),
        ("미반영 진단표", rejected_manifest_path),
        ("전체 파일 감사표", file_audit_path),
        ("변경이력", history_path),
    ]
    for label, path in paths:
        if path:
            ws.cell(row, 1).value = label
            ws.cell(row, 2).value = str(path)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
            row += 1

    for col, width in {"A":16, "B":12, "C":12, "D":12, "E":12, "F":12, "G":14, "H":12, "I":14, "J":45}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A13"

    # File audit sheet.
    audit_ws = wb.create_sheet("파일별_처리결과")
    audit_headers = [
        "파일명", "제조사", "요청번호", "처리상태", "추출행",
        "단가행", "수량행", "금액행", "적용모델", "Archive 경로", "비고",
    ]
    for col, header in enumerate(audit_headers, 1):
        audit_ws.cell(1, col).value = header
    _style_header(audit_ws, 1, len(audit_headers))

    for row_index, entry in enumerate(sorted(audits, key=lambda item: (_value(item, "manufacturer"), _value(item, "filename"))), 2):
        values = [
            _value(entry, "filename"),
            _value(entry, "manufacturer"),
            _value(entry, "request_no"),
            _value(entry, "status"),
            _value(entry, "extracted_rows", 0),
            _value(entry, "priced_rows", 0),
            _value(entry, "qty_rows", 0),
            _value(entry, "amount_rows", 0),
            _value(entry, "models"),
            _value(entry, "archive_path"),
            _value(entry, "note"),
        ]
        for col, value in enumerate(values, 1):
            audit_ws.cell(row_index, col).value = value
            audit_ws.cell(row_index, col).alignment = Alignment(vertical="center", wrap_text=True)

    audit_widths = [55, 12, 20, 20, 10, 10, 10, 10, 35, 65, 60]
    for col, width in enumerate(audit_widths, 1):
        audit_ws.column_dimensions[get_column_letter(col)].width = width
    audit_ws.freeze_panes = "A2"
    audit_ws.auto_filter.ref = f"A1:K{max(audit_ws.max_row, 1)}"
    audit_ws.sheet_view.showGridLines = False

    # Rejected detail sheet.
    reject_ws = wb.create_sheet("미반영_상세")
    reject_headers = ["파일명", "제조사", "요청번호", "단계", "원인코드", "상세원인", "권장조치", "파일경로"]
    for col, header in enumerate(reject_headers, 1):
        reject_ws.cell(1, col).value = header
    _style_header(reject_ws, 1, len(reject_headers))

    for row_index, entry in enumerate(rejected, 2):
        values = [
            _value(entry, "filename"),
            _value(entry, "manufacturer"),
            _value(entry, "request_no"),
            _value(entry, "stage"),
            _value(entry, "reason_code"),
            _value(entry, "reason"),
            _value(entry, "action"),
            _value(entry, "file_path"),
        ]
        for col, value in enumerate(values, 1):
            reject_ws.cell(row_index, col).value = value
            reject_ws.cell(row_index, col).alignment = Alignment(vertical="center", wrap_text=True)

    reject_widths = [55, 12, 20, 20, 28, 65, 65, 65]
    for col, width in enumerate(reject_widths, 1):
        reject_ws.column_dimensions[get_column_letter(col)].width = width
    reject_ws.freeze_panes = "A2"
    reject_ws.auto_filter.ref = f"A1:H{max(reject_ws.max_row, 1)}"
    reject_ws.sheet_view.showGridLines = False

    wb.save(output)
    return output
