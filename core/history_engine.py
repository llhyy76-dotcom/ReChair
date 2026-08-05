from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.parser import PartRecord, norm_model, norm_txt


CHANGE_HEADERS = [
    "이벤트ID", "실행ID", "기록일시", "제조사", "요청번호", "원본파일", "원본행",
    "변경유형", "변경항목", "모델", "부품번호", "부품명(영문)", "부품명(한글)",
    "이전값", "변경값", "기준년도", "통화",
]

RUN_HEADERS = [
    "실행ID", "실행일시", "제조사", "처리파일수", "요청번호", "원본파일",
    "업데이트전행수", "업데이트후행수", "실제신규", "실제삭제", "필드변경",
    "재처리행", "이번실행가격이력검출", "누적가격변동부품", "변경이력건수", "상태", "Master경로", "Backup경로",
]

PRICE_HISTORY_HEADERS = [
    "가격이력이벤트ID", "제조사", "모델", "부품번호", "부품명(영문)", "부품명(한글)",
    "이전단가", "최근단가", "변동액", "변동률(%)", "이전요청번호", "최근요청번호",
    "이전원본파일", "최근원본파일", "이전기준일", "최근기준일", "통화",
]


TRACKED_FIELDS = [
    ("price", "단가"),
    ("qty", "수량"),
    ("amount", "금액"),
    ("model", "적용모델"),
    ("part_name_en", "부품명(영문)"),
    ("part_name_kr", "부품명(한글)"),
    ("part_name_cn", "부품명(중문)"),
    ("part_no", "부품번호"),
    ("parameter", "SPEC/Parameter"),
    ("base_year", "기준년도"),
    ("base_date", "기준일자"),
    ("currency", "통화"),
]


@dataclass(frozen=True)
class HistoryEvent:
    event_id: str
    run_id: str
    recorded_at: str
    manufacturer: str
    request_no: str
    source_file: str
    source_row: int | None
    change_type: str
    field_name: str
    model: str
    part_no: str
    part_name_en: str
    part_name_kr: str
    old_value: str
    new_value: str
    base_year: str
    currency: str

    def as_row(self) -> list[object]:
        return [
            self.event_id, self.run_id, self.recorded_at, self.manufacturer, self.request_no,
            self.source_file, self.source_row, self.change_type, self.field_name,
            self.model, self.part_no, self.part_name_en, self.part_name_kr,
            self.old_value, self.new_value, self.base_year, self.currency,
        ]


@dataclass(frozen=True)
class HistoryRun:
    run_id: str
    recorded_at: str
    manufacturer: str
    processed_files: int
    request_nos: str
    source_files: str
    before_rows: int
    after_rows: int
    actual_new: int
    actual_deleted: int
    field_changes: int
    reprocessed_rows: int
    detected_price_histories: int
    cumulative_price_variations: int
    change_events: int
    status: str
    master_path: str
    backup_path: str

    def as_row(self) -> list[object]:
        return [
            self.run_id, self.recorded_at, self.manufacturer, self.processed_files,
            self.request_nos, self.source_files, self.before_rows, self.after_rows,
            self.actual_new, self.actual_deleted, self.field_changes,
            self.reprocessed_rows, self.detected_price_histories, self.cumulative_price_variations,
            self.change_events, self.status,
            self.master_path, self.backup_path,
        ]


def make_run_id(manufacturer: str, source_files: Sequence[str], now: datetime | None = None) -> str:
    stamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S_%f")
    suffix = hashlib.sha1("|".join(sorted(source_files)).encode("utf-8")).hexdigest()[:8]
    return f"{stamp}_{manufacturer}_{suffix}"


def _identity(record: PartRecord) -> tuple[str, int | None, str]:
    """Stable identity for a physical row in one request source."""
    return (record.source_file or "", record.source_row, record.ir_no or "")


def _fallback_identity(record: PartRecord) -> tuple[str, str, str, str]:
    return (
        norm_model(record.model), norm_txt(record.part_name_en),
        norm_txt(record.part_name_kr), norm_txt(record.part_no),
    )


def _key(record: PartRecord) -> tuple[object, ...]:
    if record.source_file:
        return ("SOURCE",) + _identity(record)
    return ("PART",) + _fallback_identity(record)


def _display(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.10g}"
    return str(value).strip()


def _event_id(payload: dict[str, object]) -> str:
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def _make_event(
    manufacturer: str,
    run_id: str,
    record: PartRecord,
    change_type: str,
    field_name: str,
    old_value: object,
    new_value: object,
    recorded_at: str,
) -> HistoryEvent:
    payload = {
        "run_id": run_id,
        "manufacturer": manufacturer,
        "request_no": record.ir_no,
        "source_file": record.source_file,
        "source_row": record.source_row,
        "change_type": change_type,
        "field_name": field_name,
        "old_value": _display(old_value),
        "new_value": _display(new_value),
    }
    return HistoryEvent(
        event_id=_event_id(payload), run_id=run_id, recorded_at=recorded_at,
        manufacturer=manufacturer, request_no=record.ir_no,
        source_file=record.source_file, source_row=record.source_row,
        change_type=change_type, field_name=field_name,
        model=record.model, part_no=record.part_no,
        part_name_en=record.part_name_en, part_name_kr=record.part_name_kr,
        old_value=_display(old_value), new_value=_display(new_value),
        base_year=str(record.base_year or ""), currency=record.currency or "USD",
    )


def compare_snapshots(
    manufacturer: str,
    before: Iterable[PartRecord],
    after: Iterable[PartRecord],
    *,
    run_id: str = "",
    recorded_at: str | None = None,
) -> list[HistoryEvent]:
    """Compare immutable pre-update and candidate post-update snapshots.

    This function must be called before Master.xlsx is overwritten. The returned events
    can be persisted only after Master save verification succeeds.
    """
    before_map = {_key(r): r for r in list(before)}
    after_map = {_key(r): r for r in list(after)}
    recorded_at = recorded_at or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    run_id = run_id or make_run_id(manufacturer, [r.source_file for r in after_map.values() if r.source_file])

    events: list[HistoryEvent] = []
    before_keys = set(before_map)
    after_keys = set(after_map)

    for key in sorted(after_keys - before_keys, key=str):
        record = after_map[key]
        events.append(_make_event(manufacturer, run_id, record, "신규", "원본행", "", "추가", recorded_at))

    for key in sorted(before_keys - after_keys, key=str):
        record = before_map[key]
        events.append(_make_event(manufacturer, run_id, record, "삭제", "원본행", "존재", "삭제", recorded_at))

    for key in sorted(before_keys & after_keys, key=str):
        old = before_map[key]
        new = after_map[key]
        for attr, label in TRACKED_FIELDS:
            old_value = getattr(old, attr)
            new_value = getattr(new, attr)
            if _display(old_value) != _display(new_value):
                events.append(_make_event(
                    manufacturer, run_id, new, "변경", label,
                    old_value, new_value, recorded_at,
                ))
    return events


def summarize_events(events: Iterable[HistoryEvent]) -> tuple[int, int, int]:
    event_list = list(events)
    return (
        sum(1 for e in event_list if e.change_type == "신규"),
        sum(1 for e in event_list if e.change_type == "삭제"),
        sum(1 for e in event_list if e.change_type == "변경"),
    )


def make_history_run(
    *, manufacturer: str, run_id: str, source_files: Sequence[str], request_nos: Sequence[str],
    before_rows: int, after_rows: int, events: Sequence[HistoryEvent], reprocessed_rows: int,
    master_path: str | Path, backup_path: str | Path, status: str = "완료",
    detected_price_histories: int = 0, cumulative_price_variations: int = 0,
    recorded_at: str | None = None,
) -> HistoryRun:
    actual_new, actual_deleted, field_changes = summarize_events(events)
    return HistoryRun(
        run_id=run_id,
        recorded_at=recorded_at or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        manufacturer=manufacturer,
        processed_files=len(source_files),
        request_nos=", ".join(sorted(set(filter(None, request_nos)))),
        source_files=", ".join(source_files),
        before_rows=before_rows, after_rows=after_rows,
        actual_new=actual_new, actual_deleted=actual_deleted,
        field_changes=field_changes, reprocessed_rows=reprocessed_rows,
        detected_price_histories=detected_price_histories,
        cumulative_price_variations=cumulative_price_variations,
        change_events=len(events), status=status,
        master_path=str(master_path), backup_path=str(backup_path),
    )


def _style_header(ws, headers: list[str], widths: list[int]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def append_history(
    history_dir: str | Path,
    events: Iterable[HistoryEvent],
    run: HistoryRun | None = None,
    price_history_events: Iterable[object] = (),
) -> tuple[Path | None, int]:
    """Persist one committed update transaction and its unique field-level changes.

    A run row is always recorded even when actual changes are zero. This prevents a
    successful reprocessing run from looking like History Engine did not execute.
    Return value remains (path, added_change_events) for compatibility.
    """
    event_list = list(events)
    price_event_list = list(price_history_events)
    if not event_list and run is None and not price_event_list:
        return None, 0

    history_dir = Path(history_dir)
    history_dir.mkdir(parents=True, exist_ok=True)
    year = datetime.now().year
    path = history_dir / f"History_{year}.xlsx"

    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "실행이력" not in wb.sheetnames:
        ws_run = wb.create_sheet("실행이력", 0)
        ws_run.append(RUN_HEADERS)
        _style_header(ws_run, RUN_HEADERS, [31, 20, 10, 12, 28, 55, 14, 14, 12, 12, 12, 12, 18, 18, 14, 12, 55, 55])
    else:
        ws_run = wb["실행이력"]

    if "변경이력" not in wb.sheetnames:
        ws_change = wb.create_sheet("변경이력")
        ws_change.append(CHANGE_HEADERS)
        _style_header(ws_change, CHANGE_HEADERS, [26, 31, 20, 10, 22, 55, 10, 12, 20, 18, 18, 35, 35, 25, 25, 12, 10])
    else:
        ws_change = wb["변경이력"]

    if "가격이력" not in wb.sheetnames:
        ws_price = wb.create_sheet("가격이력")
        ws_price.append(PRICE_HISTORY_HEADERS)
        _style_header(ws_price, PRICE_HISTORY_HEADERS, [26, 10, 18, 18, 35, 35, 14, 14, 14, 14, 22, 22, 48, 48, 14, 14, 10])
    else:
        ws_price = wb["가격이력"]

    existing_run_ids = {str(ws_run.cell(r, 1).value or "") for r in range(2, ws_run.max_row + 1)}
    if run is not None and run.run_id not in existing_run_ids:
        ws_run.append(run.as_row())

    existing_event_ids = {str(ws_change.cell(r, 1).value or "") for r in range(2, ws_change.max_row + 1)}
    added = 0
    for event in event_list:
        if event.event_id in existing_event_ids:
            continue
        ws_change.append(event.as_row())
        existing_event_ids.add(event.event_id)
        added += 1

    existing_price_ids = {str(ws_price.cell(r, 1).value or "") for r in range(2, ws_price.max_row + 1)}
    for event in price_event_list:
        event_id = str(getattr(event, "event_id", "") or "")
        if not event_id or event_id in existing_price_ids:
            continue
        ws_price.append(event.as_row())
        existing_price_ids.add(event_id)

    ws_run.auto_filter.ref = f"A1:{get_column_letter(len(RUN_HEADERS))}{ws_run.max_row}"
    ws_change.auto_filter.ref = f"A1:{get_column_letter(len(CHANGE_HEADERS))}{ws_change.max_row}"
    ws_price.auto_filter.ref = f"A1:{get_column_letter(len(PRICE_HISTORY_HEADERS))}{ws_price.max_row}"
    wb.save(path)
    return path, added
