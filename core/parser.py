from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.excel_reader import open_workbook

MANUFACTURERS = ("IR", "XC", "AC", "KA", "NC")
REQUEST_RE = re.compile(r"(?<![A-Z0-9])(IR|XC|AC|KA|NC)\s*[-_]?([0-9]{2,6}(?:\s*(?:[~\-,/&]|TO)\s*[0-9]{1,6})?)", re.I)
MODEL_RE = re.compile(r"(?:CMC|CMS|HM)?-?([A-Z]{0,3}\d{1,4}[A-Z]?)(?:\(G\))?", re.I)
DATE_RE = re.compile(r"(\d{2,4})[.\-/년 ]+(\d{1,2})[.\-/월 ]+(\d{1,2})")

@dataclass(frozen=True)
class PartRecord:
    base_year: int | None = None
    base_date: str = ""
    ir_no: str = ""  # field name kept for backward compatibility; stores IR/XC/AC/KA/NC request no.
    model: str = ""
    part_name_kr: str = ""
    part_name_en: str = ""
    part_name_cn: str = ""
    price: float | None = None
    qty: float | None = None
    amount: float | None = None
    part_no: str = ""
    parameter: str = ""
    source_file: str = ""
    source_row: int | None = None
    extraction_method: str = ""
    currency: str = "USD"
    updated: str = ""

    @property
    def part_key(self) -> tuple[str, str, str, str]:
        return (norm_model(self.model), norm_txt(self.part_name_en), norm_txt(self.part_name_kr), norm_txt(self.part_no))

    @property
    def raw_key(self) -> tuple:
        return (self.source_file, self.source_row, self.ir_no, norm_model(self.model), norm_txt(self.part_name_en), norm_txt(self.part_name_kr), self.price, self.qty)

    def as_dict(self) -> dict:
        return self.__dict__.copy()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    return re.sub(r"\s+", " ", text)


def norm_txt(value: object) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).upper()


def norm_model(value: object) -> str:
    text = clean_text(value).upper().replace("(G)", "").replace("CMC-", "").replace("CMS-", "")
    m = MODEL_RE.search(text)
    return m.group(1).upper() if m else text


def parse_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^0-9.\-]", "", str(value))
    if text in {"", ".", "-", "-."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None

parse_price = parse_number


def parse_money(value: object) -> float | None:
    """Parse USD values including legacy decimal-comma formatting.

    Examples:
    - $21.00 -> 21.0
    - $21,00 -> 21.0
    - $1,500.00 -> 1500.0
    - $1.500,00 -> 1500.0
    """
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"[^0-9,.\-]", "", text)
    if text in {"", ".", ",", "-", "-.", "-,"}:
        return None

    if "," in text and "." in text:
        # The rightmost punctuation is treated as the decimal separator.
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        pieces = text.split(",")
        if len(pieces[-1]) == 2:
            text = "".join(pieces[:-1]) + "." + pieces[-1]
        else:
            text = "".join(pieces)

    try:
        return float(text)
    except ValueError:
        return None


def _has_currency_marker(value: object, number_format: str = "") -> bool:
    """Return True when a cell visually represents money, not a quantity.

    Some legacy XC request files contain hidden/shifted columns where the
    visible header says QTY(PCS) but the cell value is $57.00. Such values must
    never be used as quantity. We check both the displayed text and Excel number
    format because openpyxl may return numeric values for currency-formatted
    cells.
    """
    text = "" if value is None else str(value).strip().upper()
    nf = (number_format or "").upper()
    return any(token in text for token in ("$", "USD", "US$", "￥", "₩")) or any(token in nf for token in ("$", "USD", "₩", "￥"))


def _cell_num_and_money(sheet: Worksheet, row_no: int, col_no: int) -> tuple[float | None, bool]:
    cell = sheet.cell(row_no, col_no)
    is_money = _has_currency_marker(cell.value, getattr(cell, "number_format", ""))
    number = parse_money(cell.value) if is_money else parse_number(cell.value)
    return number, is_money


def _valid_qty_value(num: float | None, is_money: bool) -> bool:
    """Quantity must be plain numeric, non-currency, and usually integer-like."""
    if num is None or is_money or num < 0:
        return False
    # Most request quantities are integers. Allow a tiny tolerance for Excel floats.
    return abs(num - round(num)) < 0.0001


def infer_manufacturer(path_or_text: str | Path) -> str | None:
    """Infer manufacturer even when the request number is not at filename start.

    Windows files often have prefixes such as ``수정_`` or dates before NC1974.
    The old word-boundary regex failed because Korean letters and underscores
    are Unicode word characters. The new rule searches the request number
    anywhere while still preventing matches inside ASCII product codes.
    """
    raw = str(path_or_text)
    name = Path(raw).name
    upper = name.upper()

    match = REQUEST_RE.search(upper)
    if match:
        return match.group(1).upper()

    # Label-style filenames do not always contain a formal request number:
    #   2.NC부품 단가 확인 요청 2104 - 회신.xlsx
    #   1.XC부품 단가 확인 요청 210415.xlsx
    # Korean characters are treated as a safe delimiter from ASCII product codes.
    label_match = re.search(
        r"(?:^|[^A-Z0-9])(IR|XC|AC|KA|NC)(?=[^A-Z0-9]|$)",
        upper,
    )
    if label_match:
        return label_match.group(1).upper()

    normalized = re.sub(r"[^A-Z0-9]+", "_", upper)

    # Ningbo historical report files belong to NC, but are handled separately
    # from normal supplier request sheets.
    if "DEFECTIVE_SPARE_PARTS_REPORT" in normalized:
        return "NC"

    # Narrow fallback for the 2020 CMC-930 Ningbo request whose filename has no
    # NC request number.
    if re.search(r"^CMC_?930_.*20\d{6}\.(?:XLS|XLSX|XLSM)$", normalized):
        return "NC"

    # Legacy AC file retained from v210.
    if re.fullmatch(r"FOC_SPARE_PARTS_REQUEST_20(?:17)\d{4}\.XLSX?", normalized):
        return "AC"
    return None

def infer_ir_no(path: str | Path) -> str:
    name = Path(path).stem.upper()
    match = REQUEST_RE.search(name)
    if match:
        prefix = match.group(1).upper()
        number = re.sub(r"\s+", "", match.group(2))
        number = re.sub(r"[,/&-]+", "~", number)
        return f"{prefix}{number}"

    manufacturer = infer_manufacturer(path)

    # Extract the short request token placed after words such as 요청/발주.
    # For the supplied file this produces NC2104 instead of using the full filename.
    if manufacturer:
        labelled_request = re.search(
            rf"(?:요청|발주|REQUEST)[^0-9]{{0,12}}(\d{{4,6}})",
            name,
            re.I,
        )
        if labelled_request:
            return f"{manufacturer}{labelled_request.group(1)}"

        # Fallback: manufacturer label followed later by a 4–6 digit request token.
        after_label = re.search(
            rf"(?:^|[^A-Z0-9]){manufacturer}[^0-9]{{0,24}}(\d{{4,6}})",
            name,
            re.I,
        )
        if after_label:
            return f"{manufacturer}{after_label.group(1)}"

    date_token = re.search(r"(20\d{6})", name)
    if manufacturer == "NC" and "DEFECTIVE_SPARE_PARTS_REPORT" in re.sub(r"[^A-Z0-9]+", "_", name):
        return f"NC-REPORT-{date_token.group(1) if date_token else 'UNKNOWN'}"
    if manufacturer == "NC" and re.search(r"CMC[-_ ]?930", name):
        return f"NC-REQ-{date_token.group(1) if date_token else 'UNKNOWN'}"
    if manufacturer == "AC" and date_token:
        return f"AC-FOC-{date_token.group(1)}"
    return Path(path).stem

def normalize_request_no(value: object, fallback: str = "") -> str:
    """Normalize legacy request ranges without losing the ending number.

    Examples:
    - IR18284, 285 -> IR18284~285
    - IR18284~285 -> IR18284~285
    - KA1901 -> KA1901
    """
    text = clean_text(value).upper()
    if not text:
        return fallback

    comma_range = re.search(
        r"\b(IR|XC|AC|KA|NC)\s*[-_]?\s*(\d{2,6})\s*[,/&]\s*(\d{1,6})\b",
        text,
        re.I,
    )
    if comma_range:
        prefix = comma_range.group(1).upper()
        start = comma_range.group(2)
        end = comma_range.group(3)
        return f"{prefix}{start}~{end}"

    direct = REQUEST_RE.search(text)
    if direct:
        prefix = direct.group(1).upper()
        number = re.sub(r"\s+", "", direct.group(2))
        number = re.sub(r"[,/&-]+", "~", number)
        return f"{prefix}{number}"
    return fallback


def infer_year_from_ir(ir_no: str) -> int | None:
    text = clean_text(ir_no).upper()
    m = REQUEST_RE.search(text)
    if not m:
        m2 = re.search(r"(\d{2})", text)
        if not m2:
            return None
        yy = int(m2.group(1))
    else:
        yy = int(m.group(2)[:2])
    return 2000 + yy if 0 <= yy <= 69 else 1900 + yy


def infer_date_from_filename(path: str | Path) -> str:
    name = Path(path).stem
    candidates = re.findall(r"(?:^|_)(\d{6}|\d{8})(?:$|\D)", name)
    token = candidates[-1] if candidates else ""
    if token:
        if len(token) == 6:
            yy = int(token[:2]); yyyy = 2000 + yy if yy < 70 else 1900 + yy
            mm, dd = int(token[2:4]), int(token[4:6])
        else:
            yyyy, mm, dd = int(token[:4]), int(token[4:6]), int(token[6:8])
        try:
            return date(yyyy, mm, dd).isoformat()
        except ValueError:
            pass
    return ""


def _looks_like_component_spec(token: str) -> bool:
    """Identify motor/electrical specifications and motion labels, not models."""
    compact = re.sub(r"\s+", "", clean_text(token).upper())
    if not compact:
        return False
    return bool(
        re.fullmatch(r"[234]D", compact)
        or re.search(r"(?:\d+)?ZYT\d", compact)
        or re.search(r"\bDC\d+(?:V)?\b", compact)
        or re.search(r"\bAC\d+(?:V)?\b", compact)
        or re.search(r"\d+(?:RPM|W|V|HZ)\b", compact)
    )


def infer_model(text: str) -> str:
    text = clean_text(text)
    explicit = re.search(r"(?:CMC|CMS|HM)-([A-Z]?\d{1,4}[A-Z]?)(?:\(G\))?", text, re.I)
    if explicit:
        return explicit.group(1).upper()

    # Prefer a clean comma-separated model token. Component specifications such
    # as 62ZYT110-35G and movement labels such as 3D are ignored.
    parts = [part.strip() for part in text.split(",") if part.strip()]
    for part in reversed(parts):
        if _looks_like_component_spec(part):
            continue
        compact = part.upper().replace("(G)", "").strip()
        exact = re.fullmatch(r"(?:CMC-|CMS-|HM-)?([A-Z]{0,3}\d{1,4}[A-Z]?)", compact)
        if exact:
            return exact.group(1).upper()

    if not _looks_like_component_spec(text):
        match = MODEL_RE.search(text)
        if match and not _looks_like_component_spec(match.group(0)):
            return match.group(1).upper()
    return ""


def revision_rank(path: str | Path) -> int:
    name = Path(path).stem.lower()
    rank = 0
    rev = re.search(r"rev\s*[-_ ]?(\d+)", name)
    if rev:
        rank = max(rank, 20 + int(rev.group(1)))
    if "수정" in name or "modify" in name or "revised" in name:
        rank = max(rank, 50)
    return rank


def choose_preferred_files(paths: Iterable[Path]) -> list[Path]:
    """Keep only the latest/preferred file for the same request number.

    Selection priority:
    1) revised/modified filename marker,
    2) newer modified time,
    3) larger file size.
    This prevents duplicate request files from being accumulated during normal
    update or full rebuild.
    """
    grouped: dict[str, Path] = {}

    def score(path: Path) -> tuple[int, float, int]:
        try:
            stat = path.stat()
            return (revision_rank(path), stat.st_mtime, stat.st_size)
        except OSError:
            return (revision_rank(path), 0.0, 0)

    for path in paths:
        req_no = infer_ir_no(path)
        current = grouped.get(req_no)
        if current is None or score(path) > score(current):
            grouped[req_no] = path
    return sorted(grouped.values(), key=lambda item: item.name)


def find_metadata(sheet: Worksheet, path: str | Path) -> tuple[str, str, str]:
    request_no = infer_ir_no(path)
    base_date = infer_date_from_filename(path)
    model = infer_model(Path(path).stem) or infer_model(sheet.title)
    for row in sheet.iter_rows(min_row=1, max_row=min(12, sheet.max_row)):
        for value in [cell.value for cell in row]:
            if isinstance(value, datetime) and not base_date:
                base_date = value.date().isoformat()
            text = clean_text(value)
            if not text:
                continue
            normalized_request = normalize_request_no(text, "")
            if normalized_request:
                # Do not replace a full filename range with only its first number.
                if "~" in normalized_request or "~" not in request_no:
                    request_no = normalized_request
            dm = DATE_RE.search(text)
            if dm and not base_date:
                y = int(dm.group(1)); y = y + 2000 if y < 100 else y
                try:
                    base_date = date(y, int(dm.group(2)), int(dm.group(3))).isoformat()
                except ValueError:
                    pass
            if not model:
                found = infer_model(text)
                if found:
                    model = found
    return request_no, base_date, model


def _header_norm(value: object) -> str:
    return re.sub(r"[^A-Z0-9가-힣]+", "", clean_text(value).upper())


def _is_qty_header(txt: str, nt: str) -> bool:
    return (
        nt in {"QTY", "QTYPCS", "QTYEA", "QUANTITY", "QUANTITYPCS", "PCS", "수량", "수량PCS", "수량EA"}
        or "수량" in txt
        or nt.startswith("QTY")
        or nt.startswith("QUANTITY")
    )


def _is_price_header(txt: str, nt: str) -> bool:
    return (
        nt in {"PRICE", "PRICEUSD", "UNITPRICE", "UNITPRICEUSD", "USD", "단가", "단가USD", "가격", "가격USD"}
        or "단가" in txt
        or "UNITPRICE" in nt
        or nt.startswith("PRICE")
        or ("PRICE" in nt and "QTY" not in nt)
    )


def _is_amount_header(txt: str, nt: str) -> bool:
    return (
        nt in {"TOTAL", "TOTALUSD", "AMOUNT", "AMOUNTUSD", "금액", "금액USD", "합계", "합계USD"}
        or "금액" in txt
        or nt.startswith("TOTAL")
        or nt.startswith("AMOUNT")
    )


def _looks_like_header_row(mapping: dict[str, int]) -> int:
    return (
        len(mapping)
        + (6 if "part_name_en" in mapping else 0)
        + (4 if "qty" in mapping else 0)
        + (4 if "price" in mapping else 0)
        + (2 if "amount" in mapping else 0)
    )


def _detect_part_name_columns(header_texts: list[str], mapping: dict[str, int]) -> None:
    candidates: list[tuple[int, str, str]] = []
    duplicate_parts_list: list[int] = []

    for idx, txt in enumerate(header_texts, start=1):
        stripped = txt.strip()
        nt = _header_norm(txt)
        is_parts_list = nt in {"PARTSLIST", "SPAREPARTSLIST", "SPAREPART", "SPAREPARTS"}
        is_legacy_kr = stripped in {"이름", "한글명", "한국어명"}

        if nt == "PARTSLIST":
            duplicate_parts_list.append(idx)

        if (
            "PARTNAME" in nt
            or "DESCRIPTION" in nt
            or "품명" in txt
            or "부품명" in txt
            or is_parts_list
            or is_legacy_kr
        ):
            candidates.append((idx, txt, nt))

    if not candidates:
        return

    # KA1901-type forms: the first Parts List column is Korean and the second
    # Parts List column is English.
    if len(duplicate_parts_list) >= 2:
        mapping["part_name_kr"] = duplicate_parts_list[0]
        mapping["part_name_en"] = duplicate_parts_list[1]

    # Explicit language labels always win.  In particular, a common NC layout
    # uses ``부품명`` for Korean and ``부품명(Eng)`` for English.  The old
    # generic-first rule selected the bare Korean column as English and swapped
    # both names in Master.
    for idx, txt, nt in candidates:
        stripped = txt.strip()
        if (
            "ENGLISH" in nt
            or nt.endswith("ENG")
            or "영문" in txt
            or stripped.upper() in {"PART NAME(ENG)", "PARTNAME(ENG)"}
        ):
            mapping["part_name_en"] = idx
        elif (
            "KOREAN" in nt
            or "KOR" in nt
            or "한글" in txt
            or "한국" in txt
            or "KOREA" in nt
            or stripped in {"이름", "한글명", "한국어명"}
        ):
            mapping["part_name_kr"] = idx
        elif "CHINESE" in nt or "CHINA" in nt or "중국" in txt:
            mapping["part_name_cn"] = idx

    # When an explicit English column exists, a neighbouring bare ``부품명``
    # column is the Korean/local-language column unless another Korean column
    # was already identified.
    if "part_name_en" in mapping and "part_name_kr" not in mapping:
        for idx, txt, nt in candidates:
            if idx == mapping.get("part_name_en"):
                continue
            if txt.strip() in {"부품명", "품명"} or nt in {"PARTNAME", "DESCRIPTION"}:
                mapping["part_name_kr"] = idx
                break

    for idx, _, _ in candidates:
        if idx not in {mapping.get("part_name_kr"), mapping.get("part_name_cn")}:
            mapping.setdefault("part_name_en", idx)
            break

    if "part_name_en" not in mapping and candidates:
        mapping["part_name_en"] = candidates[-1][0]

    if "part_name_kr" not in mapping and len(candidates) >= 2:
        for idx, _, _ in candidates:
            if idx != mapping.get("part_name_en"):
                mapping["part_name_kr"] = idx
                break

def locate_ir_columns(sheet: Worksheet) -> tuple[int, dict[str, object]]:
    """Find request table columns by header text, not fixed position.

    v2.0.4 핵심:
    - QTY / Price / Total 위치 고정 금지
    - Price(USD)가 2개인 구형 XC 양식 대응
      예) QTY × 첫 번째 Price = 두 번째 Price 이면
          첫 번째 Price = 단가, 두 번째 Price = 금액합계로 자동 판정
    """
    best = (1, {}, -1)
    max_r = min(sheet.max_row, 80)
    for r in range(1, max_r + 1):
        header_texts = [clean_text(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1)]
        mapping: dict[str, object] = {}
        price_candidates: list[int] = []
        amount_candidates: list[int] = []
        qty_candidates: list[int] = []
        _detect_part_name_columns(header_texts, mapping)  # type: ignore[arg-type]

        # HZH Chinese/English legacy order form:
        # A=Model No., B=Chinese/Korean article name, C=English article name,
        # F=Qty, G=Price, H=Paid Total, I=Free Amount.
        header_norms = [_header_norm(value) for value in header_texts]
        has_hzh_layout = (
            any("MODELNO" in value or value == "型号MODELNO" for value in header_norms)
            and any("ARTICLE" in value or "名称" in header_texts[i] for i, value in enumerate(header_norms))
            and any(value.startswith("QTY") or "数量" in header_texts[i] for i, value in enumerate(header_norms))
            and any("PRICE" in value or "单价" in header_texts[i] for i, value in enumerate(header_norms))
        )
        if has_hzh_layout:
            mapping["model"] = 1
            mapping["part_name_kr"] = 2
            mapping["part_name_en"] = 3
            mapping["qty"] = 6
            mapping["price"] = 7
            mapping["qty_candidates"] = [6]
            mapping["price_candidates"] = [7]
            mapping["amount_candidates"] = [8, 9]
            mapping["amount"] = 8

        for idx, txt in enumerate(header_texts, start=1):
            upper = txt.upper()
            nt = _header_norm(txt)
            if upper in {"NO", "NO."} or nt == "NO":
                mapping.setdefault("no", idx)
            if _is_qty_header(upper, nt):
                qty_candidates.append(idx)
                mapping.setdefault("qty", idx)
            elif _is_amount_header(upper, nt):
                amount_candidates.append(idx)
                mapping.setdefault("amount", idx)
            elif _is_price_header(upper, nt):
                price_candidates.append(idx)
                mapping.setdefault("price", idx)
            if "CODE" in nt or "PARTNO" in nt or "품번" in txt:
                mapping["part_no"] = idx
            if "PARAM" in nt or "비고" in txt or "REMARK" in nt:
                mapping["parameter"] = idx
            if nt in {"MODEL", "적용모델", "모델"} or ("MODEL" in nt and "PART" not in nt):
                mapping["model"] = idx
        if price_candidates:
            mapping["price_candidates"] = price_candidates
        if amount_candidates:
            mapping["amount_candidates"] = amount_candidates
        if qty_candidates:
            mapping["qty_candidates"] = qty_candidates
        # Duplicate Price(USD) with no explicit Total is very common: first is unit price, second is amount.
        if len(price_candidates) >= 2 and "amount" not in mapping:
            mapping["amount"] = price_candidates[1]
        score = _looks_like_header_row(mapping)  # type: ignore[arg-type]
        if "part_name_en" in mapping:
            nonblank_names = 0
            numeric_qty = numeric_price = numeric_amount = 0
            for rr in range(r + 1, min(sheet.max_row, r + 10) + 1):
                if clean_text(sheet.cell(rr, int(mapping["part_name_en"])).value):
                    nonblank_names += 1
                for qc in qty_candidates[:2] or ([int(mapping["qty"])] if "qty" in mapping else []):
                    if parse_number(sheet.cell(rr, qc).value) is not None:
                        numeric_qty += 1; break
                for pc in price_candidates[:3] or ([int(mapping["price"])] if "price" in mapping else []):
                    if parse_number(sheet.cell(rr, pc).value) is not None:
                        numeric_price += 1; break
                for ac in amount_candidates[:3]:
                    if parse_number(sheet.cell(rr, ac).value) is not None:
                        numeric_amount += 1; break
            score += nonblank_names + min(numeric_qty, 3) + min(numeric_price, 3) + min(numeric_amount, 2)
        if score > best[2]:
            best = (r, mapping, score)
    if "part_name_en" not in best[1]:
        raise ValueError(f"Header row not detected in sheet: {sheet.title}")
    return best[0], best[1]


def _money_equal(left: float | None, right: float | None) -> bool:
    if left is None or right is None:
        return False
    tol = max(0.05, abs(right) * 0.01)
    return abs(left - right) <= tol


def _select_qty_price_amount(row_no: int, sheet: Worksheet, col: dict[str, object]) -> tuple[float | None, float | None, float | None, list[str]]:
    """Pick QTY / Unit Price / Amount using headers, hidden columns and row values.

    v2.0.5 핵심 보정:
    - 숨김 열도 전체 스캔한다. openpyxl은 hidden column도 읽으므로, 여기서는
      열 표시 여부와 무관하게 1~max_column 전체를 후보로 본다.
    - QTY(PCS) 헤더 아래라도 값이 $57.00처럼 통화이면 수량에서 제외한다.
    - IMAGE/빈 헤더 아래에 실제 수량이 들어간 구형 XC 양식은 가격열 앞의
      일반 정수값을 수량 후보로 승격한다.
    - QTY × 단가 ≒ 금액으로 검증된 경우만 단가/금액을 확정한다.
      검증 실패 시 잘못된 단가를 반영하지 않도록 보수적으로 처리한다.
    """
    notes: list[str] = []

    def cell_num(c: object) -> float | None:
        try:
            return parse_number(sheet.cell(row_no, int(c)).value)
        except Exception:
            return None

    qty_cols = [int(c) for c in (col.get("qty_candidates", []) or ([] if "qty" not in col else [col["qty"]]))]
    price_cols = [int(c) for c in (col.get("price_candidates", []) or ([] if "price" not in col else [col["price"]]))]
    amount_cols = [int(c) for c in (col.get("amount_candidates", []) or ([] if "amount" not in col else [col["amount"]]))]

    # 1) Header-based qty candidates, but reject currency-looking values.
    qty_vals: list[tuple[int, float]] = []
    rejected_qty_money: list[int] = []
    for c in qty_cols:
        num, is_money = _cell_num_and_money(sheet, row_no, c)
        if _valid_qty_value(num, is_money):
            qty_vals.append((c, float(round(num))))
        elif num is not None and is_money:
            rejected_qty_money.append(c)

    # 2) Add row-value-based qty candidates from all columns before the first money/price area.
    #    This covers files where IMAGE column contains 20/100/216 and QTY column contains $57.00.
    first_price_col = min(price_cols + amount_cols + rejected_qty_money) if (price_cols or amount_cols or rejected_qty_money) else sheet.max_column + 1
    excluded = {int(col.get(k, 0) or 0) for k in ("no", "part_name_en", "part_name_kr", "part_name_cn", "part_no", "parameter", "model")}
    for c in range(1, sheet.max_column + 1):
        if c in excluded or c in qty_cols:
            continue
        num, is_money = _cell_num_and_money(sheet, row_no, c)
        if not _valid_qty_value(num, is_money):
            continue
        # Prefer columns before the price/money area. Still allow nearby columns just before price.
        if c < first_price_col:
            qty_vals.append((c, float(round(num))))

    # Remove duplicates while preserving order. Prefer the rightmost plain integer before price,
    # because in request forms quantity is usually immediately before unit price.
    dedup_qty: dict[int, float] = {}
    for c, v in qty_vals:
        dedup_qty[c] = v
    qty_vals = sorted(dedup_qty.items(), key=lambda item: (abs(item[0] - first_price_col), -item[0]))
    q = qty_vals[0][1] if qty_vals else None
    if rejected_qty_money:
        notes.append(f"qty_currency_rejected_cols={rejected_qty_money}")
    if q is not None and qty_vals and qty_vals[0][0] not in qty_cols:
        notes.append(f"qty_inferred_from_col={qty_vals[0][0]}")

    # 3) Price/amount candidates: include all header price/amount columns and any currency-looking
    #    numeric cells in the row. This handles duplicated Price headers and hidden currency columns.
    candidate_cols = []
    for c in price_cols + amount_cols + rejected_qty_money:
        if c not in candidate_cols:
            candidate_cols.append(c)
    for c in range(1, sheet.max_column + 1):
        num, is_money = _cell_num_and_money(sheet, row_no, c)
        if num is not None and is_money and c not in candidate_cols:
            candidate_cols.append(c)

    # Build raw values. A money-looking cell under QTY becomes a price/amount candidate, not qty.
    price_vals: list[tuple[int, float]] = []
    amount_vals: list[tuple[int, float]] = []
    for c in candidate_cols:
        num, _ = _cell_num_and_money(sheet, row_no, c)
        if num is None:
            continue
        if c in amount_cols:
            amount_vals.append((c, num))
        elif c in price_cols or c in rejected_qty_money:
            price_vals.append((c, num))
        else:
            price_vals.append((c, num))

    # If an explicit amount column is absent, duplicate price/currency columns can play both roles.
    if not amount_vals and len(price_vals) >= 2:
        amount_vals = price_vals.copy()

    best: tuple[int, float | None, float | None, str] | None = None
    if q is not None and q != 0:
        for pc, pv in price_vals:
            for ac, av in amount_vals:
                if pc == ac:
                    continue
                if _money_equal(q * pv, av):
                    score = 100
                    if ac > pc:
                        score += 10
                    if av >= pv:
                        score += 5
                    if pc in rejected_qty_money:
                        score += 3  # QTY header with $ is often the hidden unit price column.
                    candidate = (score, pv, av, f"validated:qty*price=amount(q={q},pc={pc},ac={ac})")
                    if best is None or candidate[0] > best[0]:
                        best = candidate
    if best is not None:
        notes.append(best[3])
        return q, best[1], best[2], notes

    # 4) Conservative fallback. If a value cannot be validated, avoid polluting the Master.
    #    Only keep a unit price without amount when it is clearly a single currency price and qty is valid.
    if q is not None and len(price_vals) == 1 and not amount_vals:
        notes.append("price_without_total_single_currency")
        return q, price_vals[0][1], None, notes

    if q is not None and price_vals:
        notes.append("unvalidated_price_excluded")
        return q, None, None, notes

    # No reliable quantity. Do not treat currency as quantity.
    if price_vals:
        notes.append("no_reliable_qty_price_excluded")
    return q, None, None, notes


def _coerce_qty_price_amount(qty, price, amount) -> tuple[float | None, float | None, float | None, list[str]]:
    """Backward-compatible validator for simple 3-value cases."""
    q = parse_number(qty)
    p = parse_number(price)
    a = parse_number(amount)
    notes: list[str] = []
    if a is not None:
        if q is not None and p is None and q != 0:
            derived = a / q
            if derived > 0:
                p = round(derived, 4)
                notes.append("price=total/qty")
        elif p is not None and q is None and p != 0:
            derived_q = a / p
            if abs(derived_q - round(derived_q)) < 0.0001:
                q = round(derived_q)
                notes.append("qty=total/price")
        elif p is not None and q is not None and not _money_equal(p * q, a):
            notes.append("total_mismatch")
    return q, p, a, notes



def _legacy_price_request_columns(sheet: Worksheet) -> tuple[int, dict[str, int]] | None:
    """Detect legacy price-confirmation sheets shown in 2019–2021 files."""
    for row_no in range(1, min(sheet.max_row, 80) + 1):
        headers = {
            _header_norm(clean_text(sheet.cell(row_no, col).value)): col
            for col in range(1, sheet.max_column + 1)
            if clean_text(sheet.cell(row_no, col).value)
        }

        def find(*keys: str) -> int | None:
            normalized = {_header_norm(key) for key in keys}
            for key, col in headers.items():
                if key in normalized:
                    return col
            return None

        en = find("부품명(Eng)", "부품명(영문)", "영문부품명", "PARTNAMEENG")
        kr = find("부품명(Kor)", "부품명(한글)", "한글부품명", "PARTNAMEKOR")
        annual = find("1년사용량", "연간사용량", "ANNUALUSAGE")
        qty = find("요청수량", "요청 수량", "REQUESTQTY", "QTY")
        price = find("단가", "UNITPRICE", "PRICE")
        amount = find("합계", "총액", "AMOUNT", "TOTAL")

        if (en or kr) and qty and price:
            return row_no, {
                "part_name_en": en or 0,
                "part_name_kr": kr or 0,
                "annual_usage": annual or 0,
                "qty": qty,
                "price": price,
                "amount": amount or 0,
            }
    return None


def parse_legacy_price_request_sheet(
    sheet: Worksheet,
    source_path: str | Path,
) -> list[PartRecord]:
    detected = _legacy_price_request_columns(sheet)
    if not detected:
        return []

    header_row, columns = detected
    request_no, base_date, fallback_model = find_metadata(sheet, source_path)
    base_year = int(base_date[:4]) if base_date else infer_year_from_ir(request_no)
    today = date.today().isoformat()
    records: list[PartRecord] = []

    for row_no in range(header_row + 1, sheet.max_row + 1):
        def value(name: str):
            col = columns.get(name, 0)
            return sheet.cell(row_no, col).value if col else None

        en = clean_text(value("part_name_en"))
        kr = clean_text(value("part_name_kr"))
        if not en and not kr:
            continue
        if any(
            text.upper().startswith(("TOTAL", "합계"))
            for text in (en, kr)
            if text
        ):
            continue

        qty = parse_number(value("qty"))
        price = parse_money(value("price"))
        amount = parse_money(value("amount"))
        annual_usage = parse_number(value("annual_usage"))

        if qty is None or qty <= 0:
            continue

        # Zero in these old price-confirmation forms usually means "not quoted",
        # not a valid historical unit price.
        if price is not None and price <= 0:
            price = None
        if amount is not None and amount <= 0:
            amount = None

        if price is None and amount is not None and qty:
            derived_price = amount / qty
            price = round(derived_price, 6) if derived_price > 0 else None
        if amount is None and price is not None:
            amount = round(qty * price, 6)

        part_no = ""
        if en and re.fullmatch(
            r"[A-Z]{0,5}\d{3,}(?:[-_/]\d{2,})+",
            en.upper().strip(),
        ):
            part_no, en = en, ""

        # Model inference must happen after a part-code-only English cell has
        # been removed, otherwise C101110-0400 can be mistaken for model C1011.
        model = infer_model(en) or infer_model(kr) or fallback_model

        notes = [
            f"legacy_price_request_header={header_row}",
            f"cols={columns}",
        ]
        if annual_usage is not None:
            notes.append(f"annual_usage={annual_usage}")
        if price is not None and amount is not None:
            tolerance = max(0.05, abs(amount) * 0.01)
            if abs(qty * price - amount) <= tolerance:
                notes.append("validated:qty*price=amount")
            else:
                # A mismatching amount is not trusted. Keep the unit price only.
                notes.append("amount_mismatch_excluded")
                amount = None

        records.append(PartRecord(
            base_year=base_year,
            base_date=base_date,
            ir_no=request_no,
            model=norm_model(model),
            part_name_kr=kr,
            part_name_en=en,
            price=price,
            qty=qty,
            amount=amount,
            part_no=part_no,
            parameter=f"1년 사용량: {annual_usage:g}" if annual_usage is not None else "",
            source_file=Path(source_path).name,
            source_row=row_no,
            extraction_method=";".join(notes),
            currency="USD",
            updated=today,
        ))

    return records

def parse_sheet(sheet: Worksheet, source_path: str | Path) -> list[PartRecord]:
    header_row, col = locate_ir_columns(sheet)
    request_no, base_date, fallback_model = find_metadata(sheet, source_path)
    base_year = int(base_date[:4]) if base_date else infer_year_from_ir(request_no)
    today = date.today().isoformat()
    records: list[PartRecord] = []
    for row_no in range(header_row + 1, sheet.max_row + 1):
        def val(field: str) -> object:
            c = col.get(field)
            return sheet.cell(row_no, int(c)).value if c else None
        en = clean_text(val("part_name_en")); kr = clean_text(val("part_name_kr"))
        if not en and not kr:
            continue
        if en.upper().startswith("TOTAL") or kr.upper().startswith("TOTAL") or kr.startswith("합계"):
            continue
        qty, price, amount, notes = _select_qty_price_amount(row_no, sheet, col)

        # Legacy price-check sheets sometimes return only the supplier's unit
        # price while QTY and Total remain blank. This is still valid price
        # history and must not be discarded.
        if price is None and qty is None and amount is None:
            explicit_price = parse_money(val("price"))
            if explicit_price is not None and explicit_price > 0:
                price = explicit_price
                notes.append("price_only_legacy_row")
            else:
                continue
        # Applied model is intentionally derived from the English part name first.
        explicit_model = clean_text(val("model"))
        model = infer_model(en) or infer_model(kr) or norm_model(explicit_model) or fallback_model
        if _looks_like_component_spec(model):
            model = norm_model(explicit_model) or fallback_model
        parameter = clean_text(val("parameter"))
        amount_col = col.get("amount")
        if not parameter and amount_col:
            for c in range(int(amount_col) + 1, min(sheet.max_column, int(amount_col) + 4) + 1):
                text = clean_text(sheet.cell(row_no, c).value)
                if text and not parse_number(text):
                    parameter = text; break
        extraction = f"header:{header_row};cols={col}"
        if notes:
            extraction += ";" + ";".join(notes)
        records.append(PartRecord(
            base_year=base_year, base_date=base_date, ir_no=request_no, model=norm_model(model),
            part_name_kr=kr, part_name_en=en, part_name_cn=clean_text(val("part_name_cn")), price=price, qty=qty,
            amount=amount if amount is not None else ((price or 0) * (qty or 0) if price is not None and qty is not None else None),
            part_no=clean_text(val("part_no")), parameter=parameter, source_file=Path(source_path).name,
            source_row=row_no, extraction_method=extraction, currency="USD", updated=today,
        ))
    return records



def diagnose_workbook(path: str | Path) -> dict[str, object]:
    """Return a compact, user-readable parser diagnosis without modifying files."""
    result: dict[str, object] = {
        "file": Path(path).name,
        "manufacturer": infer_manufacturer(path) or "UNKNOWN",
        "request_no": infer_ir_no(path),
        "sheets": [],
        "total_rows": 0,
        "priced_rows": 0,
    }
    with open_workbook(path, data_only=True) as workbook:
        legacy_names = {
            name for name in workbook.sheetnames
            if _legacy_price_request_columns(workbook[name])
        }
        sheet_reports: list[dict[str, object]] = []

        for name in workbook.sheetnames:
            sheet = workbook[name]
            legacy = _legacy_price_request_columns(sheet)
            report: dict[str, object] = {
                "sheet": name,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "parser": "legacy_price_request" if legacy else "standard",
                "header_row": legacy[0] if legacy else None,
                "extracted_rows": 0,
                "priced_rows": 0,
                "status": "",
            }

            if legacy_names and name not in legacy_names:
                report["parser"] = "supplementary_excluded"
                report["status"] = "단가 Master 반영 제외"
                sheet_reports.append(report)
                continue

            try:
                rows = (
                    parse_legacy_price_request_sheet(sheet, path)
                    if legacy
                    else parse_sheet(sheet, path)
                )
                report["extracted_rows"] = len(rows)
                report["priced_rows"] = sum(row.price is not None for row in rows)
                report["status"] = "성공" if rows else "대상 행 없음"
            except Exception as exc:
                report["status"] = f"실패: {exc}"
            sheet_reports.append(report)

        result["sheets"] = sheet_reports

    rows = parse_workbook(path)
    result["total_rows"] = len(rows)
    result["priced_rows"] = sum(row.price is not None for row in rows)
    result["models"] = sorted({row.model for row in rows if row.model})
    return result


def parse_workbook(path: str | Path, sheet_names: Iterable[str] | None = None) -> list[PartRecord]:
    """Parse .xlsx/.xls with one common reader.

    Hidden sheets and hidden columns are intentionally included. A sheet is
    skipped only when a request table header cannot be detected. Legacy
    quantity-only FOC requests are retained with ``price=None`` rather than
    being discarded, so the Master can show the part and requested quantity
    while clearly leaving the unit price blank.
    """
    normalized_name = re.sub(r"[^A-Z0-9]+", "_", Path(path).name.upper())
    if "DEFECTIVE_SPARE_PARTS_REPORT" in normalized_name:
        raise ValueError(
            "불량 분석 Report 파일은 공급사 부품 발주 요청서가 아니므로 "
            "부품 단가 Master 반영 대상에서 제외합니다. 별도 불량분석 모듈이 필요합니다."
        )

    records: list[PartRecord] = []
    with open_workbook(path, data_only=True) as workbook:
        names = [
            name for name in list(sheet_names or workbook.sheetnames)
            if name in workbook.sheetnames
        ]

        legacy_names = [
            name for name in names
            if _legacy_price_request_columns(workbook[name])
        ]

        # When the workbook contains a dedicated price-confirmation sheet,
        # supplementary planning/stock sheets must not be imported as a second
        # quantity-only request. This prevents the supplied NC file from being
        # read as 61 priced rows plus 41 duplicate planning rows.
        parse_names = legacy_names if legacy_names else names

        for name in parse_names:
            sheet = workbook[name]
            if name in legacy_names:
                sheet_records = parse_legacy_price_request_sheet(sheet, path)
            else:
                try:
                    sheet_records = parse_sheet(sheet, path)
                except ValueError:
                    sheet_records = []
            records.extend(sheet_records)
    return records


def parse_master(path: str | Path) -> list[PartRecord]:
    workbook = load_workbook(path, data_only=True)
    if "원본상세" not in workbook.sheetnames:
        return []
    ws = workbook["원본상세"]
    headers = {clean_text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    def find(*names: str) -> int | None:
        for n in names:
            if n in headers:
                return headers[n]
        return None
    cols = {
        "base_year": find("기준년도"), "base_date": find("기준일자"),
        "ir_no": find("요청 No.", "IR No.", "IR No", "XC No.", "AC No.", "KA No.", "NC No."),
        "model": find("모델", "적용모델"), "part_name_kr": find("부품명(한글)"),
        "part_name_en": find("부품명(영어)"), "part_name_cn": find("부품명(중국어)"),
        "price": find("단가(USD)", "Price(USD)"), "qty": find("수량", "QTY"),
        "amount": find("금액(USD)", "Amount(USD)"), "part_no": find("품번/코드", "품번"),
        "parameter": find("파라미터/비고", "비고"), "source_file": find("원본파일"),
        "source_row": find("원본행"), "extraction_method": find("추출방식"),
    }
    records: list[PartRecord] = []
    for r in range(2, ws.max_row + 1):
        def v(k: str):
            c = cols.get(k); return ws.cell(r, c).value if c else None
        en = clean_text(v("part_name_en")); kr = clean_text(v("part_name_kr"))
        if not en and not kr:
            continue
        sf = clean_text(v("source_file")); req = clean_text(v("ir_no")) or infer_ir_no(sf)
        bd = clean_text(v("base_date")) or (infer_date_from_filename(sf) if sf else "")
        by = parse_number(v("base_year"))
        if by is None and bd[:4].isdigit(): by = float(bd[:4])
        if by is None:
            y = infer_year_from_ir(req); by = float(y) if y else None
        mdl = infer_model(en) or norm_model(v("model")) or infer_model(kr) or infer_model(sf)
        records.append(PartRecord(
            base_year=int(by) if by is not None else None, base_date=bd, ir_no=req, model=norm_model(mdl),
            part_name_kr=kr, part_name_en=en, part_name_cn=clean_text(v("part_name_cn")),
            price=parse_number(v("price")), qty=parse_number(v("qty")), amount=parse_number(v("amount")),
            part_no=clean_text(v("part_no")), parameter=clean_text(v("parameter")), source_file=sf,
            source_row=int(parse_number(v("source_row")) or 0) or None,
            extraction_method=clean_text(v("extraction_method")),
        ))
    return records
