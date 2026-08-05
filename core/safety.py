from __future__ import annotations

import hashlib
import json
import shutil
import tempfile
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


IGNORED_ARCHIVE_FILES = {".keep", "thumbs.db", "desktop.ini", ".ds_store"}


@dataclass
class FileSnapshot:
    path: str
    size: int
    sha256: str


@dataclass
class MasterSnapshot:
    manufacturer: str
    path: str
    exists: bool
    size: int = 0
    sha256: str = ""
    sheets: dict[str, int] = field(default_factory=dict)
    error: str = ""


@dataclass
class SafetyReport:
    created_at: str
    passed: bool
    score: int
    master_snapshots: list[MasterSnapshot]
    archive_errors: list[str] = field(default_factory=list)
    ignored_archive_files: list[str] = field(default_factory=list)
    checks: dict[str, str] = field(default_factory=dict)
    notes: list[str] = field(default_factory=list)

    def to_json(self, path: Path) -> Path:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(asdict(self), ensure_ascii=False, indent=2), encoding="utf-8")
        return path

    @property
    def summary(self) -> str:
        return f"{'PASS' if self.passed else 'FAIL'} · {self.score}점"


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fp:
        while chunk := fp.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def snapshot_file(path: Path, root: Path | None = None) -> FileSnapshot:
    shown = str(path.relative_to(root)) if root and path.is_relative_to(root) else str(path)
    return FileSnapshot(shown, path.stat().st_size, sha256_file(path))


def snapshot_master(path: Path, manufacturer: str) -> MasterSnapshot:
    if not path.exists():
        return MasterSnapshot(manufacturer, str(path), False)
    sheets: dict[str, int] = {}
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        try:
            for ws in wb.worksheets:
                sheets[ws.title] = ws.max_row
        finally:
            wb.close()
        return MasterSnapshot(
            manufacturer=manufacturer,
            path=str(path),
            exists=True,
            size=path.stat().st_size,
            sha256=sha256_file(path),
            sheets=sheets,
        )
    except Exception as exc:
        return MasterSnapshot(
            manufacturer=manufacturer,
            path=str(path),
            exists=True,
            size=path.stat().st_size,
            error=str(exc),
        )


def _is_ignored_archive_file(path: Path) -> bool:
    name = path.name.lower()
    return name in IGNORED_ARCHIVE_FILES or name.startswith("~$")


def validate_archive(archive_root: Path) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    ignored: list[str] = []
    if not archive_root.exists():
        return errors, ignored
    for path in archive_root.rglob("*"):
        if not path.is_file():
            continue
        if _is_ignored_archive_file(path):
            ignored.append(str(path))
            continue
        try:
            if path.stat().st_size <= 0:
                errors.append(f"빈 파일: {path}")
            else:
                sha256_file(path)
        except OSError as exc:
            errors.append(f"읽기 실패: {path} ({exc})")
    return errors, ignored


def _recent_file_exists(folder: Path) -> bool:
    return folder.exists() and any(p.is_file() and not _is_ignored_archive_file(p) for p in folder.rglob("*"))


def build_safety_report(root: Path, manufacturers: Iterable[str]) -> SafetyReport:
    snapshots = [
        snapshot_master(root / "Manufacturers" / m / "Master" / "Master.xlsx", m)
        for m in manufacturers
    ]
    archive_errors, ignored_archive_files = validate_archive(root / "Archive")
    missing = [s.manufacturer for s in snapshots if not s.exists]
    unreadable = [s.manufacturer for s in snapshots if s.exists and s.error]

    checks = {
        "Master": "PASS" if not missing and not unreadable else "FAIL",
        "Archive": "PASS" if not archive_errors else "FAIL",
        "Backup": "PASS" if _recent_file_exists(root / "Backup") else "INFO",
        "History": "PASS" if _recent_file_exists(root / "History") else "INFO",
        "Log": "PASS" if (root / "Log").exists() else "INFO",
    }

    notes: list[str] = []
    if missing:
        notes.append("Master 없음: " + ", ".join(missing))
    if unreadable:
        notes.append("Master 읽기 실패: " + ", ".join(unreadable))
    if ignored_archive_files:
        notes.append(f"Archive 검사 제외 시스템 파일: {len(ignored_archive_files)}개")
    if checks["Backup"] == "INFO":
        notes.append("Backup 파일을 찾지 못했습니다. 첫 업데이트 전이라면 정상일 수 있습니다.")
    if checks["History"] == "INFO":
        notes.append("History 파일을 찾지 못했습니다. 아직 이력이 없다면 정상일 수 있습니다.")

    critical_failures = sum(1 for key in ("Master", "Archive") if checks[key] == "FAIL")
    score = max(0, 100 - critical_failures * 40)
    if checks["Backup"] == "INFO":
        score -= 5
    if checks["History"] == "INFO":
        score -= 5
    score = max(0, score)

    return SafetyReport(
        created_at=datetime.now().isoformat(timespec="seconds"),
        passed=critical_failures == 0,
        score=score,
        master_snapshots=snapshots,
        archive_errors=archive_errors,
        ignored_archive_files=ignored_archive_files,
        checks=checks,
        notes=notes,
    )


def create_simulation_workspace(root: Path) -> Path:
    """Create an isolated copy containing only runtime-critical folders/files."""
    temp_root = Path(tempfile.mkdtemp(prefix="cpms_sim_"))
    include = ["Config", "Manufacturers", "Update", "Archive", "History", "PDF_Review"]
    for name in include:
        src = root / name
        dst = temp_root / name
        if src.exists():
            shutil.copytree(src, dst)
        else:
            dst.mkdir(parents=True, exist_ok=True)
    for name in ["Backup", "Log", "Rejected", "database", "Purchase_Analysis"]:
        (temp_root / name).mkdir(parents=True, exist_ok=True)
    return temp_root
