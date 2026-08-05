from __future__ import annotations
import csv, json, re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from core.pdf_approval import process_approved_pdf_reviews

VALID_MFR={"IR","XC","AC","KA","NC"}
SAFE_VALID={"정상","금액계산"}

@dataclass
class Decision:
    review_file: Path
    request_no: str=""
    manufacturer: str=""
    row_count: int=0
    min_confidence: float=0.0
    score: int=0
    grade: str=""
    decision: str=""
    reason: str=""

@dataclass
class RunResult:
    scanned_files:int=0
    auto_approved_files:int=0
    recommended_files:int=0
    manual_review_files:int=0
    failed_files:int=0
    decisions:list[Decision]=field(default_factory=list)
    decision_log:Path|None=None
    approval_log:Path|None=None

def clean(v): return str(v or "").strip()

def num(v):
    if v in (None,""): return None
    if isinstance(v,(int,float)): return float(v)
    t=re.sub(r"[^0-9.\-]","",str(v))
    try: return float(t) if t not in {"",".","-","-."} else None
    except ValueError: return None

def threshold(root):
    p=root/"Config"/"config.json"
    try:
        with p.open("r",encoding="utf-8") as f: c=json.load(f)
        return min(100,max(0,float(c.get("pdf_auto_approve_threshold",95))))
    except Exception: return 95.0

def summary(wb):
    if "검증요약" not in wb.sheetnames: return {}
    ws=wb["검증요약"]
    return {clean(ws.cell(r,1).value):ws.cell(r,2).value for r in range(1,ws.max_row+1) if clean(ws.cell(r,1).value)}

def part_model(name):
    m=re.search(r"(?:CMC|CMS|HM)-([A-Z]?\d{2,4}[A-Z]?)(?:\(G\))?",name,re.I)
    if m: return m.group(1).upper()
    for token in reversed([x.strip() for x in name.split(",") if x.strip()]):
        m=re.fullmatch(r"([A-Z]{0,3}\d{2,4}[A-Z]?)(?:\(G\))?",token,re.I)
        if m and not re.search(r"ZYT|RPM|HZ|\dV|\dW",token,re.I): return m.group(1).upper()
    return ""

def non_part(name):
    return bool(
        re.search(
            r"\b(TOTAL|SUBTOTAL|DISCOUNT|FREIGHT|SHIPPING|PAYMENT|DELIVERY|"
            r"REMARKS?|ORDER\s*NO|ORDER\s+CONFIRMATION|SUPPLIER|DATE|ETD)\b",
            name,
            re.I,
        )
        or any(x in name for x in ("합계", "소계", "할인"))
    )


def request_manufacturer(request_no):
    match = re.match(r"\s*(IR|XC|AC|KA|NC)", clean(request_no), re.I)
    return match.group(1).upper() if match else ""


def valid_model(model):
    value = clean(model).upper()
    if not value or value in {"0", "00", "000", "UNKNOWN", "N/A"}:
        return False
    return bool(re.fullmatch(r"[A-Z]{0,3}\d{3,4}[A-Z]?", value, re.I))

def grade(score,hard):
    if hard:return "D"
    if score>=95:return "A"
    if score>=80:return "B"
    if score>=60:return "C"
    return "D"

def write_grade(wb,g,s,reason):
    ws=wb["자동판정"] if "자동판정" in wb.sheetnames else wb.create_sheet("자동판정",0)
    data=[("자동승인 등급",g),("점수",s),("판정사유",reason),("기준","A=자동승인 / B=승인추천 / C=수동검토 / D=반영금지")]
    for r,(a,b) in enumerate(data,1):
        ws.cell(r,1,a).font=Font(bold=True)
        ws.cell(r,2,b)
    ws["B1"].fill=PatternFill("solid",fgColor={"A":"C6E0B4","B":"FFF2CC","C":"FCE4D6","D":"F4CCCC"}[g])
    ws["B1"].font=Font(bold=True,size=16)
    ws.column_dimensions["A"].width=20; ws.column_dimensions["B"].width=90

def evaluate(path,th):
    d=Decision(review_file=path)
    wb=load_workbook(path)
    if "PDF_검토대기" not in wb.sheetnames:
        d.grade="D";d.decision="반영금지";d.reason="PDF_검토대기 시트 없음";return d
    ws=wb["PDF_검토대기"]
    h={clean(ws.cell(1,c).value):c for c in range(1,ws.max_column+1) if clean(ws.cell(1,c).value)}
    req=["검토상태","요청 No.","제조사","적용모델","부품명(영어)","수량","단가(USD)","금액(USD)","신뢰도","행검증"]
    miss=[x for x in req if x not in h]
    if miss:
        d.grade="D";d.decision="반영금지";d.reason="필수 열 누락: "+", ".join(miss);write_grade(wb,"D",0,d.reason);wb.save(path);return d
    existing=[clean(ws.cell(r,h["적용모델"]).value) for r in range(2,ws.max_row+1) if clean(ws.cell(r,h["적용모델"]).value)]
    default=max(set(existing),key=existing.count) if existing else ""
    rows=[];confs=[];reqs=set();mfrs=set();hard=[];ded=[];score=100
    for r in range(2,ws.max_row+1):
        part=clean(ws.cell(r,h["부품명(영어)"]).value)
        if not part or non_part(part): continue
        rows.append(r)
        rq=clean(ws.cell(r,h["요청 No."]).value); mf=clean(ws.cell(r,h["제조사"]).value).upper()
        model=clean(ws.cell(r,h["적용모델"]).value); qty=num(ws.cell(r,h["수량"]).value)
        price=num(ws.cell(r,h["단가(USD)"]).value); amount=num(ws.cell(r,h["금액(USD)"]).value)
        conf=num(ws.cell(r,h["신뢰도"]).value); val=clean(ws.cell(r,h["행검증"]).value)
        if rq:reqs.add(rq)
        if mf:mfrs.add(mf)
        if conf is not None:confs.append(conf)
        if not model:
            model=part_model(part) or default
            if model:
                ws.cell(r,h["적용모델"],model);score-=1;ded.append(f"{r}행 모델 자동보완")
            else: score-=4;ded.append(f"{r}행 적용모델 누락")
        if not rq:
            hard.append(f"{r}행 요청번호 누락")
        if mf not in VALID_MFR:
            hard.append(f"{r}행 제조사 오류")
        elif request_manufacturer(rq) and request_manufacturer(rq) != mf:
            hard.append(f"{r}행 요청번호 제조사({request_manufacturer(rq)})와 제조사({mf}) 불일치")
        if non_part(part):
            hard.append(f"{r}행 문서 제목/합계 행을 부품으로 인식")
        if not valid_model(model):
            hard.append(f"{r}행 적용모델 오류({model or '없음'})")
        if qty is None or qty <= 0:
            hard.append(f"{r}행 수량 오류")
        if price is None or price <= 0:
            hard.append(f"{r}행 단가 오류")
        if amount is None or amount <= 0:
            hard.append(f"{r}행 금액 오류")
        if qty is not None and price is not None and amount is not None:
            tolerance=max(0.1,abs(amount)*0.015)
            if abs(qty*price-amount)>tolerance:
                hard.append(f"{r}행 수량×단가와 금액 불일치")
        if conf is None:score-=5;ded.append(f"{r}행 신뢰도 없음")
        elif conf<th:score-=max(1,round((th-conf)/2));ded.append(f"{r}행 신뢰도 {conf:.1f}%")
        if val=="금액불일치":
            if amount in (0,0.0):
                ws.cell(r,h["행검증"],"정상");score-=1;ded.append(f"{r}행 FOC 금액 0 인정")
            else:score-=8;ded.append(f"{r}행 금액불일치")
        elif val not in SAFE_VALID:score-=3;ded.append(f"{r}행 검증상태 {val or '없음'}")
    if not rows:
        hard.append("추출된 부품 행 없음")
        # OCR만 성공하고 부품이 한 행도 추출되지 않은 파일은 100점으로 보이면 안 됩니다.
        score = 0
    if len(reqs)!=1:hard.append("요청번호 불일치")
    if len(mfrs)!=1:hard.append("제조사 불일치")
    s=summary(wb)
    if s.get("PDF 표기 총수량") not in (None,"") and clean(s.get("수량 합계 일치"))!="정상":
        score-=6;ded.append("PDF 총수량과 추출 합계 불일치")
    if s.get("PDF 표기 총금액") not in (None,"") and clean(s.get("금액 합계 일치"))!="정상":
        score-=10;ded.append("PDF 총금액과 추출 합계 불일치")
    score=max(0,min(100,score))
    # 치명적 검증 오류가 있으면 점수가 높더라도 승인 가능한 점수대로 표시하지 않습니다.
    if hard:
        score=min(score,59)
    g=grade(score,bool(hard))
    d.request_no=next(iter(reqs),"");d.manufacturer=next(iter(mfrs),"");d.row_count=len(rows);d.min_confidence=min(confs) if confs else 0
    d.score=score;d.grade=g;d.reason="; ".join(dict.fromkeys(hard+ded)) or "모든 검증 통과"
    if g=="A":
        for r in rows:ws.cell(r,h["검토상태"],"승인")
        d.decision="자동승인"
    elif g=="B":d.decision="승인추천"
    elif g=="C":d.decision="수동검토"
    else:d.decision="반영금지"
    write_grade(wb,g,score,d.reason);wb.save(path);return d

def run_pdf_auto_grading(root="."):
    root=Path(root);review=root/"PDF_Review";review.mkdir(exist_ok=True)
    files=sorted(review.glob("PDF_검토대기_*.xlsx"));th=threshold(root);res=RunResult(scanned_files=len(files))
    for p in files:
        try:d=evaluate(p,th)
        except Exception as e:d=Decision(p,grade="D",decision="오류",reason=str(e));res.failed_files+=1
        if d.decision=="자동승인":res.auto_approved_files+=1
        elif d.decision=="승인추천":res.recommended_files+=1
        elif d.decision!="오류":res.manual_review_files+=1
        res.decisions.append(d)
    log=root/"Log"/f"pdf_grade_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv";log.parent.mkdir(exist_ok=True)
    with log.open("w",newline="",encoding="utf-8-sig") as f:
        w=csv.writer(f);w.writerow(["file","request_no","manufacturer","rows","min_confidence","score","grade","decision","reason"])
        for d in res.decisions:w.writerow([d.review_file.name,d.request_no,d.manufacturer,d.row_count,d.min_confidence,d.score,d.grade,d.decision,d.reason])
    res.decision_log=log
    if res.auto_approved_files:
        a=process_approved_pdf_reviews(root);res.approval_log=a.log_path
    return res
