from __future__ import annotations

import json
import re
from dataclasses import dataclass, asdict, field
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook

ProgressCallback = Callable[[int, str], None]
MANUFACTURERS = ("IR", "XC", "AC", "KA", "NC")
EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}

HEADER_ALIASES = {
    "part_no": ("부품코드", "부품 코드", "part no", "part no.", "part number", "품번"),
    "part_kr": ("부품명", "부품명(kr)", "부품명(kor)", "한글부품명"),
    "part_en": ("부품명(eng)", "부품명(eng.)", "부품명(eg)", "부품명(en)", "part name", "part name(en)"),
    "price": ("단가", "부품원가", "unit price", "price", "usd"),
    "qty": ("발주수량", "신규 발주수량", "신규 부품발주 수량(ea)", "필요수량", "요청수량", "수량", "qty"),
    "amount": ("합계", "금액", "amount", "total"),
}

@dataclass
class SheetAnalysis:
    file: str
    sheet: str
    manufacturer: str = ""
    document_type: str = "검토필요"
    header_row: int = 0
    rows_estimated: int = 0
    mapped_fields: list[str] = field(default_factory=list)
    safe_action: str = "Master 반영 금지"
    reason: str = ""

@dataclass
class MultiDocReport:
    created_at: str
    root: str
    files_scanned: int
    sheets_scanned: int
    safe_import_sheets: int
    demand_only_sheets: int
    blocked_sheets: int
    items: list[SheetAnalysis]
    json_path: str = ""
    excel_path: str = ""

    def to_dict(self) -> dict:
        data = asdict(self)
        data["items"] = [asdict(x) for x in self.items]
        return data


def _norm(value: object) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text.strip()).lower()


def _manufacturer_from_sheet(sheet_name: str) -> str:
    normalized = re.sub(r"[^A-Z0-9]", "", sheet_name.upper())
    for m in MANUFACTURERS:
        if normalized == m or normalized.startswith(m):
            return m
    return ""


def _find_header(ws, max_rows: int = 40, max_cols: int = 40) -> tuple[int, dict[str, int]]:
    best_row, best_map, best_score = 0, {}, 0
    for r in range(1, min(ws.max_row, max_rows) + 1):
        values = [_norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column, max_cols) + 1)]
        mapping: dict[str, int] = {}
        for field_name, aliases in HEADER_ALIASES.items():
            for c, value in enumerate(values, start=1):
                if not value:
                    continue
                if any(value == alias or alias in value for alias in aliases):
                    # '부품명' must not steal English-specific header.
                    if field_name == "part_kr" and any(token in value for token in ("eng", "(eg)", "(en)")):
                        continue
                    mapping[field_name] = c
                    break
        score = len(mapping)
        if score > best_score:
            best_row, best_map, best_score = r, mapping, score
    return best_row, best_map


def _classify(file_name: str, sheet_name: str, mapping: dict[str, int], manufacturer: str) -> tuple[str, str, str]:
    text = f"{file_name} {sheet_name}".lower()
    fields = set(mapping)
    has_part = bool(fields & {"part_no", "part_kr", "part_en"})
    has_price = "price" in fields
    has_qty = "qty" in fields

    if any(word in text for word in ("원시 as", "raw", "부품등록", "dashboard", "종합")):
        return "참고/원시데이터", "Master 반영 금지", "원시·등록·요약 시트는 단가 원본으로 사용할 수 없음"
    if any(word in text for word in ("필요 부품", "취합", "필요수량")) or (has_part and has_qty and not has_price):
        return "수요/필요부품", "수요 이력만 저장", "단가가 없어 Master 단가 반영 불가"
    if manufacturer and has_part and has_price:
        return "제조사 발주/회신", "검토 후 단가 반영 가능", "제조사·부품·단가 필드 확인"
    if has_part and has_price:
        return "품의/통합 발주", "시트 제조사 확인 후 반영", "파일 단위 제조사 판정 금지; 시트별 확인 필요"
    return "검토필요", "Master 반영 금지", "필수 헤더(부품/단가 또는 수량) 부족"


def analyze_multidoc_folder(root: str | Path, progress: ProgressCallback | None = None) -> MultiDocReport:
    root = Path(root)
    update_dir = root / "Update"
    log_dir = root / "Log"
    log_dir.mkdir(parents=True, exist_ok=True)
    files = sorted(p for p in update_dir.iterdir() if p.is_file() and p.suffix.lower() in EXCEL_SUFFIXES and not p.name.startswith("~$")) if update_dir.exists() else []
    items: list[SheetAnalysis] = []

    for index, path in enumerate(files, start=1):
        if progress:
            progress(max(1, int((index - 1) / max(len(files), 1) * 85)), f"다중문서 분석 {index}/{len(files)}: {path.name}")
        try:
            wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
            for ws in wb.worksheets:
                header_row, mapping = _find_header(ws)
                manufacturer = _manufacturer_from_sheet(ws.title)
                doc_type, action, reason = _classify(path.name, ws.title, mapping, manufacturer)
                items.append(SheetAnalysis(
                    file=path.name,
                    sheet=ws.title,
                    manufacturer=manufacturer,
                    document_type=doc_type,
                    header_row=header_row,
                    rows_estimated=max(ws.max_row - header_row, 0) if header_row else 0,
                    mapped_fields=sorted(mapping),
                    safe_action=action,
                    reason=reason,
                ))
            wb.close()
        except Exception as exc:
            items.append(SheetAnalysis(file=path.name, sheet="", document_type="읽기실패", safe_action="Master 반영 금지", reason=str(exc)))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report = MultiDocReport(
        created_at=datetime.now().isoformat(timespec="seconds"),
        root=str(root),
        files_scanned=len(files),
        sheets_scanned=len(items),
        safe_import_sheets=sum(x.safe_action == "검토 후 단가 반영 가능" for x in items),
        demand_only_sheets=sum(x.safe_action == "수요 이력만 저장" for x in items),
        blocked_sheets=sum("금지" in x.safe_action for x in items),
        items=items,
    )
    json_path = log_dir / f"MultiDoc_Analysis_{timestamp}.json"
    excel_path = log_dir / f"MultiDoc_Analysis_{timestamp}.xlsx"
    report.json_path = str(json_path)
    report.excel_path = str(excel_path)
    json_path.write_text(json.dumps(report.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = "분석요약"
    ws.append(["생성일", report.created_at])
    ws.append(["분석파일", report.files_scanned])
    ws.append(["분석시트", report.sheets_scanned])
    ws.append(["검토 후 단가반영 가능", report.safe_import_sheets])
    ws.append(["수요이력 전용", report.demand_only_sheets])
    ws.append(["반영금지", report.blocked_sheets])
    detail = wb_out.create_sheet("시트별_진단")
    detail.append(["파일", "시트", "제조사", "문서유형", "헤더행", "추정행수", "인식필드", "안전조치", "사유"])
    for item in items:
        detail.append([item.file, item.sheet, item.manufacturer, item.document_type, item.header_row, item.rows_estimated, ", ".join(item.mapped_fields), item.safe_action, item.reason])
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    widths = {"A": 45, "B": 24, "C": 10, "D": 20, "E": 10, "F": 12, "G": 32, "H": 24, "I": 50}
    for col, width in widths.items():
        detail.column_dimensions[col].width = width
    wb_out.save(excel_path)
    if progress:
        progress(100, f"다중문서 안전 분석 완료: {excel_path.name}")
    return report
