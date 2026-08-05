from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class HistoryRow:
    category: str
    date: str
    manufacturer: str
    request_no: str
    model: str
    part_no: str
    part_name_en: str
    part_name_kr: str
    previous_value: str
    latest_value: str
    difference: str
    change_rate: str
    change_type: str
    field_name: str
    source_file: str
    currency: str

    def searchable_text(self) -> str:
        return " ".join(
            [
                self.category, self.date, self.manufacturer, self.request_no,
                self.model, self.part_no, self.part_name_en, self.part_name_kr,
                self.previous_value, self.latest_value, self.difference,
                self.change_rate, self.change_type, self.field_name,
                self.source_file, self.currency,
            ]
        ).casefold()


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.10g}"
    return str(value).strip()


def _percent(value: object) -> str:
    if value in (None, ""):
        return ""
    try:
        return f"{float(value):.1f}%"
    except (TypeError, ValueError):
        return _text(value)


def _iter_dict_rows(ws):
    """Yield worksheet rows as dictionaries in a single sequential pass.

    This is critical for openpyxl read_only mode. Calling ws.cell(row, col)
    repeatedly on a ReadOnlyWorksheet reparses XML from the beginning and turns a
    few thousand rows into an O(n²) operation. iter_rows(values_only=True) reads
    the XML once and keeps History loading fast and stable.
    """
    iterator = ws.iter_rows(values_only=True)
    try:
        header_values = next(iterator)
    except StopIteration:
        return

    headers = [_text(value) for value in header_values]
    for values in iterator:
        row = {
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }
        yield row


def _load_price_history(ws) -> list[HistoryRow]:
    result: list[HistoryRow] = []
    for data in _iter_dict_rows(ws):
        request_no = _text(data.get("최근요청번호"))
        source_file = _text(data.get("최근원본파일"))
        part_name_en = _text(data.get("부품명(영문)"))
        part_name_kr = _text(data.get("부품명(한글)"))
        if not any([request_no, source_file, part_name_en, part_name_kr]):
            continue

        result.append(
            HistoryRow(
                category="가격이력",
                date=_text(data.get("최근기준일")),
                manufacturer=_text(data.get("제조사")),
                request_no=request_no,
                model=_text(data.get("모델")),
                part_no=_text(data.get("부품번호")),
                part_name_en=part_name_en,
                part_name_kr=part_name_kr,
                previous_value=_text(data.get("이전단가")),
                latest_value=_text(data.get("최근단가")),
                difference=_text(data.get("변동액")),
                change_rate=_percent(data.get("변동률(%)")),
                change_type="가격변경",
                field_name="단가",
                source_file=source_file,
                currency=_text(data.get("통화")),
            )
        )
    return result


def _load_change_history(ws) -> list[HistoryRow]:
    result: list[HistoryRow] = []
    for data in _iter_dict_rows(ws):
        request_no = _text(data.get("요청번호"))
        source_file = _text(data.get("원본파일"))
        change_type = _text(data.get("변경유형"))
        field_name = _text(data.get("변경항목"))
        if not any([request_no, source_file, change_type, field_name]):
            continue

        result.append(
            HistoryRow(
                category="변경이력",
                date=_text(data.get("기록일시")),
                manufacturer=_text(data.get("제조사")),
                request_no=request_no,
                model=_text(data.get("모델")),
                part_no=_text(data.get("부품번호")),
                part_name_en=_text(data.get("부품명(영문)")),
                part_name_kr=_text(data.get("부품명(한글)")),
                previous_value=_text(data.get("이전값")),
                latest_value=_text(data.get("변경값")),
                difference="",
                change_rate="",
                change_type=change_type,
                field_name=field_name,
                source_file=source_file,
                currency=_text(data.get("통화")),
            )
        )
    return result


def available_history_files(history_dir: str | Path) -> list[Path]:
    history_dir = Path(history_dir)
    if not history_dir.exists():
        return []
    return sorted(
        history_dir.glob("History_*.xlsx"),
        key=lambda path: path.name,
        reverse=True,
    )


def load_history_rows(history_dir: str | Path) -> list[HistoryRow]:
    rows: list[HistoryRow] = []
    for path in available_history_files(history_dir):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if "가격이력" in workbook.sheetnames:
                rows.extend(_load_price_history(workbook["가격이력"]))
            if "변경이력" in workbook.sheetnames:
                rows.extend(_load_change_history(workbook["변경이력"]))
        finally:
            workbook.close()
    return rows


def filter_history_rows(
    rows: Iterable[HistoryRow],
    *,
    keyword: str = "",
    manufacturer: str = "전체",
    category: str = "전체",
) -> list[HistoryRow]:
    keyword = keyword.strip().casefold()
    result: list[HistoryRow] = []
    for row in rows:
        if manufacturer != "전체" and row.manufacturer != manufacturer:
            continue
        if category != "전체" and row.category != category:
            continue
        if keyword and keyword not in row.searchable_text():
            continue
        result.append(row)
    return result
