from __future__ import annotations

import csv
import json
import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from core.backup import backup_master
from core.compare import build_kr_lookup, complete_korean_names, merge_records
from core.parser import PartRecord, infer_year_from_ir, parse_master
from core.report_builder import rebuild_all_sheets


APPROVED_STATUSES = {"승인", "확인완료", "APPROVED", "Y", "YES"}
FORCE_APPROVED_STATUSES = {"강제승인", "FORCE APPROVED"}


@dataclass
class ApprovalFileResult:
    review_file: Path
    request_no: str = ""
    manufacturer: str = ""
    row_count: int = 0
    new_parts: int = 0
    updated_rows: int = 0
    price_changes: int = 0
    status: str = ""
    master_path: Path | None = None
    archived_review_path: Path | None = None
    archived_pdf_path: Path | None = None
    errors: list[str] = field(default_factory=list)


@dataclass
class ApprovalRunResult:
    scanned_files: int = 0
    approved_files: int = 0
    waiting_files: int = 0
    failed_files: int = 0
    total_rows: int = 0
    total_new_parts: int = 0
    total_updated_rows: int = 0
    total_price_changes: int = 0
    history_path: Path | None = None
    log_path: Path | None = None
    files: list[ApprovalFileResult] = field(default_factory=list)


def _clean(value) -> str:
    return str(value or "").strip()


def _number(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    import re
    text = re.sub(r"[^0-9.\-]", "", str(value))
    if text in {"", ".", "-", "-."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _request_manufacturer(request_no: str) -> str:
    match = re.match(r"\s*(IR|XC|AC|KA|NC)", _clean(request_no), re.I)
    return match.group(1).upper() if match else ""


def _is_non_part_name(name: str) -> bool:
    return bool(re.search(
        r"\b(TOTAL|SUBTOTAL|ORDER\s*NO|ORDER\s+CONFIRMATION|SUPPLIER|DATE|ETD|REMARKS?)\b",
        _clean(name),
        re.I,
    ))


def _valid_model(model: str) -> bool:
    value = _clean(model).upper()
    if value in {"", "0", "00", "000", "UNKNOWN", "N/A"}:
        return False
    return bool(re.fullmatch(r"[A-Z]{0,3}\d{3,4}[A-Z]?", value, re.I))


def _load_config(root: Path) -> dict:
    path = root / "Config" / "config.json"
    with path.open("r", encoding="utf-8") as fp:
        return json.load(fp)


def _master_path(root: Path, config: dict, manufacturer: str) -> Path:
    manufacturer_root = root / config.get("manufacturer_root", "Manufacturers")
    return manufacturer_root / manufacturer / "Master" / "Master.xlsx"


def _review_rows(path: Path) -> tuple[list[PartRecord], dict]:
    wb = load_workbook(path, data_only=True)
    if "PDF_검토대기" not in wb.sheetnames:
        raise ValueError("PDF_검토대기 시트를 찾을 수 없습니다.")

    ws = wb["PDF_검토대기"]
    headers = {
        _clean(ws.cell(1, column).value): column
        for column in range(1, ws.max_column + 1)
        if _clean(ws.cell(1, column).value)
    }
    required = ["검토상태", "요청 No.", "제조사", "적용모델", "부품명(영어)", "수량", "단가(USD)", "금액(USD)", "행검증", "원본PDF"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"필수 열 누락: {', '.join(missing)}")

    source_pdf = ""
    request_no = ""
    manufacturer = ""
    records: list[PartRecord] = []
    waiting_rows: list[int] = []
    invalid_rows: list[str] = []

    for row_no in range(2, ws.max_row + 1):
        part_name = _clean(ws.cell(row_no, headers["부품명(영어)"]).value)
        if not part_name:
            continue

        status = _clean(ws.cell(row_no, headers["검토상태"]).value).upper()
        if status not in APPROVED_STATUSES and status not in FORCE_APPROVED_STATUSES:
            waiting_rows.append(row_no)
            continue

        current_request = _clean(ws.cell(row_no, headers["요청 No."]).value)
        current_manufacturer = _clean(ws.cell(row_no, headers["제조사"]).value).upper()
        model = _clean(ws.cell(row_no, headers["적용모델"]).value)
        qty = _number(ws.cell(row_no, headers["수량"]).value)
        price = _number(ws.cell(row_no, headers["단가(USD)"]).value)
        amount = _number(ws.cell(row_no, headers["금액(USD)"]).value)
        validation = _clean(ws.cell(row_no, headers["행검증"]).value)
        current_pdf = _clean(ws.cell(row_no, headers["원본PDF"]).value)

        if not current_request or current_manufacturer not in {"IR", "XC", "AC", "KA", "NC"}:
            invalid_rows.append(f"{row_no}행: 요청번호 또는 제조사 오류")
            continue
        request_manufacturer = _request_manufacturer(current_request)
        if request_manufacturer and request_manufacturer != current_manufacturer:
            invalid_rows.append(
                f"{row_no}행: 요청번호 제조사({request_manufacturer})와 "
                f"제조사({current_manufacturer}) 불일치"
            )
            continue
        if _is_non_part_name(part_name):
            invalid_rows.append(f"{row_no}행: 문서 제목 또는 합계 행은 부품으로 승인할 수 없습니다.")
            continue
        if not _valid_model(model):
            invalid_rows.append(f"{row_no}행: 적용모델 오류({model or '없음'})")
            continue
        if qty is None or qty <= 0:
            invalid_rows.append(f"{row_no}행: 수량 오류")
            continue
        if price is None or price <= 0:
            invalid_rows.append(f"{row_no}행: 단가 오류")
            continue
        if amount is None or amount <= 0:
            invalid_rows.append(f"{row_no}행: 금액 오류")
            continue
        tolerance = max(0.1, abs(amount) * 0.015)
        if abs(qty * price - amount) > tolerance:
            invalid_rows.append(f"{row_no}행: 수량×단가와 금액 불일치")
            continue
        if validation not in {"정상", "금액계산"}:
            invalid_rows.append(f"{row_no}행: 행검증 상태 오류({validation or '없음'})")
            continue

        if price is not None and amount is None:
            amount = round(qty * price, 2)

        request_no = request_no or current_request
        manufacturer = manufacturer or current_manufacturer
        source_pdf = source_pdf or current_pdf

        if current_request != request_no:
            invalid_rows.append(f"{row_no}행: 요청번호가 다른 행과 다릅니다.")
            continue
        if current_manufacturer != manufacturer:
            invalid_rows.append(f"{row_no}행: 제조사가 다른 행과 다릅니다.")
            continue

        records.append(PartRecord(
            base_year=infer_year_from_ir(current_request),
            base_date=datetime.fromtimestamp(path.stat().st_mtime).date().isoformat(),
            ir_no=current_request,
            model=model,
            part_name_en=part_name,
            price=price,
            qty=qty,
            amount=amount,
            source_file=current_pdf or path.name,
            source_row=row_no,
            extraction_method="PDF_REVIEW_APPROVED",
            currency="USD",
            updated=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ))

    if invalid_rows:
        raise ValueError("; ".join(invalid_rows))
    if waiting_rows:
        raise PermissionError(
            f"승인되지 않은 데이터 행이 있습니다: {', '.join(map(str, waiting_rows[:20]))}"
            + (" ..." if len(waiting_rows) > 20 else "")
        )
    if not records:
        raise ValueError("승인된 부품 행이 없습니다.")

    return records, {
        "request_no": request_no,
        "manufacturer": manufacturer,
        "source_pdf": source_pdf,
    }


def _replace_source_records(master_records: list[PartRecord], incoming: list[PartRecord]) -> tuple[list[PartRecord], int]:
    sources = {record.source_file for record in incoming if record.source_file}
    filtered = [record for record in master_records if record.source_file not in sources]
    return filtered, len(master_records) - len(filtered)


def _move_with_timestamp(source: Path, destination_dir: Path) -> Path:
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / source.name
    if destination.exists():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        destination = destination_dir / f"{source.stem}_{stamp}{source.suffix}"
    shutil.move(str(source), str(destination))
    return destination


def _append_history(path: Path, result: ApprovalFileResult) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "approved_at", "review_file", "request_no", "manufacturer", "row_count",
        "new_parts", "updated_rows", "price_changes", "status",
        "master_path", "archived_review_path", "archived_pdf_path", "errors",
    ]
    exists = path.exists()
    with path.open("a", newline="", encoding="utf-8-sig") as fp:
        writer = csv.DictWriter(fp, fieldnames=fields)
        if not exists:
            writer.writeheader()
        writer.writerow({
            "approved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "review_file": result.review_file.name,
            "request_no": result.request_no,
            "manufacturer": result.manufacturer,
            "row_count": result.row_count,
            "new_parts": result.new_parts,
            "updated_rows": result.updated_rows,
            "price_changes": result.price_changes,
            "status": result.status,
            "master_path": str(result.master_path or ""),
            "archived_review_path": str(result.archived_review_path or ""),
            "archived_pdf_path": str(result.archived_pdf_path or ""),
            "errors": " | ".join(result.errors),
        })


def process_approved_pdf_reviews(root: str | Path = ".") -> ApprovalRunResult:
    root = Path(root)
    config = _load_config(root)
    review_dir = root / "PDF_Review"
    update_dir = root / config.get("update", "Update")
    archive_dir = root / config.get("archive", "Archive")
    backup_dir = root / config.get("backup", "Backup")
    log_dir = root / config.get("log", "Log")

    review_dir.mkdir(parents=True, exist_ok=True)
    review_files = sorted(review_dir.glob("PDF_검토대기_*.xlsx"))
    result = ApprovalRunResult(scanned_files=len(review_files))
    history_path = log_dir / "pdf_approval_history.csv"
    result.history_path = history_path

    for review_file in review_files:
        file_result = ApprovalFileResult(review_file=review_file)
        try:
            incoming, meta = _review_rows(review_file)
            manufacturer = meta["manufacturer"]
            request_no = meta["request_no"]
            source_pdf = meta["source_pdf"]

            file_result.manufacturer = manufacturer
            file_result.request_no = request_no
            file_result.row_count = len(incoming)

            master = _master_path(root, config, manufacturer)
            file_result.master_path = master
            if not master.exists():
                raise FileNotFoundError(f"Master 파일 없음: {master}")

            backup_master(master, backup_dir / manufacturer)
            master_records = parse_master(master)
            kr_lookup = build_kr_lookup(master_records)
            incoming, kr_completed = complete_korean_names(
                incoming,
                kr_lookup,
                int(config.get("fuzzy_threshold", 95)),
            )
            master_records, replaced_rows = _replace_source_records(master_records, incoming)
            all_records, new_records, price_changes, duplicates, updated_rows = merge_records(
                master_records,
                incoming,
            )
            updated_rows += replaced_rows

            rebuild_all_sheets(
                master,
                records=all_records,
                appended_count=len(new_records),
                selected_count=1,
                manufacturer=manufacturer,
            )
            saved_records = parse_master(master)
            if len(saved_records) < len(all_records):
                raise RuntimeError(
                    f"Master 저장 검증 실패: 저장 {len(saved_records)}행 / 예상 {len(all_records)}행"
                )

            file_result.new_parts = len(new_records)
            file_result.updated_rows = updated_rows
            file_result.price_changes = len(price_changes)
            file_result.status = "승인 반영 완료"

            approved_review_dir = review_dir / "Approved" / manufacturer
            file_result.archived_review_path = _move_with_timestamp(
                review_file,
                approved_review_dir,
            )

            if source_pdf:
                source_path = update_dir / source_pdf
                if source_path.exists():
                    pdf_archive_dir = archive_dir / manufacturer / "PDF"
                    file_result.archived_pdf_path = _move_with_timestamp(
                        source_path,
                        pdf_archive_dir,
                    )

            result.approved_files += 1
            result.total_rows += file_result.row_count
            result.total_new_parts += file_result.new_parts
            result.total_updated_rows += file_result.updated_rows
            result.total_price_changes += file_result.price_changes

        except PermissionError as exc:
            file_result.status = "승인 대기"
            file_result.errors.append(str(exc))
            result.waiting_files += 1
        except Exception as exc:
            file_result.status = "반영 실패"
            file_result.errors.append(str(exc))
            result.failed_files += 1

        _append_history(history_path, file_result)
        result.files.append(file_result)

    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"PDF_승인반영_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    lines = [
        f"검토파일: {result.scanned_files}",
        f"승인반영: {result.approved_files}",
        f"승인대기: {result.waiting_files}",
        f"실패: {result.failed_files}",
        f"반영행: {result.total_rows}",
        f"신규: {result.total_new_parts}",
        f"보정: {result.total_updated_rows}",
        f"가격변동: {result.total_price_changes}",
        f"승인이력: {history_path}",
        "",
    ]
    for item in result.files:
        lines.append(
            f"[{item.status}] {item.review_file.name} / {item.manufacturer} "
            f"/ {item.request_no} / {item.row_count}행"
        )
        for error in item.errors:
            lines.append(f"  - {error}")
    log_path.write_text("\n".join(lines), encoding="utf-8")
    result.log_path = log_path
    return result
