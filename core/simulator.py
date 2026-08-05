from __future__ import annotations

import hashlib
import json
import shutil
import tempfile
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook

from core.safety import build_safety_report
from core.updater import UpdateResult, Updater

ProgressCallback = Callable[[int, str], None]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fp:
        for chunk in iter(lambda: fp.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _snapshot_master(path: Path) -> dict:
    if not path.exists():
        return {"exists": False, "sha256": "", "size": 0, "sheets": {}}
    sheets: dict[str, int] = {}
    error = ""
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        try:
            sheets = {ws.title: int(ws.max_row) for ws in wb.worksheets}
        finally:
            wb.close()
    except Exception as exc:  # report the issue instead of hiding it
        error = str(exc)
    return {
        "exists": True,
        "sha256": _sha256(path),
        "size": path.stat().st_size,
        "sheets": sheets,
        "error": error,
    }


@dataclass
class SimulationManufacturer:
    manufacturer: str
    status: str
    selected_files: int = 0
    new_parts: int = 0
    updated_rows: int = 0
    price_changes: int = 0
    duplicates: int = 0
    before_rows: int = 0
    after_rows: int = 0
    changed: bool = False
    errors: list[str] = field(default_factory=list)


@dataclass
class SimulationReport:
    created_at: str
    passed: bool
    real_master_unchanged: bool
    update_files: int
    summary: str
    manufacturers: list[SimulationManufacturer]
    safety_summary: str
    safety_score: int
    report_json: str = ""
    report_excel: str = ""
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        payload = asdict(self)
        return payload


class UpdateSimulator:
    """Run a full CPMS update in an isolated temporary workspace.

    The simulator copies only the minimum operational folders, executes the real
    Updater against those copies, validates the simulated result, and finally
    confirms that every real Master.xlsx stayed byte-for-byte unchanged.
    """

    COPY_DIRS = ("Config", "Manufacturers", "Update", "History")

    def __init__(self, root: str | Path = ".") -> None:
        self.root = Path(root).resolve()
        self.real_updater = Updater(self.root)
        self.log_dir = self.root / self.real_updater.config.get("log", "Log")
        self.log_dir.mkdir(parents=True, exist_ok=True)

    def _emit(self, callback: ProgressCallback | None, value: int, message: str) -> None:
        if callback:
            callback(value, message)

    def _copy_workspace(self, target: Path) -> None:
        for name in self.COPY_DIRS:
            source = self.root / name
            destination = target / name
            if source.exists():
                shutil.copytree(source, destination, dirs_exist_ok=True)
            else:
                destination.mkdir(parents=True, exist_ok=True)
        for name in ("Archive", "Backup", "Log", "PDF_Review", "Excluded"):
            (target / name).mkdir(parents=True, exist_ok=True)

    def run(self, progress: ProgressCallback | None = None) -> SimulationReport:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        created_at = datetime.now().isoformat(timespec="seconds")
        self._emit(progress, 3, "시뮬레이션 준비: 실제 Master 기준값 저장")

        real_before = {
            m: _snapshot_master(self.real_updater.master_path(m))
            for m in self.real_updater.manufacturers
        }
        update_files = len(self.real_updater.scan())

        manufacturers: list[SimulationManufacturer] = []
        notes: list[str] = []
        simulation_result: UpdateResult | None = None
        safety_summary = "미실행"
        safety_score = 0
        simulation_passed = False

        with tempfile.TemporaryDirectory(prefix="cpms_simulation_") as temp_name:
            temp_root = Path(temp_name)
            self._emit(progress, 10, "임시 작업공간 생성 및 Master/Update 복제")
            self._copy_workspace(temp_root)
            sim_updater = Updater(temp_root)
            sim_before = {
                m: _snapshot_master(sim_updater.master_path(m))
                for m in sim_updater.manufacturers
            }

            try:
                self._emit(progress, 20, "임시 Master에서 실제 업데이트 로직 실행")
                simulation_result = sim_updater.run(
                    lambda value, message: self._emit(
                        progress,
                        min(85, 20 + int(value * 0.65)),
                        f"[SIM] {message}",
                    )
                )
                self._emit(progress, 88, "시뮬레이션 결과 비교 및 안전 점검")
                safety = build_safety_report(temp_root, sim_updater.manufacturers)
                safety_summary = safety.summary
                safety_score = safety.score

                for m in sim_updater.manufacturers:
                    res = simulation_result.by_manufacturer.get(m)
                    before = sim_before[m]
                    after = _snapshot_master(sim_updater.master_path(m))
                    before_rows = int(before.get("sheets", {}).get("원본상세", 0))
                    after_rows = int(after.get("sheets", {}).get("원본상세", 0))
                    changed = before.get("sha256") != after.get("sha256")
                    manufacturers.append(
                        SimulationManufacturer(
                            manufacturer=m,
                            status=res.status if res else "처리 없음",
                            selected_files=res.selected_files if res else 0,
                            new_parts=res.new_parts if res else 0,
                            updated_rows=res.updated_rows if res else 0,
                            price_changes=res.price_changes if res else 0,
                            duplicates=res.duplicates if res else 0,
                            before_rows=before_rows,
                            after_rows=after_rows,
                            changed=changed,
                            errors=list(res.errors) if res else [],
                        )
                    )

                result_errors = [
                    f"{item.manufacturer}: {error}"
                    for item in manufacturers
                    for error in item.errors
                ]
                if result_errors:
                    notes.extend(result_errors)
                simulation_passed = safety.passed and not result_errors
            except Exception as exc:
                notes.append(f"시뮬레이션 실행 실패: {exc}")
                for m in sim_updater.manufacturers:
                    manufacturers.append(
                        SimulationManufacturer(
                            manufacturer=m,
                            status="시뮬레이션 실패",
                            errors=[str(exc)],
                        )
                    )

        self._emit(progress, 94, "실제 Master 무변경 여부 최종 검증")
        real_after = {
            m: _snapshot_master(self.real_updater.master_path(m))
            for m in self.real_updater.manufacturers
        }
        real_unchanged = all(
            real_before[m].get("sha256") == real_after[m].get("sha256")
            for m in self.real_updater.manufacturers
        )
        if not real_unchanged:
            notes.append("경고: 시뮬레이션 중 실제 Master 해시가 변경되었습니다.")

        passed = simulation_passed and real_unchanged
        if update_files == 0:
            summary = "PASS · 처리할 Update 파일 없음" if passed else "FAIL · 점검 필요"
        else:
            summary = (
                f"PASS · {update_files}개 파일 실제 반영 전 검증 완료"
                if passed
                else f"FAIL · {update_files}개 파일 중 점검 필요"
            )

        report = SimulationReport(
            created_at=created_at,
            passed=passed,
            real_master_unchanged=real_unchanged,
            update_files=update_files,
            summary=summary,
            manufacturers=manufacturers,
            safety_summary=safety_summary,
            safety_score=safety_score,
            notes=notes,
        )
        json_path = self.log_dir / f"Simulation_Report_{timestamp}.json"
        excel_path = self.log_dir / f"Simulation_Report_{timestamp}.xlsx"
        report.report_json = str(json_path)
        report.report_excel = str(excel_path)
        json_path.write_text(json.dumps(report.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
        self._write_excel(report, excel_path)
        latest_path = self.log_dir / "Simulation_Report_Latest.json"
        latest_path.write_text(json.dumps(report.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
        self._emit(progress, 100, report.summary)
        return report

    @staticmethod
    def _write_excel(report: SimulationReport, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "시뮬레이션 요약"
        rows = [
            ("생성일시", report.created_at),
            ("판정", "PASS" if report.passed else "FAIL"),
            ("요약", report.summary),
            ("Update 파일", report.update_files),
            ("실제 Master 무변경", "예" if report.real_master_unchanged else "아니오"),
            ("임시 환경 안전점검", report.safety_summary),
            ("안전점수", report.safety_score),
        ]
        for row in rows:
            ws.append(row)
        ws.append([])
        ws.append([
            "제조사", "상태", "처리파일", "신규", "보정", "가격변동",
            "중복", "원본상세 전", "원본상세 후", "임시 Master 변경", "오류",
        ])
        for item in report.manufacturers:
            ws.append([
                item.manufacturer, item.status, item.selected_files, item.new_parts,
                item.updated_rows, item.price_changes, item.duplicates,
                item.before_rows, item.after_rows, "예" if item.changed else "아니오",
                " | ".join(item.errors),
            ])
        if report.notes:
            note_ws = wb.create_sheet("참고사항")
            for note in report.notes:
                note_ws.append([note])
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 55
        for col in "CDEFGHIJK":
            ws.column_dimensions[col].width = 16
        wb.save(path)
