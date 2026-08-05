from __future__ import annotations

import re
from dataclasses import dataclass
import time
from typing import Callable
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.ocr_engine import (
    OcrHealth, check_ocr_health, classify_document, ocr_document,
    run_tesseract_png,
)


@dataclass
class PdfReviewRow:
    row_no: int
    part_name_en: str
    qty: float | None
    price: float | None
    amount: float | None
    model: str
    confidence: float
    validation: str


@dataclass
class PdfReviewResult:
    source_pdf: Path
    request_no: str
    manufacturer: str
    rows: list[PdfReviewRow]
    total_qty: float | None
    total_amount: float | None
    calculated_qty: float
    calculated_amount: float
    output_xlsx: Path
    note: str = ""
    classification: str = ""
    ocr_status: str = ""
    page_count: int = 0
    elapsed_seconds: float = 0.0


def _clean(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _number(value) -> float | None:
    text = _clean(value).replace(" ", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", ".", "-", "-."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _infer_model(text: str) -> str:
    explicit = re.search(r"(?:CMC|CMS|HM)-([A-Z]?\d{2,4}[A-Z]?)(?:\(G\))?", text or "", re.I)
    if explicit:
        return explicit.group(1).upper()
    for token in reversed([x.strip() for x in (text or "").split(",") if x.strip()]):
        match = re.fullmatch(r"([A-Z]{0,3}\d{2,4}[A-Z]?)(?:\(G\))?", token, re.I)
        if match:
            return match.group(1).upper()
    return ""


def _request_no(path: Path, text: str) -> tuple[str, str]:
    combined = f"{path.stem}\n{text}"
    patterns = [
        r"\b(IR|XC|AC|KA|NC)\s*[-_]?\s*(\d{2,6})(?:\s*/?\s*SP)?",
        r"\bORDER\s+(IR|XC|AC|KA|NC)\s*(\d{2,6})",
        r"\bP/?I\s*NO\.?\s*[:：]?\s*(IR|XC|AC|KA|NC)\s*(\d{2,6})",
    ]
    for pattern in patterns:
        m = re.search(pattern, combined, re.I)
        if m:
            prefix = m.group(1).upper()
            return f"{prefix}{m.group(2)}", prefix

    order = re.search(r"ORDER\s*NO\.?\s*[:：]?\s*(\d{2,6})", text, re.I)
    upper = path.stem.upper()
    manufacturer = next((x for x in ("IR","XC","AC","KA","NC") if x in upper), "UNKNOWN")
    if order and manufacturer != "UNKNOWN":
        return f"{manufacturer}{order.group(1)}", manufacturer
    return path.stem, manufacturer


def _validate(qty, price, amount):
    if qty is not None and price is not None and amount is not None:
        tolerance = max(0.1, abs(amount) * 0.015)
        return amount, "정상" if abs(qty * price - amount) <= tolerance else "금액불일치"
    if qty is not None and price is not None:
        return round(qty * price, 2), "금액계산"
    return amount, "검토필요"


def _is_total_row(values: list[str]) -> bool:
    return any(_clean(v).upper().startswith("TOTAL") for v in values)


def _table_rows_from_native_pdf(document):
    rows = []
    total_qty = total_amount = None

    for page in document:
        tables = page.find_tables().tables
        for table in tables:
            data = table.extract()
            if not data:
                continue

            headers = [_clean(x).upper() for x in (data[0] or [])]
            joined = " | ".join(headers)

            # XC PI layout
            if "DESCRIPTION" in joined and "UNIT PRICE" in joined and "AMOUNT" in joined:
                for raw in data[1:]:
                    vals = [_clean(x) for x in raw]
                    if _is_total_row(vals):
                        nums = [_number(x) for x in vals]
                        nums = [x for x in nums if x is not None]
                        if nums:
                            total_qty = nums[-2] if len(nums) >= 2 else total_qty
                            total_amount = nums[-1]
                        continue
                    if len(vals) >= 6:
                        part = vals[2]
                        qty = _number(vals[3])
                        price = _number(vals[4])
                        amount = _number(vals[5])
                        if part and qty is not None:
                            amount, validation = _validate(qty, price, amount)
                            rows.append(PdfReviewRow(len(rows)+1, part, qty, price, amount, _infer_model(part), 100.0, validation))
                continue

            # KA PI layout
            if "NAME" in joined and "UNIT PRICE" in joined and "QUANTITY" in joined:
                start = 1 if "NO." in joined else 0
                for raw in data[start:]:
                    vals = [_clean(x) for x in raw]
                    if _is_total_row(vals):
                        nums = [_number(x) for x in vals]
                        nums = [x for x in nums if x is not None]
                        if nums:
                            total_qty = nums[-2] if len(nums) >= 2 else total_qty
                            total_amount = nums[-1]
                        continue
                    if len(vals) >= 5:
                        part = re.sub(r"^KA\d{2,6}-\d+\s+FOC\s+spare\s+parts\s*", "", vals[1], flags=re.I).strip()
                        price = _number(vals[2])
                        qty = _number(vals[3])
                        amount = 0.0 if "F.O.C" in vals[4].upper() else _number(vals[4])
                        if part and qty is not None:
                            amount, validation = _validate(qty, price, amount)
                            rows.append(PdfReviewRow(len(rows)+1, part, qty, price, amount, _infer_model(part), 100.0, validation))
                continue

            # KA price-check list
            if "PART NAME" in joined and "UNIT PRICE" in joined:
                for raw in data[1:]:
                    vals = [_clean(x) for x in raw]
                    if len(vals) >= 5:
                        part = vals[0]
                        price = _number(vals[3])
                        qty = _number(vals[4])
                        if part and qty is not None:
                            amount, validation = _validate(qty, price, None)
                            rows.append(PdfReviewRow(len(rows)+1, part, qty, price, amount, _infer_model(part), 100.0, validation))
                continue

            # IR PI layout
            if "DESCRIPTION OF GOODS" in joined and "QTY." in joined and "PRICE" in joined:
                for raw in data[2:]:
                    vals = [_clean(x) for x in raw]
                    if _is_total_row(vals):
                        nums = [_number(x) for x in vals]
                        nums = [x for x in nums if x is not None]
                        if nums:
                            total_qty = nums[-2] if len(nums) >= 2 else total_qty
                            total_amount = nums[-1]
                        continue
                    if len(vals) >= 7:
                        model = _clean(vals[0]).replace("CMC-","").replace("CMS-","")
                        part = vals[2]
                        qty = _number(vals[3])
                        price = _number(vals[5])
                        amount = _number(vals[6])
                        if part and qty is not None:
                            amount, validation = _validate(qty, price, amount)
                            rows.append(PdfReviewRow(len(rows)+1, part, qty, price, amount, model or _infer_model(part), 100.0, validation))
                continue

            # NC PI layout without a proper header split
            for raw in data:
                vals = [_clean(x) for x in raw]
                if _is_total_row(vals):
                    nums = [_number(x) for x in vals]
                    nums = [x for x in nums if x is not None]
                    if nums:
                        total_qty = nums[-2] if len(nums) >= 2 else total_qty
                        total_amount = nums[-1]
                    continue
                if len(vals) >= 6 and vals[0].isdigit():
                    part = vals[2]
                    qty = _number(vals[3])
                    price = _number(vals[4])
                    amount = _number(vals[5])
                    if part and qty is not None:
                        amount, validation = _validate(qty, price, amount)
                        rows.append(PdfReviewRow(len(rows)+1, part, qty, price, amount, _infer_model(part), 100.0, validation))

    return rows, total_qty, total_amount



_NC_NON_PART_PATTERNS = re.compile(
    r"\b(?:ORDER\s*NO|DATE|SUPPLIER|ORDER\s+CONFIRMATION|MODEL|Q['’]?\s*TY|"
    r"PRICE|AMOUNT|REMARKS|ETD|BOKJUNG\s+SCALE)\b",
    re.I,
)


def _normalize_nc_ocr_line(line: str) -> str:
    text = _clean(line)
    text = text.replace("U$", " ").replace("US", " ").replace("U S", " ")
    text = text.replace("|", " ").replace("—", "-").replace("–", "-")
    text = re.sub(r"\bPCS\b", " PCS ", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _nc_numeric_tokens(text: str) -> list[float]:
    values: list[float] = []
    for token in re.findall(r"(?<![A-Za-z])[-+]?\d[\d,]*(?:\.\d+)?", text):
        value = _number(token)
        if value is not None:
            values.append(value)
    return values


def _parse_nc_order_confirmation_text(text: str):
    """Parse legacy NC ORDER CONFIRMATION OCR output.

    The legacy form has fixed business columns:
    MODEL | Q'TY | PRICE | AMOUNT | REMARKS.
    Only rows satisfying qty × price ≈ amount are accepted. Header/date/order
    text is explicitly rejected so it can never become a part name.
    """
    rows: list[PdfReviewRow] = []
    total_qty = total_amount = None
    pending_prefix = ""

    raw_lines = [_normalize_nc_ocr_line(line) for line in text.splitlines()]
    lines = [line for line in raw_lines if line]

    in_table = False
    for line in lines:
        upper = line.upper()

        if "MODEL" in upper and ("PRICE" in upper or "AMOUNT" in upper):
            in_table = True
            pending_prefix = ""
            continue

        if "TOTAL" in upper:
            nums = _nc_numeric_tokens(line)
            if len(nums) >= 2:
                total_qty, total_amount = nums[-2], nums[-1]
            elif len(nums) == 1:
                total_amount = nums[-1]
            break

        # Some cropped OCR outputs omit the header. Start only when a plausible
        # row with at least three numeric values appears.
        nums = _nc_numeric_tokens(line)
        if not in_table and len(nums) >= 3 and not _NC_NON_PART_PATTERNS.search(line):
            in_table = True

        if not in_table:
            continue
        if _NC_NON_PART_PATTERNS.search(line):
            continue

        # Normal row: trailing numeric tokens are qty, price and amount.
        # A model number inside the part name may be numeric too, so use the
        # rightmost three values and verify the arithmetic.
        number_matches = list(
            re.finditer(r"(?<![A-Za-z])[-+]?\d[\d,]*(?:\.\d+)?", line)
        )
        if len(number_matches) < 3:
            # Multi-line descriptions are kept and prepended to the next row.
            if re.search(r"[A-Za-z]", line) and len(line) >= 8:
                pending_prefix = f"{pending_prefix} {line}".strip()
            continue

        qty = _number(number_matches[-3].group())
        price = _number(number_matches[-2].group())
        amount = _number(number_matches[-1].group())
        if qty is None or price is None or amount is None:
            continue
        if qty <= 0 or price < 0 or amount < 0:
            continue

        part = line[: number_matches[-3].start()].strip(" ,;:-")
        part = re.sub(r"\s+PCS\s*$", "", part, flags=re.I).strip()
        if pending_prefix:
            part = f"{pending_prefix} {part}".strip()
            pending_prefix = ""

        # Reject obvious document metadata and implausibly short labels.
        if not part or len(part) < 4 or _NC_NON_PART_PATTERNS.search(part):
            continue

        validated_amount, validation = _validate(qty, price, amount)
        if validation != "정상":
            # OCR sometimes drops the decimal separator in the amount. We do
            # not guess; the row is excluded from automatic extraction.
            continue

        # Require a model token at the end of the description. This is a strong
        # guard against header/date lines being accepted as parts.
        model = _infer_model(part)
        if not model:
            model_tokens = re.findall(
                r"(?<![A-Za-z0-9])([A-Z]{0,3}\d{2,4}[A-Z]?)(?![A-Za-z0-9])",
                part,
                flags=re.I,
            )
            model = model_tokens[-1].upper() if model_tokens else ""
        if not model:
            continue

        rows.append(
            PdfReviewRow(
                len(rows) + 1,
                part,
                qty,
                price,
                validated_amount,
                model,
                92.0,
                "정상",
            )
        )

    # Duplicate OCR lines may occur with some PSM modes.
    deduped: list[PdfReviewRow] = []
    seen: set[tuple[str, float | None, float | None, float | None]] = set()
    for row in rows:
        key = (
            re.sub(r"\s+", " ", row.part_name_en).strip().casefold(),
            row.qty,
            row.price,
            row.amount,
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(
            PdfReviewRow(
                len(deduped) + 1,
                row.part_name_en,
                row.qty,
                row.price,
                row.amount,
                row.model,
                row.confidence,
                row.validation,
            )
        )

    return deduped, total_qty, total_amount


def _nc_candidate_score(rows, total_qty, total_amount) -> tuple[int, int, int]:
    valid = sum(row.validation == "정상" for row in rows)
    qty_ok = (
        total_qty is not None
        and abs(sum(row.qty or 0 for row in rows) - total_qty) < 0.01
    )
    amount_ok = (
        total_amount is not None
        and abs(sum(row.amount or 0 for row in rows) - total_amount) < 0.1
    )
    return (int(qty_ok) + int(amount_ok), valid, len(rows))


def _extract_nc_order_confirmation(
    document,
    executable: str | Path,
    language: str,
    timeout_seconds: int,
):
    """Run a dedicated cropped OCR pass for the legacy NC order form."""
    best = ([], None, None, "")
    best_score = (-1, -1, -1)

    for page_no, page in enumerate(document, start=1):
        rect = page.rect
        # Header/title and footer are excluded. The table occupies the central
        # portion in all supplied NC1909/1911/1930 forms.
        clip = type(rect)(
            rect.x0 + rect.width * 0.01,
            rect.y0 + rect.height * 0.18,
            rect.x1 - rect.width * 0.01,
            rect.y0 + rect.height * 0.72,
        )
        pix = page.get_pixmap(matrix=__import__("fitz").Matrix(3.0, 3.0), clip=clip, alpha=False)
        png = pix.tobytes("png")

        for psm in (6, 4, 11):
            text, _elapsed = run_tesseract_png(
                png,
                executable,
                language=language,
                timeout_seconds=timeout_seconds,
                psm=psm,
            )
            rows, total_qty, total_amount = _parse_nc_order_confirmation_text(text)
            score = _nc_candidate_score(rows, total_qty, total_amount)
            if score > best_score:
                best_score = score
                best = (rows, total_qty, total_amount, f"NC 전용 표 OCR (PSM {psm})")

    return best

def _rows_from_ocr(text: str):
    rows = []
    total_qty = total_amount = None
    pattern = re.compile(
        r"^(?:\d+\s+)?(?P<part>.+?)\s+(?P<qty>\d[\d,]*)\s*(?:PCS|Pcs|pcs)?\s*(?:US|U\$|\$)?\s*(?P<price>\d[\d,.]*)\s*(?:US|U\$|\$)?\s*(?P<amount>\d[\d,.]*)\s*$"
    )
    for line in [_clean(x) for x in text.splitlines() if _clean(x)]:
        if "TOTAL" in line.upper():
            nums = re.findall(r"\d[\d,.]*", line)
            if nums:
                values = [_number(x) for x in nums]
                values = [x for x in values if x is not None]
                if values:
                    total_qty = values[-2] if len(values) >= 2 else total_qty
                    total_amount = values[-1]
            continue
        m = pattern.match(line)
        if not m:
            continue
        part = m.group("part")
        qty = _number(m.group("qty"))
        price = _number(m.group("price"))
        amount = _number(m.group("amount"))
        if part and qty is not None:
            amount, validation = _validate(qty, price, amount)
            rows.append(PdfReviewRow(len(rows)+1, part, qty, price, amount, _infer_model(part), 75.0, validation))
    return rows, total_qty, total_amount


def _write_review(
    pdf_path, output_dir, request_no, manufacturer, rows, total_qty, total_amount,
    note, classification="", ocr_status="", page_count=0, elapsed_seconds=0.0,
):
    calculated_qty = sum(x.qty or 0 for x in rows)
    calculated_amount = sum(x.amount or 0 for x in rows)
    output_xlsx = output_dir / f"PDF_검토대기_{request_no}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "PDF_검토대기"
    headers = ["검토상태","요청 No.","제조사","적용모델","부품명(영어)","수량","단가(USD)","금액(USD)","신뢰도","행검증","원본PDF"]
    for c,h in enumerate(headers,1):
        cell=ws.cell(1,c,h)
        cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",fgColor="1F4E79")
        cell.alignment=Alignment(horizontal="center")
    for r_idx,row in enumerate(rows,2):
        vals=["미확인",request_no,manufacturer,row.model,row.part_name_en,row.qty,row.price,row.amount,row.confidence,row.validation,pdf_path.name]
        for c,v in enumerate(vals,1):
            ws.cell(r_idx,c,v)
        ws.cell(r_idx,1).fill=PatternFill("solid",fgColor="FFF2CC")
    ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:K{max(ws.max_row,1)}"
    widths=[12,18,10,14,58,12,14,16,12,16,45]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

    summary=wb.create_sheet("검증요약")
    items=[
        ("원본 PDF",pdf_path.name),("요청번호",request_no),("제조사",manufacturer),
        ("추출 행",len(rows)),("PDF 표기 총수량",total_qty),("추출 합산 수량",calculated_qty),
        ("PDF 표기 총금액",total_amount),("추출 합산 금액",calculated_amount),
        ("수량 합계 일치","정상" if total_qty is not None and abs(total_qty-calculated_qty)<0.01 else "검토필요"),
        ("금액 합계 일치","정상" if total_amount is not None and abs(total_amount-calculated_amount)<0.1 else "검토필요"),
        ("추출 방식",note),("PDF 분류",classification),("OCR 상태",ocr_status),
        ("페이지 수",page_count),("처리시간(초)",round(elapsed_seconds,2)),
        ("처리 원칙","검토 후 승인 기능을 통해서만 Master 반영"),
    ]
    for i,(a,b) in enumerate(items,1):
        summary.cell(i,1,a).font=Font(bold=True)
        summary.cell(i,2,b)
    summary.column_dimensions["A"].width=25
    summary.column_dimensions["B"].width=75
    wb.save(output_xlsx)
    return PdfReviewResult(
        pdf_path, request_no, manufacturer, rows, total_qty, total_amount,
        calculated_qty, calculated_amount, output_xlsx, note,
        classification, ocr_status, page_count, elapsed_seconds,
    )


def create_pdf_review(
    pdf_path: str | Path,
    output_dir: str | Path,
    *,
    tesseract_path: str | Path | None = None,
    ocr_language: str = "eng",
    timeout_seconds: int = 45,
    progress: Callable[[int, int, str], None] | None = None,
    health: OcrHealth | None = None,
) -> PdfReviewResult:
    try:
        import fitz
    except ImportError as exc:
        raise RuntimeError("PDF 모듈이 없습니다: py -m pip install pymupdf") from exc

    started = time.perf_counter()
    pdf_path = Path(pdf_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    document = fitz.open(pdf_path)
    try:
        native_text = "\n".join(
            page.get_text("text", sort=True) or "" for page in document
        )
        classification = classify_document(document, native_text)
        rows, total_qty, total_amount = _table_rows_from_native_pdf(document)
        note = "PyMuPDF 표 좌표 직접 추출"
        ocr_status = "미사용"

        # Text PDFs sometimes have usable text but no detected table.
        if not rows and native_text.strip():
            rows, total_qty, total_amount = _rows_from_ocr(native_text)
            if rows:
                note = "PDF 내장 텍스트 행 추출"

        if not rows:
            health = health or check_ocr_health(tesseract_path, ocr_language)
            if not health.available:
                request_no, manufacturer = _request_no(pdf_path, native_text)
                note = "OCR 미설치 - 수동 입력 필요"
                ocr_status = health.message
                return _write_review(
                    pdf_path, output_dir, request_no, manufacturer, [],
                    total_qty, total_amount, note,
                    classification=classification.mode,
                    ocr_status=ocr_status,
                    page_count=classification.page_count,
                    elapsed_seconds=time.perf_counter() - started,
                )

            try:
                ocr_text, page_results = ocr_document(
                    document,
                    health,
                    language=ocr_language,
                    timeout_seconds=timeout_seconds,
                    progress=progress,
                )
                rows, total_qty, total_amount = _rows_from_ocr(ocr_text)
                native_text = ocr_text
                success_pages = sum(item.status == "성공" for item in page_results)
                ocr_status = f"성공 {success_pages}/{len(page_results)}페이지"
                note = "Tesseract OCR 추출"

                detected_request, detected_manufacturer = _request_no(pdf_path, ocr_text)
                if detected_manufacturer == "NC":
                    nc_rows, nc_total_qty, nc_total_amount, nc_note = _extract_nc_order_confirmation(
                        document,
                        health.executable,
                        ocr_language,
                        timeout_seconds,
                    )
                    if _nc_candidate_score(nc_rows, nc_total_qty, nc_total_amount) > _nc_candidate_score(
                        rows, total_qty, total_amount
                    ):
                        rows, total_qty, total_amount = (
                            nc_rows, nc_total_qty, nc_total_amount
                        )
                        note = nc_note
            except Exception as exc:
                request_no, manufacturer = _request_no(pdf_path, native_text)
                note = "OCR 실패 - 수동 입력 필요"
                ocr_status = str(exc)
                return _write_review(
                    pdf_path, output_dir, request_no, manufacturer, [],
                    total_qty, total_amount, note,
                    classification=classification.mode,
                    ocr_status=ocr_status,
                    page_count=classification.page_count,
                    elapsed_seconds=time.perf_counter() - started,
                )

        request_no, manufacturer = _request_no(pdf_path, native_text)
        if not rows:
            note = "자동 추출 실패 - 수동 입력 필요"

        return _write_review(
            pdf_path, output_dir, request_no, manufacturer, rows,
            total_qty, total_amount, note,
            classification=classification.mode,
            ocr_status=ocr_status,
            page_count=classification.page_count,
            elapsed_seconds=time.perf_counter() - started,
        )
    finally:
        document.close()
