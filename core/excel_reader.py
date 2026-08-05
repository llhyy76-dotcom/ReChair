from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Iterator, Protocol

from openpyxl import load_workbook


class CellLike(Protocol):
    value: object
    number_format: str


class SheetLike(Protocol):
    title: str
    max_row: int
    max_column: int

    def cell(self, row: int, column: int) -> CellLike: ...
    def iter_rows(self, min_row: int = 1, max_row: int | None = None) -> Iterator[tuple[CellLike, ...]]: ...


@dataclass(frozen=True)
class ReaderInfo:
    source_format: str
    hidden_sheets: tuple[str, ...]
    hidden_columns: dict[str, tuple[int, ...]]


class WorkbookReader:
    """Unified read-only workbook wrapper for .xlsx/.xls files.

    The parser intentionally reads hidden sheets and hidden columns as data.
    Visibility is captured only for diagnostics; it never suppresses values.
    """

    def __init__(self, path: str | Path, *, data_only: bool = True):
        self.path = Path(path)
        suffix = self.path.suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            self._impl: _BaseWorkbook = _OpenPyxlWorkbook(self.path, data_only=data_only)
        elif suffix == ".xls":
            self._impl = _XlrdWorkbook(self.path)
        else:
            raise ValueError(f"Unsupported Excel format: {self.path.suffix}")

    @property
    def sheetnames(self) -> list[str]:
        return self._impl.sheetnames

    @property
    def info(self) -> ReaderInfo:
        return self._impl.info

    def __getitem__(self, name: str) -> SheetLike:
        return self._impl[name]

    def close(self) -> None:
        self._impl.close()

    def __enter__(self) -> "WorkbookReader":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()


class _BaseWorkbook:
    @property
    def sheetnames(self) -> list[str]:
        raise NotImplementedError

    @property
    def info(self) -> ReaderInfo:
        raise NotImplementedError

    def __getitem__(self, name: str) -> SheetLike:
        raise NotImplementedError

    def close(self) -> None:
        pass


class _OpenPyxlWorkbook(_BaseWorkbook):
    def __init__(self, path: Path, *, data_only: bool):
        self.book = load_workbook(
            path, data_only=data_only, read_only=False, keep_links=False
        )
        hidden_sheets = tuple(ws.title for ws in self.book.worksheets if ws.sheet_state != "visible")
        hidden_columns: dict[str, tuple[int, ...]] = {}
        for ws in self.book.worksheets:
            cols: list[int] = []
            for key, dim in ws.column_dimensions.items():
                if dim.hidden:
                    try:
                        from openpyxl.utils.cell import column_index_from_string

                        cols.append(column_index_from_string(key))
                    except Exception:
                        continue
            hidden_columns[ws.title] = tuple(sorted(set(cols)))
        self._info = ReaderInfo("xlsx", hidden_sheets, hidden_columns)

    @property
    def sheetnames(self) -> list[str]:
        return list(self.book.sheetnames)

    @property
    def info(self) -> ReaderInfo:
        return self._info

    def __getitem__(self, name: str):
        return self.book[name]

    def close(self) -> None:
        self.book.close()


@dataclass
class _XlrdCell:
    value: object
    number_format: str = ""


class _XlrdSheet:
    def __init__(self, book, sheet):
        self._book = book
        self._sheet = sheet
        self.title = sheet.name
        self.max_row = sheet.nrows
        self.max_column = sheet.ncols

    def _number_format(self, row0: int, col0: int) -> str:
        try:
            xf_index = self._sheet.cell_xf_index(row0, col0)
            xf = self._book.xf_list[xf_index]
            fmt = self._book.format_map.get(xf.format_key)
            return fmt.format_str if fmt is not None else ""
        except Exception:
            return ""

    def cell(self, row: int, column: int) -> _XlrdCell:
        row0, col0 = row - 1, column - 1
        if row0 < 0 or col0 < 0 or row0 >= self.max_row or col0 >= self.max_column:
            return _XlrdCell(None, "")
        cell = self._sheet.cell(row0, col0)
        value = cell.value
        # xlrd date cells are represented as floats. Convert only when the cell type says date.
        try:
            import xlrd

            if cell.ctype == xlrd.XL_CELL_DATE:
                value = xlrd.xldate.xldate_as_datetime(value, self._book.datemode)
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                value = bool(value)
            elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                value = None
        except Exception:
            pass
        return _XlrdCell(value, self._number_format(row0, col0))

    def iter_rows(self, min_row: int = 1, max_row: int | None = None):
        end = min(max_row or self.max_row, self.max_row)
        for row in range(max(1, min_row), end + 1):
            yield tuple(self.cell(row, col) for col in range(1, self.max_column + 1))


class _XlrdWorkbook(_BaseWorkbook):
    def __init__(self, path: Path):
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError(".xls support requires xlrd==2.0.1") from exc

        # formatting_info allows currency formats and hidden-column diagnostics.
        self.book = xlrd.open_workbook(path, on_demand=False, formatting_info=True)
        self._sheets = {sheet.name: _XlrdSheet(self.book, sheet) for sheet in self.book.sheets()}
        hidden_sheets: list[str] = []
        hidden_columns: dict[str, tuple[int, ...]] = {}
        for sheet in self.book.sheets():
            try:
                if getattr(sheet, "visibility", 0) != 0:
                    hidden_sheets.append(sheet.name)
            except Exception:
                pass
            cols = [idx + 1 for idx, info in getattr(sheet, "colinfo_map", {}).items() if getattr(info, "hidden", 0)]
            hidden_columns[sheet.name] = tuple(sorted(cols))
        self._info = ReaderInfo("xls", tuple(hidden_sheets), hidden_columns)

    @property
    def sheetnames(self) -> list[str]:
        return list(self._sheets.keys())

    @property
    def info(self) -> ReaderInfo:
        return self._info

    def __getitem__(self, name: str) -> _XlrdSheet:
        return self._sheets[name]

    def close(self) -> None:
        try:
            self.book.release_resources()
        except Exception:
            pass


def open_workbook(path: str | Path, *, data_only: bool = True) -> WorkbookReader:
    return WorkbookReader(path, data_only=data_only)
