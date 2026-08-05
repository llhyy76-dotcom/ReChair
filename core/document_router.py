from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

EXCEL_SUFFIXES = {'.xlsx', '.xlsm', '.xltx', '.xltm'}
SYSTEM_NAMES = {'.keep', 'thumbs.db', 'desktop.ini', '.ds_store'}
MANUFACTURERS = {'IR', 'XC', 'AC', 'KA', 'NC'}


@dataclass(frozen=True)
class RouteItem:
    path: str
    filename: str
    route: str
    manufacturers: list[str]
    reason: str
    sheet_count: int = 0
    large_file: bool = False


@dataclass(frozen=True)
class RouteReport:
    created_at: str
    items: list[RouteItem]
    counts: dict[str, int]
    report_json: str


def _name_route(path: Path) -> tuple[str | None, str]:
    name = path.name.lower()
    if path.suffix.lower() == '.pdf':
        return 'PDF_REVIEW', 'PDF 문서'
    if any(token in name for token in ('필요 부품', '필요부품', '수요', '소요량', '취합')):
        return 'DEMAND_HISTORY', '수요·필요부품 문서명 패턴'
    if any(token in name for token in ('품의서', '종합', '발주 자료', '발주자료', '유상발주 정리')):
        return 'MULTI_DOCUMENT', '복합·품의 문서명 패턴'
    if re.search(r'\b(?:IR|XC|AC|KA|NC)\s*\d{4,}\b', path.stem, re.I):
        return 'STANDARD_UPDATE', '파일명에서 제조사 요청번호 확인'
    return None, ''


def _inspect_excel(path: Path) -> tuple[str, list[str], str, int]:
    manufacturers: set[str] = set()
    sheet_count = 0
    has_price = False
    has_part = False
    has_qty = False
    has_raw = False

    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet_count = len(wb.sheetnames)
        for ws in wb.worksheets:
            sheet_name = str(ws.title).strip().upper()
            if sheet_name in MANUFACTURERS:
                manufacturers.add(sheet_name)
            if any(token in sheet_name for token in ('원시', 'RAW', '부품등록', '사용량', '종합')):
                has_raw = True
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), max_col=min(ws.max_column, 20), values_only=True):
                joined = ' | '.join(str(v or '') for v in row).upper()
                has_part = has_part or any(t in joined for t in ('부품명', 'PART NAME', 'DESCRIPTION', '품명'))
                has_price = has_price or any(t in joined for t in ('단가', 'UNIT PRICE', 'PRICE(USD)', 'PRICE'))
                has_qty = has_qty or any(t in joined for t in ('수량', 'QTY', 'QUANTITY', "Q'TY"))
                for m in MANUFACTURERS:
                    if re.search(rf'\b{m}\s*\d{{4,}}\b', joined):
                        manufacturers.add(m)
    finally:
        wb.close()

    if len(manufacturers) > 1:
        return 'MULTI_DOCUMENT', sorted(manufacturers), '여러 제조사 시트 또는 요청번호 확인', sheet_count
    if len(manufacturers) == 1 and has_part and has_price and sheet_count <= 3 and not has_raw:
        return 'STANDARD_UPDATE', sorted(manufacturers), '단일 제조사·단가 문서 구조', sheet_count
    if has_part and has_qty and not has_price:
        return 'DEMAND_HISTORY', sorted(manufacturers), '부품·수량은 있으나 단가가 없음', sheet_count
    if has_part and (has_price or has_qty):
        return 'MULTI_DOCUMENT', sorted(manufacturers), '복합 시트 또는 통합 발주 문서 구조', sheet_count
    return 'REVIEW_REQUIRED', sorted(manufacturers), '자동 분류 근거 부족', sheet_count


def route_folder(update_dir: Path, log_dir: Path | None = None) -> RouteReport:
    items: list[RouteItem] = []
    for path in sorted(update_dir.iterdir(), key=lambda p: p.name.lower()):
        if not path.is_file() or path.name.lower() in SYSTEM_NAMES or path.name.startswith('~$'):
            continue
        large = path.stat().st_size >= 50 * 1024 * 1024
        route, reason = _name_route(path)
        manufacturers: list[str] = []
        sheet_count = 0
        if route is None and path.suffix.lower() in EXCEL_SUFFIXES:
            try:
                route, manufacturers, reason, sheet_count = _inspect_excel(path)
            except Exception as exc:
                route, reason = 'REVIEW_REQUIRED', f'엑셀 구조 확인 실패: {exc}'
        elif route is None:
            route, reason = 'UNSUPPORTED', f'지원하지 않는 확장자: {path.suffix.lower() or "없음"}'
        items.append(RouteItem(str(path), path.name, route, manufacturers, reason, sheet_count, large))

    counts: dict[str, int] = {}
    for item in items:
        counts[item.route] = counts.get(item.route, 0) + 1
    created = datetime.now().isoformat(timespec='seconds')
    report_path = ''
    if log_dir is not None:
        log_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        target = log_dir / f'Document_Route_{stamp}.json'
        payload = {'created_at': created, 'counts': counts, 'items': [asdict(i) for i in items]}
        target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
        report_path = str(target)
    return RouteReport(created, items, counts, report_path)
