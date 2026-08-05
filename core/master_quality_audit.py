from __future__ import annotations

import json
import re
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string


VALID_MANUFACTURERS = ("IR", "XC", "AC", "KA", "NC")
HEADER_FILL = "1F4E79"
WHITE = "FFFFFF"
LIGHT_GREEN = "E2F0D9"
LIGHT_YELLOW = "FFF2CC"
LIGHT_RED = "FCE4D6"
GRID = "D9E2F3"


@dataclass
class MasterRow:
    manufacturer: str
    master_path: Path
    sheet_row: int
    base_year: str = ""
    base_date: str = ""
    request_no: str = ""
    model: str = ""
    part_name_kr: str = ""
    part_name_en: str = ""
    price: float | None = None
    qty: float | None = None
    amount: float | None = None
    source_file: str = ""
    source_row: str = ""


@dataclass
class Issue:
    manufacturer: str
    severity: str
    issue_code: str
    issue_name: str
    model: str = ""
    part_name_kr: str = ""
    part_name_en: str = ""
    request_no: str = ""
    current_value: str = ""
    related_values: str = ""
    source_file: str = ""
    master_row: int = 0
    recommendation: str = ""


@dataclass
class AuditResult:
    master_files: int = 0
    total_rows: int = 0
    issues: list[Issue] = field(default_factory=list)
    report_path: Path | None = None


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _norm(value: Any) -> str:
    return re.sub(r"[^A-Z0-9가-힣]", "", _clean(value).upper())


def _num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[^0-9.\-]", "", str(value))
    if text in {"", ".", "-", "-."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _load_config(root: Path) -> dict:
    path = root / "Config" / "config.json"
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as fp:
            return json.load(fp)
    except Exception:
        return {}


def _find_original_sheet(wb):
    preferred = ["원본상세", "Original_Detail", "원본 상세"]
    for name in preferred:
        if name in wb.sheetnames:
            return wb[name]
    for ws in wb.worksheets:
        header_text = " ".join(_clean(ws.cell(1, c).value) for c in range(1, min(ws.max_column, 30) + 1))
        if "원본" in header_text and ("부품" in header_text or "PART" in header_text.upper()):
            return ws
    return None


def _header_map(ws) -> tuple[int, dict[str, int]]:
    aliases = {
        "base_year": ["기준년도", "BASE YEAR", "YEAR"],
        "base_date": ["기준일자", "BASE DATE", "DATE"],
        "request_no": ["요청 NO.", "요청NO", "IR NO.", "REQUEST NO", "REQUEST_NO"],
        "model": ["적용모델", "MODEL", "MODEL NO."],
        "part_name_kr": ["부품명(한글)", "PART NAME(KR)", "한글명", "이름"],
        "part_name_en": ["부품명(영어)", "PART NAME(EN)", "PART NAME", "DESCRIPTION", "SPARE PARTS LIST"],
        "price": ["단가(USD)", "PRICE", "UNIT PRICE"],
        "qty": ["수량", "QTY", "QTY(PCS)", "REQUEST QTY"],
        "amount": ["금액(USD)", "AMOUNT", "TOTAL"],
        "source_file": ["원본파일", "SOURCE FILE", "SOURCE_FILE"],
        "source_row": ["원본행", "SOURCE ROW", "SOURCE_ROW"],
    }

    for row in range(1, min(ws.max_row, 15) + 1):
        values = {_norm(ws.cell(row, c).value): c for c in range(1, ws.max_column + 1)}
        mapping: dict[str, int] = {}
        for key, names in aliases.items():
            for name in names:
                normalized = _norm(name)
                if normalized in values:
                    mapping[key] = values[normalized]
                    break
        if "part_name_en" in mapping and ("request_no" in mapping or "model" in mapping):
            return row, mapping
    raise ValueError("원본상세 헤더를 찾지 못했습니다.")


def _xlsx_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    """Load shared strings only when the workbook uses them."""
    try:
        source = archive.open("xl/sharedStrings.xml")
    except KeyError:
        return []

    strings: list[str] = []
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    for event, element in ET.iterparse(source, events=("end",)):
        if element.tag == f"{namespace}si":
            parts = [node.text or "" for node in element.iter(f"{namespace}t")]
            strings.append("".join(parts))
            element.clear()
    return strings


def _xlsx_original_sheet_xml(archive: zipfile.ZipFile) -> str:
    """Resolve the XML path for 원본상세 without opening all workbook sheets."""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    pkg_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    relation_id = ""
    preferred = {"원본상세", "Original_Detail", "원본 상세"}

    sheets = workbook.find(f"{{{main_ns}}}sheets")
    if sheets is None:
        raise ValueError("workbook.xml에 시트 목록이 없습니다.")

    fallback_ids: list[tuple[str, str]] = []
    for sheet in sheets:
        name = sheet.attrib.get("name", "")
        rid = sheet.attrib.get(f"{{{rel_ns}}}id", "")
        fallback_ids.append((name, rid))
        if name in preferred:
            relation_id = rid
            break

    if not relation_id:
        for name, rid in fallback_ids:
            normalized = re.sub(r"\s+", "", name).lower()
            if "원본" in normalized or "original" in normalized:
                relation_id = rid
                break

    if not relation_id:
        raise ValueError("원본상세 시트를 찾지 못했습니다.")

    relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    for relation in relationships:
        if relation.attrib.get("Id") == relation_id:
            target = relation.attrib.get("Target", "")
            if target.startswith("/"):
                return target.lstrip("/")
            target = target.replace("\\", "/")
            if target.startswith("xl/"):
                return target
            return f"xl/{target.lstrip('./')}"

    raise ValueError("원본상세 시트 관계를 찾지 못했습니다.")


def _xlsx_cell_value(cell, shared_strings: list[str]) -> object:
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    cell_type = cell.attrib.get("t", "")
    value_node = cell.find(f"{namespace}v")

    if cell_type == "inlineStr":
        parts = [node.text or "" for node in cell.iter(f"{namespace}t")]
        return "".join(parts)

    if value_node is None:
        formula = cell.find(f"{namespace}f")
        return "" if formula is None else ""

    raw = value_node.text or ""
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except (ValueError, IndexError):
            return raw
    if cell_type == "b":
        return raw == "1"
    if cell_type in {"str", "e"}:
        return raw

    try:
        number = float(raw)
        return int(number) if number.is_integer() else number
    except ValueError:
        return raw


def _read_master_fast(
    manufacturer: str,
    path: Path,
    progress: Callable[[str], None] | None = None,
) -> list[MasterRow]:
    """Read only the original-detail worksheet directly from XLSX XML."""
    def emit(message: str) -> None:
        if progress:
            progress(message)

    rows: list[MasterRow] = []
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

    with zipfile.ZipFile(path) as archive:
        emit(f"{manufacturer}: XLSX 내부 구조 확인...")
        shared_strings = _xlsx_shared_strings(archive)
        sheet_path = _xlsx_original_sheet_xml(archive)
        emit(f"{manufacturer}: 원본상세 XML 스트리밍 시작 ({sheet_path})")

        header_row_number = 0
        mapping: dict[str, int] = {}
        processed_data_rows = 0

        aliases = {
            "base_year": ["기준년도", "BASE YEAR", "YEAR"],
            "base_date": ["기준일자", "BASE DATE", "DATE"],
            "request_no": ["요청 NO.", "요청NO", "IR NO.", "REQUEST NO", "REQUEST_NO"],
            "model": ["적용모델", "MODEL", "MODEL NO."],
            "part_name_kr": ["부품명(한글)", "PART NAME(KR)", "한글명", "이름"],
            "part_name_en": ["부품명(영어)", "PART NAME(EN)", "PART NAME", "DESCRIPTION", "SPARE PARTS LIST"],
            "price": ["단가(USD)", "PRICE", "UNIT PRICE"],
            "qty": ["수량", "QTY", "QTY(PCS)", "REQUEST QTY"],
            "amount": ["금액(USD)", "AMOUNT", "TOTAL"],
            "source_file": ["원본파일", "SOURCE FILE", "SOURCE_FILE"],
            "source_row": ["원본행", "SOURCE ROW", "SOURCE_ROW"],
        }

        with archive.open(sheet_path) as source:
            for event, element in ET.iterparse(source, events=("end",)):
                if element.tag != f"{namespace}row":
                    continue

                row_number = int(element.attrib.get("r", "0") or 0)
                values: dict[int, object] = {}

                for cell in element.findall(f"{namespace}c"):
                    reference = cell.attrib.get("r", "")
                    match = re.match(r"([A-Z]+)", reference)
                    if not match:
                        continue
                    column = column_index_from_string(match.group(1))
                    values[column] = _xlsx_cell_value(cell, shared_strings)

                if not mapping and row_number <= 15:
                    normalized = {_norm(value): column for column, value in values.items()}
                    candidate: dict[str, int] = {}
                    for key, names in aliases.items():
                        for name in names:
                            name_norm = _norm(name)
                            if name_norm in normalized:
                                candidate[key] = normalized[name_norm]
                                break
                    if "part_name_en" in candidate and (
                        "request_no" in candidate or "model" in candidate
                    ):
                        mapping = candidate
                        header_row_number = row_number
                        emit(f"{manufacturer}: 헤더 발견 — Excel {header_row_number}행")
                    element.clear()
                    continue

                if not mapping or row_number <= header_row_number:
                    element.clear()
                    continue

                def val(key: str):
                    column = mapping.get(key)
                    return values.get(column) if column else None

                en = _clean(val("part_name_en"))
                kr = _clean(val("part_name_kr"))
                request = _clean(val("request_no"))
                if en or kr or request:
                    rows.append(MasterRow(
                        manufacturer=manufacturer,
                        master_path=path,
                        sheet_row=row_number,
                        base_year=_clean(val("base_year")),
                        base_date=_clean(val("base_date")),
                        request_no=request,
                        model=_clean(val("model")),
                        part_name_kr=kr,
                        part_name_en=en,
                        price=_num(val("price")),
                        qty=_num(val("qty")),
                        amount=_num(val("amount")),
                        source_file=_clean(val("source_file")),
                        source_row=_clean(val("source_row")),
                    ))

                processed_data_rows += 1
                if processed_data_rows % 10000 == 0:
                    emit(
                        f"{manufacturer}: 원본상세 {processed_data_rows:,}행 확인 "
                        f"/ 유효 데이터 {len(rows):,}행"
                    )

                element.clear()

        if not mapping:
            raise ValueError("원본상세 헤더를 찾지 못했습니다.")

    emit(
        f"{manufacturer}: XML 스트리밍 완료 — 확인 {processed_data_rows:,}행 "
        f"/ 유효 데이터 {len(rows):,}행"
    )
    return rows


def _read_master(
    manufacturer: str,
    path: Path,
    progress: Callable[[str], None] | None = None,
) -> list[MasterRow]:
    """Use the fast XML reader, with openpyxl as a compatibility fallback."""
    try:
        return _read_master_fast(manufacturer, path, progress)
    except (zipfile.BadZipFile, KeyError, ET.ParseError, ValueError) as fast_error:
        if progress:
            progress(
                f"{manufacturer}: 빠른 XML 읽기 실패 — 호환 모드로 재시도 "
                f"({type(fast_error).__name__}: {fast_error})"
            )

        wb = load_workbook(path, data_only=True, read_only=True)
        ws = _find_original_sheet(wb)
        if ws is None:
            raise ValueError("원본상세 시트 없음")
        header_row, h = _header_map(ws)

        rows: list[MasterRow] = []
        checked = 0
        for row_no in range(header_row + 1, ws.max_row + 1):
            def val(key: str):
                col = h.get(key)
                return ws.cell(row_no, col).value if col else None

            en = _clean(val("part_name_en"))
            kr = _clean(val("part_name_kr"))
            request = _clean(val("request_no"))
            if en or kr or request:
                rows.append(MasterRow(
                    manufacturer=manufacturer,
                    master_path=path,
                    sheet_row=row_no,
                    base_year=_clean(val("base_year")),
                    base_date=_clean(val("base_date")),
                    request_no=request,
                    model=_clean(val("model")),
                    part_name_kr=kr,
                    part_name_en=en,
                    price=_num(val("price")),
                    qty=_num(val("qty")),
                    amount=_num(val("amount")),
                    source_file=_clean(val("source_file")),
                    source_row=_clean(val("source_row")),
                ))

            checked += 1
            if progress and checked % 10000 == 0:
                progress(
                    f"{manufacturer}: 호환 모드 {checked:,}행 확인 "
                    f"/ 유효 데이터 {len(rows):,}행"
                )

        wb.close()
        return rows

def _part_key(row: MasterRow) -> tuple[str, str]:
    return (_norm(row.model), _norm(row.part_name_en or row.part_name_kr))


def _audit_rows(rows: list[MasterRow]) -> list[Issue]:
    issues: list[Issue] = []

    for row in rows:
        common = dict(
            manufacturer=row.manufacturer,
            model=row.model,
            part_name_kr=row.part_name_kr,
            part_name_en=row.part_name_en,
            request_no=row.request_no,
            source_file=row.source_file,
            master_row=row.sheet_row,
        )

        if not row.model:
            issues.append(Issue(
                severity="경고", issue_code="MODEL_MISSING", issue_name="적용모델 누락",
                current_value="", recommendation="파일명 또는 부품명의 모델 정보를 확인해 적용모델을 입력하세요.",
                **common,
            ))
        if not row.part_name_en and not row.part_name_kr:
            issues.append(Issue(
                severity="오류", issue_code="PART_NAME_MISSING", issue_name="부품명 누락",
                recommendation="원본 요청서를 확인해 부품명을 입력하세요.", **common,
            ))
        if row.price is not None and row.price < 0:
            issues.append(Issue(
                severity="오류", issue_code="NEGATIVE_PRICE", issue_name="음수 단가",
                current_value=str(row.price), recommendation="단가 부호와 원본 셀 위치를 확인하세요.", **common,
            ))
        if row.qty is not None and row.qty < 0:
            issues.append(Issue(
                severity="오류", issue_code="NEGATIVE_QTY", issue_name="음수 수량",
                current_value=str(row.qty), recommendation="반품/차감 행인지 확인하고 일반 발주 수량과 분리하세요.", **common,
            ))
        if row.price is not None and row.qty is not None and row.amount is not None:
            expected = row.price * row.qty
            tolerance = max(0.1, abs(row.amount) * 0.01)
            if abs(expected - row.amount) > tolerance and row.amount != 0:
                issues.append(Issue(
                    severity="경고", issue_code="AMOUNT_MISMATCH", issue_name="수량×단가와 금액 불일치",
                    current_value=f"{row.qty} × {row.price} = {expected:.2f}",
                    related_values=f"기록금액: {row.amount:.2f}",
                    recommendation="FOC, 할인, 분할금액 여부를 확인하고 잘못된 값이면 수정하세요.",
                    **common,
                ))
        if not row.request_no:
            issues.append(Issue(
                severity="경고", issue_code="REQUEST_NO_MISSING", issue_name="요청번호 누락",
                recommendation="원본파일명 또는 문서 내부 요청번호를 확인하세요.", **common,
            ))
        if not row.base_year:
            issues.append(Issue(
                severity="정보", issue_code="BASE_YEAR_MISSING", issue_name="기준년도 누락",
                recommendation="요청번호나 기준일자에서 연도를 보완하세요.", **common,
            ))

    by_key: dict[tuple[str, str], list[MasterRow]] = defaultdict(list)
    by_source: dict[tuple[str, str, str], list[MasterRow]] = defaultdict(list)
    for row in rows:
        if _part_key(row)[1]:
            by_key[_part_key(row)].append(row)
        if row.source_file and row.source_row:
            by_source[(row.source_file.lower(), row.source_row, row.request_no)].append(row)

    for key, group in by_key.items():
        prices = sorted({round(row.price, 6) for row in group if row.price is not None})
        kr_names = sorted({_clean(row.part_name_kr) for row in group if _clean(row.part_name_kr)})
        models = sorted({_clean(row.model) for row in group if _clean(row.model)})
        latest = group[-1]

        if len(prices) > 1:
            issues.append(Issue(
                manufacturer=latest.manufacturer, severity="검토", issue_code="MULTIPLE_PRICES",
                issue_name="동일 부품 복수 단가", model=latest.model,
                part_name_kr=latest.part_name_kr, part_name_en=latest.part_name_en,
                request_no=latest.request_no, current_value=str(prices[-1]),
                related_values=", ".join(map(str, prices)),
                source_file=latest.source_file, master_row=latest.sheet_row,
                recommendation="정상적인 가격이력인지 확인하고 최신 단가와 요청번호 매칭 색상을 검토하세요.",
            ))

        if len(kr_names) > 1:
            issues.append(Issue(
                manufacturer=latest.manufacturer, severity="검토", issue_code="KR_NAME_CONFLICT",
                issue_name="동일 부품 한글명 불일치", model=latest.model,
                part_name_kr=latest.part_name_kr, part_name_en=latest.part_name_en,
                request_no=latest.request_no, current_value=latest.part_name_kr,
                related_values=" | ".join(kr_names),
                source_file=latest.source_file, master_row=latest.sheet_row,
                recommendation="표준 한글 부품명을 하나로 통일하세요.",
            ))

    for _, group in by_source.items():
        if len(group) > 1:
            first = group[0]
            issues.append(Issue(
                manufacturer=first.manufacturer, severity="오류", issue_code="DUPLICATE_SOURCE_ROW",
                issue_name="동일 원본행 중복", model=first.model,
                part_name_kr=first.part_name_kr, part_name_en=first.part_name_en,
                request_no=first.request_no, current_value=f"{len(group)}건",
                related_values=", ".join(str(row.sheet_row) for row in group),
                source_file=first.source_file, master_row=first.sheet_row,
                recommendation="동일 원본파일·원본행·요청번호 중복을 제거하세요.",
            ))

    return issues


def _style_header(ws, row: int, max_col: int) -> None:
    thin = Side(style="thin", color=GRID)
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(name="맑은 고딕", bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_report(root: Path, rows: list[MasterRow], issues: list[Issue]) -> Path:
    report_dir = root / "Log"
    report_dir.mkdir(parents=True, exist_ok=True)
    output = report_dir / f"Master_Quality_Audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "품질검사_요약"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    ws["A1"] = "CPMS MASTER 품질 자동검사"
    ws["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws["A1"].font = Font(name="맑은 고딕", size=20, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    severity_counts = defaultdict(int)
    manufacturer_counts = defaultdict(int)
    for issue in issues:
        severity_counts[issue.severity] += 1
        manufacturer_counts[issue.manufacturer] += 1

    summary = [
        ("검사 Master", len({row.master_path for row in rows})),
        ("검사 원본행", len(rows)),
        ("전체 이슈", len(issues)),
        ("오류", severity_counts["오류"]),
        ("경고", severity_counts["경고"]),
        ("검토", severity_counts["검토"]),
        ("정보", severity_counts["정보"]),
        ("실행일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for index, (label, value) in enumerate(summary):
        col = 1 + (index % 4) * 2
        row = 4 + (index // 4) * 2
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.cell(row, col, label)
        ws.cell(row, col).fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
        ws.cell(row, col).font = Font(bold=True)
        ws.cell(row, col).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        ws.cell(row + 1, col, value)
        ws.cell(row + 1, col).font = Font(size=15, bold=True)
        ws.cell(row + 1, col).alignment = Alignment(horizontal="center")

    row = 9
    headers = ["제조사", "검사행", "이슈", "오류", "경고", "검토", "정보", "상태"]
    for col, header in enumerate(headers, 1):
        ws.cell(row, col, header)
    _style_header(ws, row, len(headers))
    row += 1

    for manufacturer in VALID_MANUFACTURERS:
        m_rows = [item for item in rows if item.manufacturer == manufacturer]
        m_issues = [item for item in issues if item.manufacturer == manufacturer]
        values = [
            manufacturer, len(m_rows), len(m_issues),
            sum(item.severity == "오류" for item in m_issues),
            sum(item.severity == "경고" for item in m_issues),
            sum(item.severity == "검토" for item in m_issues),
            sum(item.severity == "정보" for item in m_issues),
            "정상" if not m_issues else "검토필요",
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
            ws.cell(row, col).alignment = Alignment(horizontal="center")
        row += 1

    issue_ws = wb.create_sheet("품질이슈_상세")
    issue_headers = [
        "제조사", "심각도", "이슈코드", "이슈명", "적용모델",
        "부품명(한글)", "부품명(영어)", "요청 No.", "현재값",
        "관련값", "원본파일", "Master 행", "권장조치",
    ]
    for col, header in enumerate(issue_headers, 1):
        issue_ws.cell(1, col, header)
    _style_header(issue_ws, 1, len(issue_headers))

    for row_no, issue in enumerate(issues, 2):
        values = [
            issue.manufacturer, issue.severity, issue.issue_code, issue.issue_name,
            issue.model, issue.part_name_kr, issue.part_name_en, issue.request_no,
            issue.current_value, issue.related_values, issue.source_file,
            issue.master_row, issue.recommendation,
        ]
        for col, value in enumerate(values, 1):
            issue_ws.cell(row_no, col, value)
            issue_ws.cell(row_no, col).alignment = Alignment(vertical="center", wrap_text=True)
        fill = {
            "오류": LIGHT_RED, "경고": LIGHT_YELLOW,
            "검토": "DDEBF7", "정보": LIGHT_GREEN,
        }.get(issue.severity)
        if fill:
            issue_ws.cell(row_no, 2).fill = PatternFill("solid", fgColor=fill)

    widths = [10, 10, 25, 28, 16, 30, 48, 20, 25, 55, 55, 12, 65]
    for col, width in enumerate(widths, 1):
        issue_ws.column_dimensions[get_column_letter(col)].width = width
    issue_ws.freeze_panes = "A2"
    issue_ws.auto_filter.ref = f"A1:M{max(issue_ws.max_row, 1)}"
    issue_ws.sheet_view.showGridLines = False

    raw_ws = wb.create_sheet("검사대상_원본행")
    raw_headers = [
        "제조사", "기준년도", "기준일자", "요청 No.", "적용모델",
        "부품명(한글)", "부품명(영어)", "단가", "수량", "금액",
        "원본파일", "원본행", "Master 행",
    ]
    for col, header in enumerate(raw_headers, 1):
        raw_ws.cell(1, col, header)
    _style_header(raw_ws, 1, len(raw_headers))
    for row_no, item in enumerate(rows, 2):
        values = [
            item.manufacturer, item.base_year, item.base_date, item.request_no,
            item.model, item.part_name_kr, item.part_name_en, item.price,
            item.qty, item.amount, item.source_file, item.source_row, item.sheet_row,
        ]
        for col, value in enumerate(values, 1):
            raw_ws.cell(row_no, col, value)
    raw_ws.freeze_panes = "A2"
    raw_ws.auto_filter.ref = f"A1:M{max(raw_ws.max_row, 1)}"

    for col, width in {"A":14, "B":14, "C":14, "D":14, "E":14, "F":14, "G":14, "H":24}.items():
        ws.column_dimensions[col].width = width

    wb.save(output)
    return output


def run_master_quality_audit(
    root: str | Path = ".",
    progress: Callable[[str], None] | None = None,
) -> AuditResult:
    root = Path(root)
    config = _load_config(root)
    manufacturer_root = root / config.get("manufacturer_root", "Manufacturers")

    def emit(message: str) -> None:
        if progress:
            progress(message)

    all_rows: list[MasterRow] = []
    all_issues: list[Issue] = []
    master_files = 0

    emit("품질검사를 시작합니다.")
    emit(f"제조사 Master 경로: {manufacturer_root}")

    for index, manufacturer in enumerate(VALID_MANUFACTURERS, 1):
        master = manufacturer_root / manufacturer / "Master" / "Master.xlsx"
        emit("")
        emit(f"[{index}/{len(VALID_MANUFACTURERS)}] {manufacturer} Master 확인 중...")
        emit(f"파일: {master}")

        if not master.exists():
            emit(f"{manufacturer}: Master 파일 없음 — 건너뜁니다.")
            all_issues.append(Issue(
                manufacturer=manufacturer, severity="오류",
                issue_code="MASTER_FILE_MISSING", issue_name="Master 파일 없음",
                current_value=str(master),
                recommendation="제조사 Master 경로와 파일 생성 상태를 확인하세요.",
            ))
            continue

        master_files += 1
        try:
            emit(f"{manufacturer}: 원본상세 고속 읽기 준비...")
            rows = _read_master(manufacturer, master, progress=emit)
            emit(f"{manufacturer}: 원본상세 {len(rows):,}행 읽기 완료")
            emit(f"{manufacturer}: 품질 규칙 검사 중...")
            issues = _audit_rows(rows)
            all_rows.extend(rows)
            all_issues.extend(issues)
            emit(f"{manufacturer}: 검사 완료 — 이슈 {len(issues):,}건")
        except PermissionError as exc:
            emit(f"{manufacturer}: 접근 실패 — Excel에서 파일이 열려 있을 수 있습니다.")
            emit(f"상세: {exc}")
            all_issues.append(Issue(
                manufacturer=manufacturer, severity="오류",
                issue_code="MASTER_LOCKED", issue_name="Master 파일 잠김",
                current_value=str(exc), source_file=str(master),
                recommendation="해당 Master.xlsx를 닫고 다시 실행하세요.",
            ))
        except Exception as exc:
            emit(f"{manufacturer}: 읽기/검사 실패 — 나머지 제조사는 계속 검사합니다.")
            emit(f"상세: {type(exc).__name__}: {exc}")
            all_issues.append(Issue(
                manufacturer=manufacturer, severity="오류",
                issue_code="MASTER_READ_FAILED", issue_name="Master 읽기 실패",
                current_value=f"{type(exc).__name__}: {exc}", source_file=str(master),
                recommendation="Excel 파일 잠금, 원본상세 시트와 헤더 구조를 확인하세요.",
            ))

    emit("")
    emit("전체 제조사 검사가 끝났습니다.")
    emit(f"검사 원본행: {len(all_rows):,}행")
    emit(f"발견 이슈: {len(all_issues):,}건")
    emit("Excel 품질검사 보고서 생성 중...")

    report = _write_report(root, all_rows, all_issues)
    emit(f"보고서 생성 완료: {report}")

    return AuditResult(
        master_files=master_files,
        total_rows=len(all_rows),
        issues=all_issues,
        report_path=report,
    )

